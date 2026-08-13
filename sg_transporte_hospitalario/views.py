import json
import re
import unicodedata
from datetime import date, datetime, time, timedelta
from decimal import Decimal
import calendar
from collections import defaultdict

from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.db.models import Count, Exists, OuterRef, Q, Prefetch
from django.utils.decorators import method_decorator
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.utils import timezone
from django.urls import reverse
from django.views.decorators.clickjacking import xframe_options_exempt
from django.views.decorators.http import require_GET, require_POST
from django.views.generic import ListView, TemplateView, View
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from core.services.reporte.EXCEL.reporte_service_excel import ServiceExcel
from core.services.usuario_service import UsuarioService
from core.utils.utilidades_fechas import mes_nombre
from core.utils.utilidades_mensajes import mostrar_mensaje
from ingreso.models import Ingreso
from paciente.models import Paciente
from rrhh.models import Empleado

from .forms import EjecucionViajeForm, SolicitudCreateForm, SolicitudForm, ViajeProgramacionForm, _empleados_operativos_disponibles_qs
from .models import EjecucionViaje, Solicitud, SolicitudPaciente, SolicitudPersonal, TipoSolicitud, Viaje, ViajeSolicitud, ViajePersonal, ViajeViatico, Vehiculo, Motorista, Viatico


def puede_ver_modulo(usuario):
    return bool(
        usuario
        and (
            usuario.is_superuser
            or UsuarioService.es_directivo(usuario)
            or UsuarioService.es_admin_global(usuario)
        )
    )


def puede_gestionar_modulo(usuario):
    return bool(usuario and (usuario.is_superuser or UsuarioService.es_admin_global(usuario)))


def construir_contexto_dashboard(usuario):
    return {
        "titulo": "SG-transporte_hospitalario",
        "subtitulo": "Módulo de gestión de transporte hospitalario.",
        "fecha_actual": timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M"),
        "puede_ver": puede_ver_modulo(usuario),
        "puede_escribir": puede_gestionar_modulo(usuario),
    }


TAB_SECTIONS = [
    {"key": "solicitud", "label": "Solicitud", "icon": "bi-journal-text"},
    {"key": "autorizacion", "label": "Autorización", "icon": "bi-check2-circle"},
    {"key": "viaje_construccion", "label": "Viaje en Construcción", "icon": "bi-graph-up-arrow"},
    {"key": "ejecucion", "label": "Ejecución", "icon": "bi-truck"},
    {"key": "resumen", "label": "Resumen", "icon": "bi-card-list"},
]


def _area_solicitante_desde_punto(punto):
    if not punto:
        return "-"
    if punto.unidad_clinica:
        return str(punto.unidad_clinica)
    if punto.unidad:
        return punto.unidad.nombre_unidad
    return "-"


def _punto_solicitud_corto(punto):
    if not punto:
        return "-"
    if punto.unidad_clinica:
        unidad_clinica = punto.unidad_clinica
        if getattr(unidad_clinica, "area_atencion", None):
            area = unidad_clinica.area_atencion
            if getattr(area, "nombre_corto_area_atencion", None):
                return area.nombre_corto_area_atencion
        if getattr(unidad_clinica, "sala", None):
            sala = unidad_clinica.sala
            if getattr(sala, "nombre_corto_sala", None):
                return sala.nombre_corto_sala
        if getattr(unidad_clinica, "servicio_aux", None):
            servicio_aux = unidad_clinica.servicio_aux
            if getattr(servicio_aux, "nombre_corto_servicio_a", None):
                return servicio_aux.nombre_corto_servicio_a
        establecimiento = getattr(unidad_clinica, "establecimiento_ext", None)
        if establecimiento and getattr(establecimiento, "nivel_complejidad_institucional", None):
            siglas = getattr(establecimiento.nivel_complejidad_institucional, "siglas", None)
            if siglas:
                return siglas
    if punto.unidad:
        return punto.unidad.nombre_corto_unidad or punto.unidad.nombre_unidad
    return "-"


def _nombre_paciente_completo(paciente):
    if not paciente:
        return "-"
    return " ".join(
        x for x in [
            paciente.primer_nombre,
            paciente.segundo_nombre,
            paciente.primer_apellido,
            paciente.segundo_apellido,
        ] if x
    ) or "-"


def _inicio_dia_local(fecha_base):
    return timezone.make_aware(datetime.combine(fecha_base, time.min))


def _fin_dia_local(fecha_base):
    return timezone.make_aware(datetime.combine(fecha_base, time.max))


def _limites_periodo_resumen(request):
    hoy = timezone.localdate()
    mes_inicio = (request.GET.get("mes_inicio") or hoy.strftime("%Y-%m")).strip()
    mes_fin = (request.GET.get("mes_fin") or mes_inicio or hoy.strftime("%Y-%m")).strip()

    try:
        inicio_anio, inicio_mes = [int(valor) for valor in mes_inicio.split("-", 1)]
        fin_anio, fin_mes = [int(valor) for valor in mes_fin.split("-", 1)]
        fecha_inicio = date(inicio_anio, inicio_mes, 1)
        fecha_fin_mes = date(fin_anio, fin_mes, 1)
    except (TypeError, ValueError):
        fecha_inicio = hoy.replace(day=1)
        fecha_fin_mes = fecha_inicio
        mes_inicio = hoy.strftime("%Y-%m")
        mes_fin = mes_inicio

    if fecha_fin_mes < fecha_inicio:
        ultimo_dia_inicio = calendar.monthrange(fecha_inicio.year, fecha_inicio.month)[1]
        ultimo_dia_fin = calendar.monthrange(fecha_fin_mes.year, fecha_fin_mes.month)[1]
        return {
            "valido": False,
            "error": "El mes final no puede ser anterior al mes inicio.",
            "mes_inicio": mes_inicio,
            "mes_fin": mes_fin,
            "inicio": _inicio_dia_local(fecha_inicio),
            "fin": _fin_dia_local(date(fecha_fin_mes.year, fecha_fin_mes.month, ultimo_dia_fin)),
            "fin_mes": fecha_fin_mes,
        }

    ultimo_dia_fin = calendar.monthrange(fecha_fin_mes.year, fecha_fin_mes.month)[1]
    return {
        "valido": True,
        "mes_inicio": mes_inicio,
        "mes_fin": mes_fin,
        "inicio": _inicio_dia_local(fecha_inicio),
        "fin": _fin_dia_local(date(fecha_fin_mes.year, fecha_fin_mes.month, ultimo_dia_fin)),
        "fin_mes": fecha_fin_mes,
    }


def _normalizar_nombre_hoja_excel(nombre, usados):
    texto = unicodedata.normalize("NFKD", str(nombre or "")).encode("ascii", "ignore").decode("ascii")
    texto = re.sub(r"[\\\/?\*:\[\]]", "", texto)
    texto = re.sub(r"\s+", " ", texto).strip() or "Hoja"
    texto = texto[:31]

    base = texto
    indice = 2
    while texto.lower() in usados:
        sufijo = f"_{indice}"
        texto = f"{base[:31 - len(sufijo)]}{sufijo}".strip()
        indice += 1
    usados.add(texto.lower())
    return texto


def _formatear_fecha_excel(fecha_dt):
    if not fecha_dt:
        return ""
    return timezone.localtime(fecha_dt).strftime("%d/%m/%Y %H:%M")


def _formatear_hora_excel(fecha_dt):
    if not fecha_dt:
        return ""
    return timezone.localtime(fecha_dt).strftime("%H:%M")


def _calcular_total_horas_excel(fecha_salida, fecha_retorno):
    if not fecha_salida or not fecha_retorno:
        return None
    delta = fecha_retorno - fecha_salida
    if delta.total_seconds() < 0:
        delta += timedelta(days=1)
    if delta.total_seconds() < 0:
        return None
    return round(delta.total_seconds() / 3600.0, 2)


def _nombre_motorista_excel(viaje):
    motorista = getattr(viaje, "motorista", None)
    empleado = getattr(motorista, "empleado", None)
    if empleado:
        return _nombre_empleado(empleado)
    return "SIN MOTORISTA"


def _texto_vehiculo_excel(viaje):
    vehiculo = getattr(viaje, "vehiculo", None)
    if vehiculo:
        return f"{vehiculo.codigo} - {vehiculo.placa}"
    return ""


def _texto_motorista_excel(viaje):
    motorista = getattr(viaje, "motorista", None)
    empleado = getattr(motorista, "empleado", None)
    if empleado:
        return _nombre_empleado(empleado)
    return ""


def _texto_personal_viaje_excel(viaje):
    personal = []
    for item in viaje.viaje_personal.all():
        nombre = _nombre_empleado(item.empleado)
        participacion = item.tipo_participacion or ""
        personal.append(f"{nombre} ({participacion})" if participacion else nombre)
    return "\n".join(personal)


def _texto_paciente_excel(item_paciente):
    paciente = getattr(item_paciente, "paciente", None)
    if not paciente:
        return "", ""
    nombre = _nombre_paciente_completo(paciente)
    identidad = str(getattr(paciente, "dni", "") or "")
    return nombre, identidad


def _viaje_filas_excel(viaje):
    solicitudes = list(viaje.viaje_solicitudes.all())
    ejecucion = getattr(viaje, "ejecucion_viaje", None)
    total_viaticos = sum(
        (Decimal(str(item.monto_aplicado)) for item in viaje.viaje_viaticos.all() if item.monto_aplicado is not None),
        Decimal("0.00"),
    )
    total_km = None
    if ejecucion and ejecucion.kilometraje_salida is not None and ejecucion.kilometraje_retorno is not None:
        total_km = Decimal(str(ejecucion.kilometraje_retorno)) - Decimal(str(ejecucion.kilometraje_salida))
        if total_km < 0:
            total_km = None
    total_horas = _calcular_total_horas_excel(getattr(ejecucion, "fecha_salida", None), getattr(ejecucion, "fecha_retorno", None))

    filas = []
    total_pacientes = 0

    for solicitud_relacion in solicitudes:
        solicitud = solicitud_relacion.solicitud
        pacientes = list(solicitud.solicitud_pacientes.all())
        if not pacientes:
            pacientes = [None]

        for paciente_item in pacientes:
            total_pacientes += 1 if paciente_item is not None else 0
            nombre_paciente, identidad_paciente = ("", "")
            if paciente_item is not None:
                nombre_paciente, identidad_paciente = _texto_paciente_excel(paciente_item)

            filas.append(
                {
                    "numero_orden": viaje.numero_viaje if not filas else "",
                    "fecha_solicitud": _formatear_fecha_excel(solicitud.fecha_solicitud),
                    "fecha_salida": _formatear_fecha_excel(viaje.fecha_programacion) if not filas else "",
                    "prioridad": solicitud.prioridad.nombre if solicitud.prioridad_id else "",
                    "departamento_sala": _punto_solicitud_corto(solicitud.punto_solicitud),
                    "centro_costo": viaje.centro_costo if not filas and viaje.centro_costo is not None else "",
                    "salida": _info_institucion(solicitud.lugar_salida)["nombre"] if solicitud.lugar_salida_id else "",
                    "destino": _info_institucion(solicitud.lugar_destino)["nombre"] if solicitud.lugar_destino_id else "",
                    "acompanantes": _texto_personal_viaje_excel(viaje) if not filas else "",
                    "motorista": _texto_motorista_excel(viaje) if not filas else "",
                    "vehiculo": _texto_vehiculo_excel(viaje) if not filas else "",
                    "nombre_paciente": nombre_paciente,
                    "identidad_paciente": identidad_paciente,
                    "costo_viaticos": total_viaticos if not filas else None,
                    "tipo_viaje": dict(Viaje.TipoProgramacion.choices).get(viaje.tipo_viaje, "") if not filas else "",
                    "hora_salida": _formatear_hora_excel(getattr(ejecucion, "fecha_salida", None)) if not filas else "",
                    "hora_llegada": _formatear_hora_excel(getattr(ejecucion, "fecha_retorno", None)) if not filas else "",
                    "total_horas": total_horas if not filas else None,
                    "km_salida": ejecucion.kilometraje_salida if ejecucion and ejecucion.kilometraje_salida is not None and not filas else None,
                    "km_entrada": ejecucion.kilometraje_retorno if ejecucion and ejecucion.kilometraje_retorno is not None and not filas else None,
                    "total_km": total_km if not filas else None,
                    "pacientes": total_pacientes if not filas else "",
                }
            )

    if not filas:
        filas.append(
            {
                "numero_orden": viaje.numero_viaje,
                "fecha_solicitud": "",
                "fecha_salida": _formatear_fecha_excel(viaje.fecha_programacion),
                "prioridad": "",
                "departamento_sala": "",
                "centro_costo": viaje.centro_costo or "",
                "salida": "",
                "destino": "",
                "acompanantes": _texto_personal_viaje_excel(viaje),
                "motorista": _texto_motorista_excel(viaje),
                "vehiculo": _texto_vehiculo_excel(viaje),
                "nombre_paciente": "",
                "identidad_paciente": "",
                "costo_viaticos": total_viaticos,
                "tipo_viaje": dict(Viaje.TipoProgramacion.choices).get(viaje.tipo_viaje, ""),
                "hora_salida": _formatear_hora_excel(getattr(ejecucion, "fecha_salida", None)),
                "hora_llegada": _formatear_hora_excel(getattr(ejecucion, "fecha_retorno", None)),
                "total_horas": total_horas,
                "km_salida": ejecucion.kilometraje_salida if ejecucion and ejecucion.kilometraje_salida is not None else None,
                "km_entrada": ejecucion.kilometraje_retorno if ejecucion and ejecucion.kilometraje_retorno is not None else None,
                "total_km": total_km,
                "pacientes": 0,
            }
        )

    return filas, total_pacientes, total_km, total_viaticos, total_horas


def _estilizar_encabezado_excel(ws, fila_encabezado, total_columnas):
    rango = f"A{fila_encabezado}:{get_column_letter(total_columnas)}{fila_encabezado}"
    ServiceExcel.aplicar_estilo_rango(
        ws,
        rango,
        font=ServiceExcel.FONT_TITULO_TABLA,
        fill=ServiceExcel.FILL_TITULO_TABLA,
        align=ServiceExcel.ALIGN_TITULO,
        border=ServiceExcel.BORDER_GENERAL,
    )
    ws.row_dimensions[fila_encabezado].height = ServiceExcel.ROW_HEIGHT_ENCABEZADO


def _escribir_filas_excel(ws, filas, columnas_texto, columnas_numero):
    fila_actual = ws.max_row + 1
    for indice, fila in enumerate(filas, start=1):
        valores = [fila[columna] for columna in [
            "numero_orden",
            "fecha_solicitud",
            "fecha_salida",
            "prioridad",
            "departamento_sala",
            "centro_costo",
            "salida",
            "destino",
            "acompanantes",
            "motorista",
            "vehiculo",
            "nombre_paciente",
            "identidad_paciente",
            "costo_viaticos",
            "tipo_viaje",
            "hora_salida",
            "hora_llegada",
            "total_horas",
            "km_salida",
            "km_entrada",
            "total_km",
            "pacientes",
        ]]
        ws.append(valores)
        fill = ServiceExcel.FILL_ZEBRA_1 if fila_actual % 2 == 0 else ServiceExcel.FILL_ZEBRA_2
        ServiceExcel.aplicar_estilo_rango(
            ws,
            f"A{fila_actual}:V{fila_actual}",
            fill=fill,
            border=ServiceExcel.BORDER_GENERAL,
            font=ServiceExcel.FONT_GENERAL,
            align=ServiceExcel.ALIGN_GENERAL,
        )
        for col in columnas_texto:
            ws.cell(row=fila_actual, column=col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for col in columnas_numero:
            ws.cell(row=fila_actual, column=col).alignment = Alignment(horizontal="center", vertical="center")
        fila_actual += 1


def _generar_excel_viajes_historial(request, resumen_config):
    if not resumen_config.get("valido", True):
        return JsonResponse({"error": resumen_config.get("error", "Rango inválido")}, status=400)

    resumen_vista = (request.GET.get("vista") or "viajes").strip().lower()
    if resumen_vista != "viajes":
        return JsonResponse({"error": "El reporte Excel solo está disponible para la vista Por viajes."}, status=400)

    qs_viajes = _obtener_qs_viajes_historial(request.user).filter(
        fecha_programacion__gte=resumen_config["inicio"],
        fecha_programacion__lte=resumen_config["fin"],
    )

    viajes = list(qs_viajes)
    if not viajes:
        return JsonResponse({"error": "No existen viajes para el período seleccionado."}, status=404)

    viajes_por_motorista = defaultdict(list)
    consolidado = defaultdict(lambda: {
        "motorista": "SIN MOTORISTA",
        "total_viajes": 0,
        "total_pacientes": 0,
        "total_km": Decimal("0.00"),
        "total_viaticos": Decimal("0.00"),
    })

    for viaje in viajes:
        nombre_motorista = _nombre_motorista_excel(viaje)
        filas, total_pacientes, total_km, total_viaticos, _ = _viaje_filas_excel(viaje)
        viajes_por_motorista[nombre_motorista].append(
            {
                "viaje": viaje,
                "filas": filas,
                "total_pacientes": total_pacientes,
                "total_km": total_km,
                "total_viaticos": total_viaticos,
            }
        )

        resumen = consolidado[nombre_motorista]
        resumen["motorista"] = nombre_motorista
        resumen["total_viajes"] += 1
        resumen["total_pacientes"] += total_pacientes
        resumen["total_km"] += total_km if isinstance(total_km, Decimal) else Decimal(str(total_km or 0))
        resumen["total_viaticos"] += total_viaticos if isinstance(total_viaticos, Decimal) else Decimal(str(total_viaticos or 0))

    wb = Workbook()
    ws_consolidado = wb.active
    ws_consolidado.title = "CONSOLIDADO"

    fecha_inicio = resumen_config["inicio"]
    fecha_fin = resumen_config["fin"]
    mes_inicio_nombre = mes_nombre(fecha_inicio.month)
    mes_fin_nombre = mes_nombre(fecha_fin.month)
    usuario = getattr(request.user, "username", "")
    usuario_nombre = f"{getattr(request.user, 'first_name', '')} {getattr(request.user, 'last_name', '')}".strip() or usuario
    fecha_actual = timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")

    def preparar_hoja(ws, total_columnas, titulo, subtitulo):
        ServiceExcel.dibujar_encabezado_excel(ws, col_fin=get_column_letter(total_columnas))
        fila_titulo = ws.max_row + 2
        ws.merge_cells(f"A{fila_titulo}:{get_column_letter(total_columnas)}{fila_titulo}")
        titulo_celda = ws.cell(row=fila_titulo, column=1)
        titulo_celda.value = titulo
        titulo_celda.font = ServiceExcel.FONT_TITULO
        titulo_celda.alignment = ServiceExcel.ALIGN_TITULO
        ws.row_dimensions[fila_titulo].height = 22

        fila_subtitulo = fila_titulo + 1
        ws.merge_cells(f"A{fila_subtitulo}:{get_column_letter(total_columnas)}{fila_subtitulo}")
        subtitulo_celda = ws.cell(row=fila_subtitulo, column=1)
        subtitulo_celda.value = subtitulo
        subtitulo_celda.font = ServiceExcel.FONT_SUBTITULO
        subtitulo_celda.alignment = ServiceExcel.ALIGN_TITULO
        ws.row_dimensions[fila_subtitulo].height = 20

    def agregar_pie(ws):
        ServiceExcel.dibujar_pie_excel(ws, fecha_actual, usuario, usuario_nombre)

    columnas_consolidado = ["MOTORISTA", "TOTAL VIAJES", "TOTAL PACIENTES", "TOTAL KM", "TOTAL VIÁTICOS"]
    preparar_hoja(
        ws_consolidado,
        len(columnas_consolidado),
        "REPORTE EXCEL HISTÓRICO DE VIAJES POR MOTORISTA",
        f"Período: {mes_inicio_nombre} {fecha_inicio.year} - {mes_fin_nombre} {fecha_fin.year}",
    )
    ws_consolidado.append([])
    ws_consolidado.append(columnas_consolidado)
    fila_encabezado = ws_consolidado.max_row
    _estilizar_encabezado_excel(ws_consolidado, fila_encabezado, len(columnas_consolidado))

    for col, ancho in {1: 30, 2: 14, 3: 16, 4: 14, 5: 16}.items():
        ws_consolidado.column_dimensions[get_column_letter(col)].width = ancho

    fila_actual = ws_consolidado.max_row + 1
    for nombre_motorista in sorted(consolidado.keys(), key=lambda valor: valor.lower()):
        item = consolidado[nombre_motorista]
        ws_consolidado.append([
            item["motorista"],
            item["total_viajes"],
            item["total_pacientes"],
            float(item["total_km"]),
            float(item["total_viaticos"]),
        ])
        fill = ServiceExcel.FILL_ZEBRA_1 if fila_actual % 2 == 0 else ServiceExcel.FILL_ZEBRA_2
        ServiceExcel.aplicar_estilo_rango(
            ws_consolidado,
            f"A{fila_actual}:E{fila_actual}",
            fill=fill,
            border=ServiceExcel.BORDER_GENERAL,
            font=ServiceExcel.FONT_GENERAL,
            align=ServiceExcel.ALIGN_GENERAL,
        )
        ws_consolidado.cell(row=fila_actual, column=1).alignment = Alignment(horizontal="left", vertical="center")
        fila_actual += 1

    agregar_pie(ws_consolidado)
    ServiceExcel.configurar_impresion(ws_consolidado, fila_encabezado)

    usados = {"consolidado"}
    columnas_texto = {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 15, 16, 17, 22}
    columnas_numero = {14, 18, 19, 20, 21}

    for nombre_motorista in sorted(viajes_por_motorista.keys(), key=lambda valor: valor.lower()):
        ws = wb.create_sheet(title=_normalizar_nombre_hoja_excel(nombre_motorista, usados))
        preparar_hoja(
            ws,
            22,
            f"REPORTE DE VIAJES - {nombre_motorista}",
            f"Período: {mes_inicio_nombre} {fecha_inicio.year} - {mes_fin_nombre} {fecha_fin.year}",
        )
        ws.append([])
        headers = [
            "N° ORDEN",
            "FECHA DE SOLICITUD",
            "FECHA SALIDA",
            "PRIORIDAD",
            "DEPARTAMENTO / SALA",
            "CENTRO DE COSTO",
            "SALIDA",
            "DESTINO",
            "ACOMPAÑANTES",
            "MOTORISTA",
            "VEHÍCULO",
            "NOMBRE DEL PACIENTE",
            "NÚMERO DE IDENTIDAD DEL PACIENTE",
            "COSTO DE VIÁTICOS",
            "TIPO DE VIAJE",
            "HORA SALIDA",
            "HORA LLEGADA",
            "TOTAL HORAS",
            "KM. SALIDA",
            "KM. ENTRADA",
            "TOTAL KM",
            "PACIENTES",
        ]
        ws.append(headers)
        fila_encabezado = ws.max_row
        _estilizar_encabezado_excel(ws, fila_encabezado, len(headers))
        ws.freeze_panes = f"A{fila_encabezado + 1}"

        anchos = {
            1: 14,
            2: 18,
            3: 18,
            4: 14,
            5: 24,
            6: 14,
            7: 24,
            8: 24,
            9: 30,
            10: 24,
            11: 24,
            12: 26,
            13: 22,
            14: 16,
            15: 16,
            16: 12,
            17: 12,
            18: 12,
            19: 12,
            20: 12,
            21: 12,
            22: 10,
        }
        for col, ancho in anchos.items():
            ws.column_dimensions[get_column_letter(col)].width = ancho

        for bundle in viajes_por_motorista[nombre_motorista]:
            _escribir_filas_excel(ws, bundle["filas"], columnas_texto, columnas_numero)

        agregar_pie(ws)
        ServiceExcel.configurar_impresion(ws, fila_encabezado)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    nombre_archivo = (
        f"Reporte_Viajes_{mes_inicio_nombre}_{fecha_inicio.year}_{mes_fin_nombre}_{fecha_fin.year}.xlsx"
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response


def _viaje_asociado_solicitud(solicitud):
    for relacion in solicitud.viaje_solicitudes.all():
        if relacion.viaje_id:
            return relacion.viaje
    return None


def _resumen_pacientes_solicitud(solicitud):
    pacientes = []
    for item in solicitud.solicitud_pacientes.all():
        if item.paciente:
            pacientes.append(_nombre_paciente_completo(item.paciente))
    texto = ", ".join(pacientes[:2]) if pacientes else "Sin pacientes"
    if len(pacientes) > 2:
        texto += f" (+{len(pacientes) - 2})"
    return {
        "total": len(pacientes),
        "texto": texto,
    }


def _detalle_solicitud_payload(solicitud):
    pacientes = []
    for item in solicitud.solicitud_pacientes.select_related("paciente", "ingreso__sala"):
        paciente_obj = item.paciente
        pacientes.append(
            {
                "paciente": _nombre_paciente_completo(paciente_obj),
                "paciente_obj": paciente_obj,
                "expediente": str(getattr(paciente_obj, "expediente_numero", "-") or "-"),
                "ingreso": f"ING-{item.ingreso_id}" if item.ingreso_id else "-",
                "sala": item.ingreso.sala.nombre_sala if item.ingreso and item.ingreso.sala else "-",
            }
        )

    personal = [
        {
            "id": p.empleado_id,
            "nombre": _nombre_empleado(p.empleado),
            "cargo": _cargo_empleado(p.empleado),
            "unidad": getattr(_unidad_empleado(p.empleado), "nombre_corto_unidad", None)
            or getattr(_unidad_empleado(p.empleado), "nombre_unidad", None)
            or "-",
            "observacion": p.observacion or "-",
        }
        for p in solicitud.solicitud_personal.select_related("empleado")
    ]

    punto_solicitud = _info_punto_solicitud(solicitud.punto_solicitud)
    origen = _info_institucion(solicitud.lugar_salida)
    destino = _info_institucion(solicitud.lugar_destino)
    solicitante = _info_empleado_solicitante(solicitud.solicitante_empleado)
    pacientes_detalle = []
    for item in pacientes:
        paciente_obj = item.get("paciente_obj")
        pacientes_detalle.append(
            {
                "paciente": item["paciente"],
                "identidad": str(getattr(paciente_obj, "dni", "") or ""),
                "expediente": item["expediente"],
                "ingreso": item["ingreso"],
                "sala": item["sala"],
            }
        )

    viaje = _viaje_asociado_solicitud(solicitud)

    return {
        "id": solicitud.id,
        "numero_solicitud": solicitud.numero_solicitud,
        "fecha_solicitud": timezone.localtime(solicitud.fecha_solicitud).strftime("%d/%m/%Y %H:%M"),
        "origen": origen["nombre"],
        "destino": destino["nombre"],
        "tipo_solicitud": solicitud.tipo_solicitud.nombre,
        "prioridad": solicitud.prioridad.nombre,
        "estado": solicitud.proceso_funcional,
        "proceso": solicitud.proceso_funcional,
        "puede_editar": solicitud.puede_editar,
        "motivo": solicitud.motivo,
        "observaciones": solicitud.observaciones or "",
        "motivo_anulacion": solicitud.motivo_anulacion or "",
        "observacion_anulacion": solicitud.observacion_anulacion or "",
        "anulada_por": str(solicitud.anulada_por) if solicitud.anulada_por else "",
        "anulada_en": timezone.localtime(solicitud.anulada_en).strftime("%d/%m/%Y %H:%M") if solicitud.anulada_en else "",
        "punto_solicitud": punto_solicitud["nombre"],
        "punto_corto": punto_solicitud["nombre_corto"],
        "solicitante": solicitante,
        "pacientes": pacientes_detalle,
        "personal": personal,
        "viaje_asociado": {
            "id": viaje.id,
            "numero_viaje": viaje.numero_viaje,
            "estado": viaje.estado,
        } if viaje else None,
    }


def _generar_numero_solicitud():
    ultimo = Solicitud.objects.order_by("-id").values_list("id", flat=True).first() or 0
    return f"SOL-{ultimo + 1:06d}"


def _proceso_solicitud_label(estado):
    etiquetas = {
        Solicitud.Estado.PENDIENTE: "PENDIENTE",
        Solicitud.Estado.PROGRAMADA: "PROGRAMADA",
        Solicitud.Estado.EN_EJECUCION: "EN_EJECUCION",
        Solicitud.Estado.FINALIZADA: "FINALIZADA",
        Solicitud.Estado.ANULADA: "ANULADA",
    }
    return etiquetas.get(estado, estado or "-")


def _obtener_qs_solicitudes_base(usuario):
    viaje_solicitudes_activas = ViajeSolicitud.objects.filter(
        solicitud_id=OuterRef("pk"),
        activo=True,
    )
    qs = (
        Solicitud.objects
        .annotate(
            tiene_viaje_solicitud_activa=Exists(viaje_solicitudes_activas),
        )
        .select_related(
            "solicitante_empleado",
            "solicitante_empleado__personal_salud_empleado__servicio_unidad",
            "solicitante_empleado__personal_no_clinico__servicio_unidad",
            "punto_solicitud__unidad",
            "punto_solicitud__unidad_clinica",
            "punto_solicitud__unidad_clinica__area_atencion__servicio",
            "punto_solicitud__unidad_clinica__sala__servicio",
            "punto_solicitud__unidad_clinica__servicio_aux",
            "punto_solicitud__unidad_clinica__establecimiento_ext__nivel_complejidad_institucional",
            "punto_solicitud__unidad_clinica__establecimiento_ext__region_salud",
            "tipo_solicitud",
            "prioridad",
            "lugar_salida__nivel_complejidad_institucional",
            "lugar_salida__region_salud",
            "lugar_destino__nivel_complejidad_institucional",
            "lugar_destino__region_salud",
        )
        .prefetch_related(
            Prefetch(
                "viaje_solicitudes",
                queryset=ViajeSolicitud.objects.select_related("viaje").filter(activo=True),
            ),
            Prefetch(
                "solicitud_pacientes",
                queryset=SolicitudPaciente.objects.select_related("paciente", "ingreso__sala"),
            ),
            Prefetch(
                "solicitud_personal",
                queryset=SolicitudPersonal.objects.select_related(
                    "empleado",
                    "empleado__personal_salud_empleado__tipo_personal_salud",
                    "empleado__personal_no_clinico",
                ),
            ),
        )
        .order_by("-fecha_solicitud")
    )

    if _puede_ver_todas_solicitudes(usuario):
        return qs

    empleado_id = getattr(getattr(usuario, "empleado", None), "id", None)
    if not empleado_id:
        return qs.none()

    return qs.filter(solicitante_empleado_id=empleado_id)


def _obtener_qs_solicitudes_activas(usuario):
    return _obtener_qs_solicitudes_base(usuario).filter(
        activo=True,
    ).exclude(
        estado__in=[Solicitud.Estado.ANULADA, Solicitud.Estado.FINALIZADA],
    )


def _obtener_qs_solicitudes_historial(usuario):
    return _obtener_qs_solicitudes_base(usuario)


def _obtener_qs_solicitudes_resumen(usuario, vista):
    qs = _obtener_qs_solicitudes_historial(usuario)
    if vista == "denegadas":
        return qs.filter(estado=Solicitud.Estado.ANULADA).order_by("-fecha_solicitud", "numero_solicitud")
    return qs.filter(estado__in=[Solicitud.Estado.PROGRAMADA, Solicitud.Estado.EN_EJECUCION, Solicitud.Estado.FINALIZADA]).order_by("-fecha_solicitud", "numero_solicitud")


def _serializar_solicitud_resumen(solicitud):
    pacientes = _resumen_pacientes_solicitud(solicitud)
    viaje = _viaje_asociado_solicitud(solicitud)
    detalle = _detalle_solicitud_payload(solicitud)
    return {
        "id": solicitud.id,
        "numero": solicitud.numero_solicitud,
        "fecha": timezone.localtime(solicitud.fecha_solicitud).strftime("%d/%m/%Y %H:%M"),
        "punto": _punto_solicitud_corto(solicitud.punto_solicitud),
        "tipo": solicitud.tipo_solicitud.nombre,
        "prioridad": solicitud.prioridad.nombre,
        "prioridad_nivel": solicitud.prioridad.nivel,
        "pacientes": pacientes["texto"],
        "pacientes_total": pacientes["total"],
        "estado": solicitud.proceso_funcional,
        "viaje_asociado": viaje.numero_viaje if viaje else "-",
        "motivo_anulacion": solicitud.motivo_anulacion or "",
        "observacion_anulacion": solicitud.observacion_anulacion or "",
        "anulada_por": str(solicitud.anulada_por) if solicitud.anulada_por else "",
        "anulada_en": timezone.localtime(solicitud.anulada_en).strftime("%d/%m/%Y %H:%M") if solicitud.anulada_en else "",
        "detalle_json": json.dumps(detalle, ensure_ascii=False),
    }


def _puede_ver_todas_solicitudes(usuario):
    return bool(
        usuario
        and (
            usuario.is_superuser
            or UsuarioService.es_admin_global(usuario)
            or UsuarioService.es_directivo(usuario)
        )
    )


def _obtener_qs_solicitudes_propias(usuario):
    qs = _obtener_qs_solicitudes_activas(usuario)
    empleado_id = getattr(getattr(usuario, "empleado", None), "id", None)
    if not empleado_id:
        return qs.none()
    return qs.filter(solicitante_empleado_id=empleado_id)


def _obtener_qs_solicitudes_autorizacion(usuario):
    return (
        _obtener_qs_solicitudes_activas(usuario)
        .filter(estado=Solicitud.Estado.PENDIENTE)
        .exclude(viaje_solicitudes__activo=True)
        .annotate(cantidad_pacientes=Count("solicitud_pacientes", distinct=True))
        .distinct()
        .order_by("-fecha_solicitud")
    )


def _obtener_qs_viaje_construccion(usuario):
    return (
        ViajeSolicitud.objects
        .select_related(
            "solicitud__prioridad",
            "solicitud__tipo_solicitud",
            "solicitud__punto_solicitud",
            "solicitud__lugar_destino",
        )
        .filter(activo=True, viaje__isnull=True, solicitud__activo=True)
        .order_by("fecha_asignacion", "id")
    )


def _obtener_qs_viajes_ejecucion(usuario):
    return (
        Viaje.objects
        .filter(activo=True, estado__in=["PROGRAMADA", "EN_EJECUCION", "FINALIZADA"])
        .select_related(
            "vehiculo",
            "motorista__empleado",
            "ejecucion_viaje",
        )
        .prefetch_related(
            Prefetch(
                "viaje_viaticos",
                queryset=ViajeViatico.objects.select_related("viatico", "creado_por"),
            ),
            Prefetch(
                "viaje_solicitudes",
                queryset=ViajeSolicitud.objects.select_related(
                    "solicitud",
                    "solicitud__prioridad",
                    "solicitud__tipo_solicitud",
                    "solicitud__punto_solicitud",
                    "solicitud__lugar_destino",
                ).prefetch_related(
                    Prefetch(
                        "solicitud__solicitud_pacientes",
                        queryset=SolicitudPaciente.objects.select_related("paciente", "ingreso__sala"),
                    ),
                    Prefetch(
                        "solicitud__solicitud_personal",
                        queryset=SolicitudPersonal.objects.select_related(
                            "empleado",
                            "empleado__personal_salud_empleado__tipo_personal_salud",
                            "empleado__personal_no_clinico",
                        ),
                    ),
                ).filter(activo=True),
            ),
            Prefetch(
                "viaje_personal",
                queryset=ViajePersonal.objects.select_related(
                    "empleado",
                    "empleado__personal_salud_empleado__tipo_personal_salud",
                    "empleado__personal_no_clinico",
                ),
            )
        )
        .order_by("-fecha_programacion", "numero_viaje")
    )


def _obtener_qs_viajes_historial(usuario):
    return (
        Viaje.objects
        .select_related(
            "vehiculo",
            "motorista__empleado",
            "ejecucion_viaje",
        )
        .prefetch_related(
            Prefetch(
                "viaje_solicitudes",
                queryset=ViajeSolicitud.objects.select_related(
                    "solicitud",
                    "solicitud__prioridad",
                    "solicitud__tipo_solicitud",
                    "solicitud__punto_solicitud",
                    "solicitud__lugar_destino",
                ).prefetch_related(
                    Prefetch(
                        "solicitud__solicitud_pacientes",
                        queryset=SolicitudPaciente.objects.select_related("paciente", "ingreso__sala"),
                    ),
                    Prefetch(
                        "solicitud__solicitud_personal",
                        queryset=SolicitudPersonal.objects.select_related(
                            "empleado",
                            "empleado__personal_salud_empleado__tipo_personal_salud",
                            "empleado__personal_no_clinico",
                        ),
                    ),
                ).filter(activo=True),
            ),
            Prefetch(
                "viaje_personal",
                queryset=ViajePersonal.objects.select_related(
                    "empleado",
                    "empleado__personal_salud_empleado__tipo_personal_salud",
                    "empleado__personal_no_clinico",
                ),
            )
        )
        .order_by("-fecha_programacion", "numero_viaje")
    )


@require_GET
def api_resumen_exportar_excel(request):
    if not puede_ver_modulo(request.user):
        return JsonResponse({"error": "Acceso denegado."}, status=403)

    resumen_config = _limites_periodo_resumen(request)
    return _generar_excel_viajes_historial(request, resumen_config)


def _parse_payload_json(valor):
    if not valor:
        return []
    try:
        data = json.loads(valor)
        return data if isinstance(data, list) else []
    except (TypeError, ValueError, json.JSONDecodeError):
        return []


def _generar_numero_viaje():
    ultimo = Viaje.objects.order_by("-id").values_list("id", flat=True).first() or 0
    return f"VIA-{ultimo + 1:06d}"


def _serializar_viaje_solicitud(viaje_solicitud):
    solicitud = viaje_solicitud.solicitud
    personal = [
        {
            "id": item.empleado_id,
            "nombre": _nombre_empleado(item.empleado),
            "cargo": _cargo_empleado(item.empleado),
            "unidad": getattr(_unidad_empleado(item.empleado), "nombre_corto_unidad", None)
            or getattr(_unidad_empleado(item.empleado), "nombre_unidad", None)
            or "-",
        }
        for item in solicitud.solicitud_personal.select_related("empleado")
    ]
    pacientes = []
    for item in solicitud.solicitud_pacientes.select_related("paciente", "ingreso__sala"):
        nombre_paciente = "-"
        if item.paciente:
            nombre_paciente = " ".join(
                x for x in [
                    item.paciente.primer_nombre,
                    item.paciente.segundo_nombre,
                    item.paciente.primer_apellido,
                    item.paciente.segundo_apellido,
                ] if x
            ) or "-"
        pacientes.append(
            {
                "id": item.paciente_id,
                "nombre": nombre_paciente,
                "identidad": str(getattr(item.paciente, "dni", "-") or "-"),
                "expediente": str(getattr(item.paciente, "expediente_numero", "-") or "-"),
                "ingreso": f"ING-{item.ingreso_id}" if item.ingreso_id else "-",
                "sala": item.ingreso.sala.nombre_sala if item.ingreso and item.ingreso.sala else "-",
            }
        )
    return {
        "id": viaje_solicitud.id,
        "solicitud_id": solicitud.id,
        "numero": solicitud.numero_solicitud,
        "fecha": timezone.localtime(solicitud.fecha_solicitud).strftime("%d/%m/%Y %H:%M"),
        "punto": _area_solicitante_desde_punto(solicitud.punto_solicitud),
        "punto_corto": _punto_solicitud_corto(solicitud.punto_solicitud),
        "tipo": solicitud.tipo_solicitud.nombre,
        "prioridad": solicitud.prioridad.nombre,
        "prioridad_nivel": solicitud.prioridad.nivel,
        "destino": solicitud.lugar_destino.nombre_institucion_salud,
        "pacientes": solicitud.solicitud_pacientes.count(),
        "pacientesCount": solicitud.solicitud_pacientes.count(),
        "pacientes_detalle": pacientes,
        "personalCount": len(personal),
        "personal": personal,
    }


def _serializar_viatico(viatico):
    return {
        "id": viatico.id,
        "codigo": viatico.codigo,
        "nombre": viatico.nombre,
        "monto_vigente": str(viatico.monto_vigente) if viatico.monto_vigente is not None else "",
    }


def _estado_ejecucion_viaje(viaje):
    ejecucion = getattr(viaje, "ejecucion_viaje", None)
    if not ejecucion or not ejecucion.fecha_salida:
        return "PROGRAMADO"
    if not ejecucion.fecha_retorno:
        return "EN_EJECUCION"
    return "FINALIZADO"


def _serializar_viaje_ejecucion(viaje):
    ejecucion = getattr(viaje, "ejecucion_viaje", None)
    solicitudes = list(viaje.viaje_solicitudes.all())
    solicitud_principal = solicitudes[0].solicitud if solicitudes else None
    estado_ejecucion = _estado_ejecucion_viaje(viaje)
    total_combustible = getattr(ejecucion, "total_combustible", None) if ejecucion else None
    if total_combustible is None and ejecucion and ejecucion.precio_litro_salida is not None and ejecucion.litros_cargados_salida is not None:
        total_combustible = Decimal(str(ejecucion.precio_litro_salida)) * Decimal(str(ejecucion.litros_cargados_salida))

    tipo_viaje = dict(Viaje.TipoProgramacion.choices).get(viaje.tipo_viaje, "-")
    personal_operativo = []
    for item in viaje.viaje_personal.all():
        personal_operativo.append(
            {
                "id": item.empleado_id,
                "nombre": _nombre_empleado(item.empleado),
                "cargo": _cargo_empleado(item.empleado),
                "unidad": getattr(_unidad_empleado(item.empleado), "nombre_corto_unidad", None)
                or getattr(_unidad_empleado(item.empleado), "nombre_unidad", None)
                or "-",
                "participacion": item.tipo_participacion or "-",
            }
        )

    pacientes_consolidados = []
    solicitudes_detalle = []
    motivos_consolidados = []
    observaciones_viaje = []
    if ejecucion:
        if ejecucion.observaciones:
            observaciones_viaje.append({"etapa": "General", "texto": ejecucion.observaciones})
        if ejecucion.observaciones_salida:
            observaciones_viaje.append({"etapa": "Salida", "texto": ejecucion.observaciones_salida})
        if ejecucion.observaciones_retorno:
            observaciones_viaje.append({"etapa": "Retorno", "texto": ejecucion.observaciones_retorno})
    for viaje_solicitud in solicitudes:
        solicitud = viaje_solicitud.solicitud
        pacientes_solicitud = []
        for item in solicitud.solicitud_pacientes.all():
            nombre_paciente = "-"
            if item.paciente:
                nombre_paciente = " ".join(
                    x for x in [
                        item.paciente.primer_nombre,
                        item.paciente.segundo_nombre,
                        item.paciente.primer_apellido,
                        item.paciente.segundo_apellido,
                    ] if x
                ) or "-"
            paciente_data = {
                "solicitud": solicitud.numero_solicitud,
                "paciente": nombre_paciente,
                "identidad": str(getattr(item.paciente, "dni", "-") or "-"),
            }
            pacientes_consolidados.append(paciente_data)
            pacientes_solicitud.append(paciente_data)

        motivo_textos = []
        if solicitud.motivo:
            motivo_textos.append(solicitud.motivo)
        solicitudes_detalle.append(
            {
                "numero": solicitud.numero_solicitud,
                "punto": _punto_solicitud_corto(solicitud.punto_solicitud),
                "prioridad": solicitud.prioridad.nombre,
                "motivos": motivo_textos,
                "pacientes": pacientes_solicitud,
            }
        )
        motivos_consolidados.append(
            {
                "numero": solicitud.numero_solicitud,
                "items": motivo_textos,
            }
        )

    detalle = {
        "info_general": {
            "numero": viaje.numero_viaje,
            "estado": viaje.estado,
            "estado_etiqueta": {
                "PROGRAMADO": "Programado",
                "EN_EJECUCION": "En ejecución",
                "FINALIZADO": "Finalizado",
            }.get(estado_ejecucion, "Programado"),
            "tipo_viaje": tipo_viaje,
            "ambulancia": f"{viaje.vehiculo.codigo} - {viaje.vehiculo.placa}" if viaje.vehiculo else "-",
            "motorista": _nombre_empleado(viaje.motorista.empleado) if viaje.motorista and viaje.motorista.empleado else "-",
        },
        "personal_operativo": personal_operativo,
        "pacientes": pacientes_consolidados,
        "solicitudes": solicitudes_detalle,
        "motivos": motivos_consolidados,
        "observaciones_viaje": observaciones_viaje,
    }
    return {
        "id": viaje.id,
        "numero": viaje.numero_viaje,
        "fecha": timezone.localtime(viaje.fecha_programacion).strftime("%d/%m/%Y %H:%M"),
        "vehiculo": f"{viaje.vehiculo.codigo} - {viaje.vehiculo.placa}" if viaje.vehiculo else "-",
        "motorista": _nombre_empleado(viaje.motorista.empleado) if viaje.motorista and viaje.motorista.empleado else "-",
        "tipo_viaje": tipo_viaje,
        "ambulancia": f"{viaje.vehiculo.codigo} - {viaje.vehiculo.placa}" if viaje.vehiculo else "-",
        "estado": viaje.estado,
        "estado_ejecucion": estado_ejecucion,
        "estado_etiqueta": {
            "PROGRAMADO": "Programado",
            "EN_EJECUCION": "En ejecución",
            "FINALIZADO": "Finalizado",
        }.get(estado_ejecucion, "Programado"),
        "solicitudes": len(solicitudes),
        "solicitud_principal": solicitud_principal.numero_solicitud if solicitud_principal else "-",
        "salida_registrada": bool(ejecucion and ejecucion.fecha_salida),
        "entrada_registrada": bool(ejecucion and ejecucion.fecha_retorno),
        "fecha_salida_input": timezone.localtime(ejecucion.fecha_salida).strftime("%Y-%m-%dT%H:%M") if ejecucion and ejecucion.fecha_salida else "",
        "fecha_salida": timezone.localtime(ejecucion.fecha_salida).strftime("%d/%m/%Y %H:%M") if ejecucion and ejecucion.fecha_salida else "",
        "kilometraje_salida": str(ejecucion.kilometraje_salida) if ejecucion and ejecucion.kilometraje_salida is not None else "",
        "precio_litro_salida": str(ejecucion.precio_litro_salida) if ejecucion and ejecucion.precio_litro_salida is not None else "",
        "litros_cargados_salida": str(ejecucion.litros_cargados_salida) if ejecucion and ejecucion.litros_cargados_salida is not None else "",
        "total_combustible": str(total_combustible) if total_combustible is not None else "",
        "observaciones_salida": ejecucion.observaciones_salida if ejecucion and ejecucion.observaciones_salida else "",
        "fecha_retorno_input": timezone.localtime(ejecucion.fecha_retorno).strftime("%Y-%m-%dT%H:%M") if ejecucion and ejecucion.fecha_retorno else "",
        "fecha_retorno": timezone.localtime(ejecucion.fecha_retorno).strftime("%d/%m/%Y %H:%M") if ejecucion and ejecucion.fecha_retorno else "",
        "kilometraje_retorno": str(ejecucion.kilometraje_retorno) if ejecucion and ejecucion.kilometraje_retorno is not None else "",
        "observaciones_retorno": ejecucion.observaciones_retorno if ejecucion and ejecucion.observaciones_retorno else "",
        "puede_marcar_salida": bool(not ejecucion or not ejecucion.fecha_salida),
        "puede_marcar_entrada": bool(ejecucion and ejecucion.fecha_salida and not ejecucion.fecha_retorno),
        "puede_editar": True,
        "detalle_json": json.dumps(detalle, ensure_ascii=False),
    }


def _serializar_vehiculo(vehiculo):
    return {
        "id": vehiculo.id,
        "texto": f"{vehiculo.codigo} - {vehiculo.placa} | {vehiculo.descripcion or 'Sin descripción'}",
    }


def _serializar_motorista(motorista):
    empleado = motorista.empleado
    nombre = " ".join(x for x in [empleado.primer_nombre, empleado.primer_apellido] if x) or str(empleado)
    return {
        "id": motorista.id,
        "texto": f"{empleado.dni} - {nombre}",
    }


def _nombre_empleado(empleado):
    if not empleado:
        return "-"
    return " ".join(x for x in [empleado.primer_nombre, empleado.primer_apellido] if x) or "-"


def _cargo_empleado(empleado):
    if not empleado:
        return "Personal"
    try:
        return empleado.personal_salud_empleado.tipo_personal_salud.descripcion_tipo_personal_salud
    except Exception:
        try:
            return empleado.personal_no_clinico.get_tipo_display()
        except Exception:
            return "Personal"


def _unidad_empleado(empleado):
    if not empleado:
        return None
    try:
        return empleado.personal_salud_empleado.servicio_unidad
    except Exception:
        try:
            return empleado.personal_no_clinico.servicio_unidad
        except Exception:
            return None


def _info_empleado_solicitante(empleado):
    unidad = _unidad_empleado(empleado)
    return {
        "nombre": _nombre_empleado(empleado),
        "unidad": getattr(unidad, "nombre_unidad", "-") or "-",
        "nombre_corto": getattr(unidad, "nombre_corto_unidad", "-") or "-",
    }


def _texto_nivel_complejidad(institucion):
    nivel = getattr(institucion, "nivel_complejidad_institucional", None)
    if not nivel:
        return "-"
    if getattr(nivel, "siglas", None):
        return f"Nivel {nivel.siglas}"
    if getattr(nivel, "nivel_complejidad", None):
        return f"Nivel {nivel.nivel_complejidad}"
    return "-"


def _info_institucion(institucion):
    if not institucion:
        return {"nombre": "-", "alias": "", "nivel": "-", "region": "-"}
    return {
        "nombre": institucion.nombre_institucion_salud,
        "alias": "",
        "nivel": _texto_nivel_complejidad(institucion),
        "region": getattr(institucion.region_salud, "nombre_region_salud", "-") or "-",
    }


def _info_punto_solicitud(punto):
    if not punto:
        return {
            "nombre": "-",
            "nombre_corto": "-",
            "tipo": "-",
            "servicio": "-",
            "servicio_corto": "-",
            "nivel": "-",
            "region": "-",
        }

    if punto.unidad:
        return {
            "nombre": punto.unidad.nombre_unidad,
            "nombre_corto": punto.unidad.nombre_corto_unidad or punto.unidad.nombre_unidad,
            "tipo": "Unidad Administrativa",
            "servicio": "-",
            "servicio_corto": "-",
            "nivel": "-",
            "region": "-",
        }

    unidad_clinica = punto.unidad_clinica
    if not unidad_clinica:
        return {
            "nombre": "-",
            "nombre_corto": "-",
            "tipo": "Unidad Clínica",
            "servicio": "-",
            "servicio_corto": "-",
            "nivel": "-",
            "region": "-",
        }

    tipo_unidad = unidad_clinica.get_tipo_unidad()[1]

    if unidad_clinica.area_atencion:
        servicio = unidad_clinica.area_atencion.servicio
        return {
            "nombre": unidad_clinica.area_atencion.nombre_area_atencion,
            "nombre_corto": unidad_clinica.area_atencion.nombre_corto_area_atencion or unidad_clinica.area_atencion.servicio.nombre_corto,
            "tipo": tipo_unidad,
            "servicio": servicio.nombre_servicio,
            "servicio_corto": servicio.nombre_corto,
            "nivel": "-",
            "region": "-",
        }

    if unidad_clinica.sala:
        servicio = unidad_clinica.sala.servicio
        return {
            "nombre": unidad_clinica.sala.nombre_sala,
            "nombre_corto": unidad_clinica.sala.nombre_corto_sala or unidad_clinica.sala.nombre_sala,
            "tipo": tipo_unidad,
            "servicio": servicio.nombre_servicio,
            "servicio_corto": servicio.nombre_corto,
            "nivel": "-",
            "region": "-",
        }

    if unidad_clinica.servicio_aux:
        return {
            "nombre": unidad_clinica.servicio_aux.nombre_servicio_a,
            "nombre_corto": unidad_clinica.servicio_aux.nombre_corto_servicio_a or unidad_clinica.servicio_aux.nombre_servicio_a,
            "tipo": tipo_unidad,
            "servicio": unidad_clinica.servicio_aux.nombre_servicio_a,
            "servicio_corto": unidad_clinica.servicio_aux.nombre_corto_servicio_a or unidad_clinica.servicio_aux.nombre_servicio_a,
            "nivel": "-",
            "region": "-",
        }

    if unidad_clinica.establecimiento_ext:
        establecimiento = unidad_clinica.establecimiento_ext
        return {
            "nombre": establecimiento.nombre_institucion_salud,
            "nombre_corto": getattr(establecimiento.nivel_complejidad_institucional, "siglas", None) or establecimiento.nombre_institucion_salud,
            "tipo": tipo_unidad,
            "servicio": establecimiento.nombre_institucion_salud,
            "servicio_corto": getattr(establecimiento.nivel_complejidad_institucional, "siglas", None) or establecimiento.nombre_institucion_salud,
            "nivel": getattr(establecimiento.nivel_complejidad_institucional, "detalle_nivel_complejidad", None) or "-",
            "region": getattr(establecimiento.region_salud, "nombre_region_salud", None) or "-",
        }

    return {
        "nombre": str(unidad_clinica),
        "nombre_corto": str(unidad_clinica),
        "tipo": tipo_unidad,
        "servicio": "-",
        "servicio_corto": "-",
        "nivel": "-",
        "region": "-",
    }


def _metadata_opciones_formulario(form, solicitud, usuario):
    destinos = {
        str(destino.pk): _info_institucion(destino)
        for destino in form.fields["lugar_destino"].queryset
    }
    puntos = {
        str(punto.pk): _info_punto_solicitud(punto)
        for punto in form.fields["punto_solicitud"].queryset
    }
    empleado_solicitante = getattr(solicitud, "solicitante_empleado", None) or getattr(usuario, "empleado", None)
    return {
        "destinos_info_json": json.dumps(destinos),
        "puntos_info_json": json.dumps(puntos),
        "solicitante_info": _info_empleado_solicitante(empleado_solicitante),
    }


def _normalizar_pacientes_payload(pacientes_payload):
    pacientes_ids_guardados = set()
    pacientes_unicos = []

    for item in pacientes_payload or []:
        paciente_id = item.get("paciente_id")
        ingreso_id = item.get("ingreso_id")
        if not paciente_id and not ingreso_id:
            continue

        key = f"{paciente_id or 0}:{ingreso_id or 0}"
        if key in pacientes_ids_guardados:
            continue

        pacientes_ids_guardados.add(key)
        pacientes_unicos.append(item)

    return pacientes_unicos


def _validar_reglas_pacientes(form, pacientes_payload, solicitud_excluida_id=None):
    pacientes_unicos = _normalizar_pacientes_payload(pacientes_payload)
    tipo_solicitud = form.cleaned_data.get("tipo_solicitud")
    tipo_codigo = (getattr(tipo_solicitud, "codigo", "") or "").upper()

    if len(pacientes_unicos) > 3:
        form.add_error(None, "Una solicitud permite como máximo 3 pacientes.")

    if tipo_codigo == "PACIENTES" and not pacientes_unicos:
        form.add_error("tipo_solicitud", "Debe seleccionar al menos un paciente para este tipo de solicitud.")

    if tipo_codigo != "PACIENTES" and pacientes_unicos:
        form.add_error("tipo_solicitud", "Solo el tipo de solicitud PACIENTES permite asociar pacientes.")

    pacientes_ids = [item.get("paciente_id") for item in pacientes_unicos if item.get("paciente_id")]
    if pacientes_ids:
        conflictos = (
            SolicitudPaciente.objects
            .select_related("solicitud", "paciente")
            .filter(paciente_id__in=pacientes_ids, solicitud__activo=True)
            .order_by("-created_at", "-id")
        )
        if solicitud_excluida_id is not None:
            conflictos = conflictos.exclude(solicitud_id=solicitud_excluida_id)

        solicitudes_reportadas = {}
        for item in conflictos:
            if item.paciente_id in solicitudes_reportadas:
                continue
            paciente = item.paciente
            nombre_paciente = " ".join(
                x for x in [
                    getattr(paciente, "primer_nombre", None),
                    getattr(paciente, "segundo_nombre", None),
                    getattr(paciente, "primer_apellido", None),
                    getattr(paciente, "segundo_apellido", None),
                ] if x
            ) or f"Paciente {item.paciente_id}"
            solicitudes_reportadas[item.paciente_id] = (nombre_paciente, item.solicitud.numero_solicitud)

        for nombre_paciente, numero_solicitud in solicitudes_reportadas.values():
            form.add_error(None, f"El paciente {nombre_paciente} ya pertenece a la solicitud {numero_solicitud}.")

    return pacientes_unicos


class SGTransporteHospitalarioDashboardView(LoginRequiredMixin, TemplateView):
    template_name = "sg_transporte_hospitalario/dashboard.html"

    def _active_tab(self):
        active_tab = self.request.GET.get("tab", "solicitud")
        tab_keys = {tab["key"] for tab in TAB_SECTIONS}
        return active_tab if active_tab in tab_keys else "solicitud"

    def dispatch(self, request, *args, **kwargs):
        if not puede_ver_modulo(request.user):
            return redirect("acceso_denegado")
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        if not puede_gestionar_modulo(request.user):
            return JsonResponse({"ok": False, "error": "Acceso denegado."}, status=403)

        form = SolicitudCreateForm(request.POST)
        pacientes_payload = _parse_payload_json(request.POST.get("pacientes_json"))
        personal_payload = _parse_payload_json(request.POST.get("empleados_json"))

        empleado_usuario = getattr(request.user, "empleado", None)
        if empleado_usuario is None:
            form.add_error(None, "El usuario actual no tiene empleado asociado.")

        if form.is_valid() and empleado_usuario is not None:
            pacientes_unicos = _validar_reglas_pacientes(form, pacientes_payload)

        if form.is_valid() and empleado_usuario is not None:
            with transaction.atomic():
                solicitud = form.save(commit=False)
                solicitud.numero_solicitud = _generar_numero_solicitud()
                solicitud.solicitante_empleado = empleado_usuario
                solicitud.estado = "PENDIENTE"
                solicitud.activo = True
                solicitud.save()

                for item in pacientes_unicos:
                    paciente_id = item.get("paciente_id")
                    ingreso_id = item.get("ingreso_id")

                    SolicitudPaciente.objects.create(
                        solicitud=solicitud,
                        paciente_id=paciente_id or None,
                        ingreso_id=ingreso_id or None,
                    )

                empleados_ids_guardados = set()
                for item in personal_payload:
                    empleado_id = item.get("empleado_id")
                    if not empleado_id or empleado_id in empleados_ids_guardados:
                        continue
                    empleados_ids_guardados.add(empleado_id)
                    SolicitudPersonal.objects.create(
                        solicitud=solicitud,
                        empleado_id=empleado_id,
                        observacion=item.get("observacion") or None,
                    )

            mostrar_mensaje(request, "success", f"Solicitud {solicitud.numero_solicitud} creada con estado PENDIENTE.")
            return redirect(f"{reverse('sg_transporte_hospitalario_dashboard')}?tab=solicitud")

        mostrar_mensaje(request, "error", "No se pudo guardar la solicitud. Revisa los datos ingresados.")
        self.form = form
        self.pacientes_payload = pacientes_payload
        self.personal_payload = personal_payload
        return self.get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(construir_contexto_dashboard(self.request.user))
        active_tab = self._active_tab()
        resumen_config = None
        resumen_vista = (self.request.GET.get("vista") or "viajes").strip().lower()
        if resumen_vista not in {"viajes", "autorizadas", "denegadas"}:
            resumen_vista = "viajes"
        solicitudes_activas = list(_obtener_qs_solicitudes_activas(self.request.user))
        solicitudes_autorizacion = list(_obtener_qs_solicitudes_autorizacion(self.request.user))
        viaje_construccion_items = list(_obtener_qs_viaje_construccion(self.request.user))
        viajes_ejecucion_items = list(_obtener_qs_viajes_ejecucion(self.request.user))
        usuario_empleado_id = getattr(getattr(self.request.user, "empleado", None), "id", None)

        for solicitud in solicitudes_activas:
            solicitud.es_propia = solicitud.solicitante_empleado_id == usuario_empleado_id
            solicitud.puede_editar_propia = bool(solicitud.es_propia and solicitud.puede_editar)
            solicitud.puede_eliminar_propia = bool(solicitud.es_propia and solicitud.puede_editar)

        for solicitud in solicitudes_autorizacion:
            personal = [
                {
                    "id": item.empleado_id,
                    "nombre": _nombre_empleado(item.empleado),
                    "cargo": _cargo_empleado(item.empleado),
                    "unidad": getattr(_unidad_empleado(item.empleado), "nombre_corto_unidad", None)
                    or getattr(_unidad_empleado(item.empleado), "nombre_unidad", None)
                    or "-",
                }
                for item in solicitud.solicitud_personal.select_related("empleado")
            ]
            solicitud.personal_json = json.dumps(personal, ensure_ascii=False)

        context["tabs"] = TAB_SECTIONS
        context["active_tab"] = active_tab
        context["tabs_url_base"] = reverse("sg_transporte_hospitalario_dashboard")
        context["usuario_empleado_id"] = usuario_empleado_id
        context["solicitud_form"] = getattr(self, "form", SolicitudCreateForm())
        context["solicitudes_activas"] = solicitudes_activas
        context["filtro_procesos"] = [
            ("Pendiente", "Pendiente"),
            ("En proceso", "En proceso"),
            ("Programada", "Programada"),
            ("En ejecución", "En ejecución"),
        ]
        context["solicitudes_autorizacion"] = solicitudes_autorizacion
        context["viaje_construccion_items"] = [
            _serializar_viaje_solicitud(item)
            for item in viaje_construccion_items
        ]
        context["viaje_construccion_total"] = len(viaje_construccion_items)
        context["viajes_ejecucion_items"] = [
            _serializar_viaje_ejecucion(item)
            for item in viajes_ejecucion_items
        ]
        context["viajes_ejecucion_total"] = len(viajes_ejecucion_items)
        context["programacion_form"] = ViajeProgramacionForm()
        context["filtro_estados"] = list(Solicitud.Estado.choices)
        context["filtro_areas"] = sorted(
            {
                _area_solicitante_desde_punto(solicitud.punto_solicitud)
                for solicitud in solicitudes_activas
                if _area_solicitante_desde_punto(solicitud.punto_solicitud) != "-"
            }
        )
        context["filtro_tipos"] = sorted(
            {
                solicitud.tipo_solicitud.nombre
                for solicitud in solicitudes_activas
                if solicitud.tipo_solicitud_id
            }
        )
        context["vehiculos_operativos"] = [
            _serializar_vehiculo(vehiculo)
            for vehiculo in Vehiculo.objects.filter(activo=True).order_by("codigo", "placa")
        ]
        context["motoristas_operativos"] = [
            _serializar_motorista(motorista)
            for motorista in Motorista.objects.filter(activo=True).select_related("empleado").order_by("id")
        ]
        context["viaticos_operativos"] = [
            _serializar_viatico(viatico)
            for viatico in Viatico.objects.filter(activo=True).order_by("codigo", "nombre")
        ]
        context["autorizacion_total"] = len(solicitudes_autorizacion)
        context["pacientes_payload"] = json.dumps(getattr(self, "pacientes_payload", []))
        context["personal_payload"] = json.dumps(getattr(self, "personal_payload", []))
        context["resumen_vista"] = resumen_vista
        context["resumen_vista_label"] = {
            "viajes": "Por viajes",
            "autorizadas": "Solicitudes autorizadas",
            "denegadas": "Solicitudes denegadas",
        }[resumen_vista]
        context["resumen_tabs"] = [
            {"key": "viajes", "label": "Por viajes"},
            {"key": "autorizadas", "label": "Solicitudes autorizadas"},
            {"key": "denegadas", "label": "Solicitudes denegadas"},
        ]
        context["resumen_titulo"] = "Resumen / Historial"
        context["resumen_config"] = resumen_config
        if active_tab == "resumen":
            resumen_config = _limites_periodo_resumen(self.request)
            context["resumen_config"] = resumen_config
            context["resumen_error"] = resumen_config.get("error") if not resumen_config.get("valido", True) else ""
            if resumen_config.get("valido", True):
                if resumen_vista == "viajes":
                    qs_resumen = _obtener_qs_viajes_historial(self.request.user).filter(
                        fecha_programacion__gte=resumen_config["inicio"],
                        fecha_programacion__lte=resumen_config["fin"],
                    )
                    context["resumen_items"] = [
                        _serializar_viaje_ejecucion(viaje)
                        for viaje in qs_resumen
                    ]
                    context["resumen_total"] = len(context["resumen_items"])
                    context["resumen_empty"] = "No hay viajes para los filtros seleccionados."
                else:
                    qs_resumen = _obtener_qs_solicitudes_resumen(self.request.user, resumen_vista).filter(
                        fecha_solicitud__gte=resumen_config["inicio"],
                        fecha_solicitud__lte=resumen_config["fin"],
                    )
                    context["resumen_items"] = [
                        _serializar_solicitud_resumen(solicitud)
                        for solicitud in qs_resumen
                    ]
                    context["resumen_total"] = len(context["resumen_items"])
                    context["resumen_empty"] = "No hay solicitudes para los filtros seleccionados."
            else:
                context["resumen_items"] = []
                context["resumen_total"] = 0
                context["resumen_empty"] = resumen_config["error"]
            context["resumen_mes_inicio"] = resumen_config["mes_inicio"]
            context["resumen_mes_fin"] = resumen_config["mes_fin"]
        else:
            context["resumen_items"] = []
            context["resumen_total"] = 0
            context["resumen_empty"] = "No hay registros para los filtros seleccionados."
            hoy = timezone.localdate().strftime("%Y-%m")
            context["resumen_mes_inicio"] = hoy
            context["resumen_mes_fin"] = hoy
            context["resumen_error"] = ""
        return context


@require_GET
def api_estado_modulo(request):
    if not puede_ver_modulo(request.user):
        return JsonResponse({"ok": False, "error": "Acceso denegado."}, status=403)

    return JsonResponse(
        {
            "ok": True,
            "modulo": "SG-transporte_hospitalario",
            "lectura": True,
            "escritura": puede_gestionar_modulo(request.user),
        }
    )


@require_GET
def api_buscar_pacientes(request):
    if not puede_ver_modulo(request.user):
        return JsonResponse({"ok": False, "error": "Acceso denegado."}, status=403)

    termino = (request.GET.get("q") or "").strip()
    tipo = (request.GET.get("tipo") or "nombre").strip().lower()
    if len(termino) < 2:
        return JsonResponse({"data": []})

    if tipo == "identidad":
        q_base = Q(dni__icontains=termino)
    else:
        # Reutiliza el patrón de búsqueda por nombre compuesto usado en s_exp:
        # divide por palabras y exige coincidencia de cada palabra en algún campo nominal.
        palabras = [p for p in termino.split() if p]
        q_base = Q()
        for palabra in palabras:
            q_base &= (
                Q(primer_nombre__icontains=palabra)
                | Q(segundo_nombre__icontains=palabra)
                | Q(primer_apellido__icontains=palabra)
                | Q(segundo_apellido__icontains=palabra)
            )
        q_base = q_base | Q(dni__icontains=termino)

    if termino.isdigit():
        q_base = q_base | Q(expediente_numero=int(termino))

    ingresos_prefetch = Prefetch(
        "pacientes_ingresados",
        queryset=Ingreso.objects.filter(estado=1).select_related("sala").order_by("-fecha_ingreso"),
    )
    solicitudes_prefetch = Prefetch(
        "solicitudes_transporte",
        queryset=(
            SolicitudPaciente.objects
            .select_related("solicitud")
            .filter(solicitud__activo=True)
            .order_by("-created_at", "-id")
        ),
        to_attr="solicitudes_activas_transporte",
    )

    pacientes = (
        Paciente.objects
        .filter(q_base)
        .prefetch_related(ingresos_prefetch, solicitudes_prefetch)
        .order_by("primer_nombre", "primer_apellido")[:20]
    )

    data = []
    for p in pacientes:
        ingreso_activo = p.pacientes_ingresados.first()
        solicitud_activa = (getattr(p, "solicitudes_activas_transporte", []) or [None])[0]
        solicitud_activa = solicitud_activa.solicitud if solicitud_activa else None
        data.append(
            {
                "paciente_id": p.id,
                "nombre": " ".join(
                    x for x in [p.primer_nombre, p.segundo_nombre, p.primer_apellido, p.segundo_apellido] if x
                ),
                "expediente": str(p.expediente_numero or "-"),
                "ingreso_id": ingreso_activo.id if ingreso_activo else None,
                "ingreso": f"ING-{ingreso_activo.id}" if ingreso_activo else "-",
                "sala": ingreso_activo.sala.nombre_sala if ingreso_activo and ingreso_activo.sala else "-",
                "solicitud_numero": solicitud_activa.numero_solicitud if solicitud_activa else None,
                "solicitud_estado": solicitud_activa.estado if solicitud_activa else None,
            }
        )

    return JsonResponse({"data": data})


@require_GET
def api_buscar_empleados(request):
    if not puede_ver_modulo(request.user):
        return JsonResponse({"ok": False, "error": "Acceso denegado."}, status=403)

    termino = (request.GET.get("q") or "").strip()
    tipo = (request.GET.get("tipo") or "nombre").strip().lower()
    solo_disponibles = str(request.GET.get("disponibles") or "").lower() in {"1", "true", "si", "yes"}
    if len(termino) < 2:
        return JsonResponse({"data": []})

    base_qs = _empleados_operativos_disponibles_qs() if solo_disponibles else Empleado.objects.all()

    if tipo == "identidad":
        filtro = Q(dni__icontains=termino)
    else:
        palabras = [p for p in termino.split() if p]
        filtro = Q()
        for palabra in palabras:
            filtro &= (
                Q(primer_nombre__icontains=palabra)
                | Q(segundo_nombre__icontains=palabra)
                | Q(primer_apellido__icontains=palabra)
                | Q(segundo_apellido__icontains=palabra)
            )
        filtro = filtro | Q(dni__icontains=termino)

    empleados = (
        base_qs
        .filter(filtro)
        .select_related("personal_salud_empleado__tipo_personal_salud", "personal_no_clinico")
        .order_by("primer_nombre", "primer_apellido")[:20]
    )

    data = []
    for e in empleados:
        data.append(
            {
                "empleado_id": e.id,
                "nombre": _nombre_empleado(e),
                "cargo": _cargo_empleado(e),
                "dni": e.dni,
                "unidad": getattr(_unidad_empleado(e), "nombre_unidad", "-") or "-",
                "unidad_corta": getattr(_unidad_empleado(e), "nombre_corto_unidad", "-") or "-",
            }
        )

    return JsonResponse({"data": data})


@require_GET
def api_solicitudes_activas(request):
    if not puede_ver_modulo(request.user):
        return JsonResponse({"ok": False, "error": "Acceso denegado."}, status=403)

    draw = int(request.GET.get("draw", 0))
    start = int(request.GET.get("start", 0))
    length = int(request.GET.get("length", 10))
    search = (request.GET.get("search[value]") or "").strip()
    estado_filtro = (request.GET.get("estado") or "").strip()

    qs = _obtener_qs_solicitudes_activas(request.user)
    usuario_empleado_id = getattr(getattr(request.user, "empleado", None), "id", None)

    total = qs.count()
    if estado_filtro:
        qs = qs.filter(estado=estado_filtro)

    if search:
        qs = qs.filter(
            Q(numero_solicitud__icontains=search)
            | Q(tipo_solicitud__nombre__icontains=search)
            | Q(prioridad__nombre__icontains=search)
            | Q(estado__icontains=search)
        )

    filtered = qs.count()
    rows = qs[start:start + length]

    data = []
    for s in rows:
        es_propia = s.solicitante_empleado_id == usuario_empleado_id
        data.append(
            {
                "id": s.id,
                "numero_solicitud": s.numero_solicitud,
                "fecha": timezone.localtime(s.fecha_solicitud).strftime("%d/%m/%Y %H:%M"),
                "creado_por": str(s.solicitante_empleado) if s.solicitante_empleado_id else "-",
                "area_solicitante": _area_solicitante_desde_punto(s.punto_solicitud),
                "tipo_solicitud": s.tipo_solicitud.nombre,
                "prioridad": s.prioridad.nombre,
                "proceso": s.proceso_funcional,
                "es_propia": es_propia,
                "puede_editar": bool(es_propia and s.puede_editar),
                "puede_eliminar": bool(es_propia and s.puede_editar),
                "url_ver": reverse("sg_transporte_hospitalario_solicitud_ver", kwargs={"pk": s.id}),
                "url_editar": reverse("sg_transporte_hospitalario_solicitud_editar", kwargs={"pk": s.id}),
                "url_eliminar": reverse("sg_transporte_hospitalario_solicitud_eliminar", kwargs={"pk": s.id}),
            }
        )

    return JsonResponse(
        {
            "draw": draw,
            "recordsTotal": total,
            "recordsFiltered": filtered,
            "data": data,
        }
    )


@require_GET
def api_detalle_solicitud(request):
    if not puede_ver_modulo(request.user):
        return JsonResponse({"ok": False, "error": "Acceso denegado."}, status=403)

    solicitud_id = request.GET.get("id")
    if not solicitud_id:
        return JsonResponse({"ok": False, "error": "El parametro id es requerido."}, status=400)

    try:
        solicitud = _obtener_qs_solicitudes_historial(request.user).get(pk=solicitud_id)
    except Solicitud.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Solicitud no encontrada."}, status=404)
    payload = _detalle_solicitud_payload(solicitud)
    return JsonResponse({"ok": True, "data": payload})
@require_POST
def api_solicitud_autorizar(request):
    if not puede_ver_modulo(request.user):
        return JsonResponse({"ok": False, "error": "Acceso denegado."}, status=403)

    try:
        body = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "Datos inválidos."}, status=400)

    solicitud_id = body.get("solicitud_id")
    if not solicitud_id:
        return JsonResponse({"ok": False, "error": "La solicitud es obligatoria."}, status=400)

    with transaction.atomic():
        try:
            solicitud = Solicitud.objects.select_for_update().get(
                pk=solicitud_id,
                activo=True,
                estado=Solicitud.Estado.PENDIENTE,
            )
        except Solicitud.DoesNotExist:
            return JsonResponse({"ok": False, "error": "Solicitud no encontrada o no disponible."}, status=404)

        if ViajeSolicitud.objects.filter(solicitud=solicitud).exists():
            return JsonResponse({"ok": False, "error": "La solicitud ya fue autorizada."}, status=400)

        ViajeSolicitud.objects.create(
            solicitud=solicitud,
            creado_por=request.user,
            viaje=None,
            activo=True,
        )

    return JsonResponse(
        {
            "ok": True,
            "data": {
                "id": solicitud_id,
                "estado": Solicitud.Estado.PENDIENTE,
            },
        }
    )


@require_GET
def api_programacion_viaje(request):
    if not puede_ver_modulo(request.user):
        return JsonResponse({"ok": False, "error": "Acceso denegado."}, status=403)

    seleccionadas = (
        ViajeSolicitud.objects
        .select_related(
            "solicitud__prioridad",
            "solicitud__tipo_solicitud",
            "solicitud__punto_solicitud",
            "solicitud__lugar_destino",
        )
        .filter(activo=True, viaje__isnull=True, solicitud__activo=True)
        .order_by("fecha_asignacion", "id")
    )

    data = [_serializar_viaje_solicitud(item) for item in seleccionadas]
    return JsonResponse({"ok": True, "data": {"seleccionadas": data, "total": len(data)}})


@require_POST
def api_programacion_agregar(request):
    if not puede_ver_modulo(request.user):
        return JsonResponse({"ok": False, "error": "Acceso denegado."}, status=403)

    try:
        body = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "Datos inválidos."}, status=400)

    solicitud_id = body.get("solicitud_id")
    if not solicitud_id:
        return JsonResponse({"ok": False, "error": "La solicitud es obligatoria."}, status=400)
    return JsonResponse({"ok": False, "error": "El flujo actual ya no usa agregación intermedia por esta vía."}, status=400)


@require_POST
def api_programacion_quitar(request):
    if not puede_ver_modulo(request.user):
        return JsonResponse({"ok": False, "error": "Acceso denegado."}, status=403)

    try:
        body = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "Datos inválidos."}, status=400)

    solicitud_id = body.get("solicitud_id")
    if not solicitud_id:
        return JsonResponse({"ok": False, "error": "La solicitud es obligatoria."}, status=400)
    return JsonResponse({"ok": False, "error": "El flujo actual ya no usa desasignación intermedia por esta vía."}, status=400)


@require_POST
def api_programacion_anular(request):
    if not puede_ver_modulo(request.user):
        return JsonResponse({"ok": False, "error": "Acceso denegado."}, status=403)

    try:
        body = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "Datos inválidos."}, status=400)

    solicitud_id = body.get("solicitud_id")
    motivo = (body.get("motivo") or "").strip()
    observacion = (body.get("observacion") or "").strip() or None

    if not solicitud_id:
        return JsonResponse({"ok": False, "error": "La solicitud es obligatoria."}, status=400)
    if not motivo:
        return JsonResponse({"ok": False, "error": "El motivo de anulación es obligatorio."}, status=400)

    with transaction.atomic():
        try:
            solicitud = _obtener_qs_solicitudes_activas(request.user).get(pk=solicitud_id)
        except Solicitud.DoesNotExist:
            return JsonResponse({"ok": False, "error": "Solo se pueden anular solicitudes pendientes."}, status=404)

        solicitud.estado = Solicitud.Estado.ANULADA
        solicitud.activo = False
        solicitud.motivo_anulacion = motivo
        solicitud.observacion_anulacion = observacion
        solicitud.anulada_por = request.user
        solicitud.anulada_en = timezone.now()
        solicitud.save(update_fields=[
            "estado",
            "activo",
            "motivo_anulacion",
            "observacion_anulacion",
            "anulada_por",
            "anulada_en",
        ])

        ViajeSolicitud.objects.filter(solicitud=solicitud, activo=True).update(activo=False)

    return JsonResponse({"ok": True})


@require_POST
def api_programacion_confirmar(request):
    if not puede_ver_modulo(request.user):
        return JsonResponse({"ok": False, "error": "Acceso denegado."}, status=403)

    try:
        body = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "Datos inválidos."}, status=400)

    viaje_solicitud_ids_raw = body.get("viaje_solicitud_ids")
    if not isinstance(viaje_solicitud_ids_raw, list) or not viaje_solicitud_ids_raw:
        return JsonResponse({"ok": False, "error": "Debe seleccionar al menos una solicitud para crear el viaje."}, status=400)
    if len(viaje_solicitud_ids_raw) != len({str(v) for v in viaje_solicitud_ids_raw}):
        return JsonResponse({"ok": False, "error": "Una solicitud no puede agregarse dos veces al mismo viaje."}, status=400)

    programacion_form = ViajeProgramacionForm(data=body)
    if not programacion_form.is_valid():
        errores = {
            campo: [item.get("message", "") for item in mensajes]
            for campo, mensajes in programacion_form.errors.get_json_data().items()
        }
        return JsonResponse(
            {
                "ok": False,
                "error": "Revise los campos de programación del viaje.",
                "errors": errores,
            },
            status=400,
        )

    viaticos_json = programacion_form.cleaned_data.get("viaticos_json") or []
    viaticos_por_id = {}
    if viaticos_json:
        viatico_ids = [item["viatico_id"] for item in viaticos_json]
        viaticos = list(
            Viatico.objects.filter(id__in=viatico_ids, activo=True).order_by("codigo", "nombre")
        )
        viaticos_por_id = {viatico.id: viatico for viatico in viaticos}
        if len(viaticos_por_id) != len(viatico_ids):
            return JsonResponse({"ok": False, "error": "Uno o más viáticos seleccionados no están disponibles."}, status=400)

    try:
        with transaction.atomic():
            try:
                viaje_solicitudes = list(
                    ViajeSolicitud.objects.select_for_update().select_related("solicitud").filter(
                        id__in=viaje_solicitud_ids_raw,
                        activo=True,
                        viaje__isnull=True,
                        solicitud__activo=True,
                    )
                )
            except Exception:
                return JsonResponse({"ok": False, "error": "No se pudieron cargar las solicitudes seleccionadas."}, status=400)

            if len(viaje_solicitudes) != len({str(v) for v in viaje_solicitud_ids_raw}):
                return JsonResponse({"ok": False, "error": "Una o más solicitudes seleccionadas no están disponibles."}, status=400)

            viaje = Viaje.objects.create(
                numero_viaje=_generar_numero_viaje(),
                fecha_programacion=timezone.now(),
                motorista=programacion_form.cleaned_data["motorista"],
                vehiculo=programacion_form.cleaned_data["vehiculo"],
                tipo_viaje=programacion_form.cleaned_data["tipo_viaje"],
                centro_costo=programacion_form.cleaned_data["centro_costo"],
                estado="PROGRAMADA",
                activo=True,
            )

            ViajePersonal.objects.bulk_create(
                [
                    ViajePersonal(
                        viaje=viaje,
                        empleado_id=empleado_id,
                        tipo_participacion="OPERATIVO",
                    )
                    for empleado_id in programacion_form.cleaned_data["personal_operativo_ids"]
                ]
            )

            if viaticos_json:
                for item in viaticos_json:
                    viatico = viaticos_por_id.get(item["viatico_id"])
                    if not viatico:
                        raise ValueError("Uno o más viáticos seleccionados no están disponibles.")
                    ViajeViatico.registrar_asignacion(
                        viaje=viaje,
                        viatico=viatico,
                        creado_por=request.user,
                        observacion=item.get("observacion"),
                    )

            for viaje_solicitud in viaje_solicitudes:
                viaje_solicitud.viaje = viaje
                viaje_solicitud.activo = True
                viaje_solicitud.save(update_fields=["viaje", "activo"])

                solicitud = viaje_solicitud.solicitud
                solicitud.estado = Solicitud.Estado.PROGRAMADA
                solicitud.save(update_fields=["estado"])
    except ValueError as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=400)

    return JsonResponse({"ok": True, "data": {"viaje_id": viaje.id, "numero_viaje": viaje.numero_viaje, "cantidad": len(viaje_solicitudes)}})


@require_POST
def api_ejecucion_guardar(request):
    if not puede_ver_modulo(request.user):
        return JsonResponse({"ok": False, "error": "Acceso denegado."}, status=403)

    try:
        body = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "Datos inválidos."}, status=400)

    viaje_id = body.get("viaje_id")
    modo = (body.get("modo") or "").strip().lower()
    if not viaje_id:
        return JsonResponse({"ok": False, "error": "El viaje es obligatorio."}, status=400)
    if modo not in {"ver", "editar", "salida", "entrada"}:
        return JsonResponse({"ok": False, "error": "El modo de ejecución no es válido."}, status=400)

    try:
        viaje = Viaje.objects.select_related("ejecucion_viaje").get(pk=viaje_id, activo=True)
    except Viaje.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Viaje no encontrado."}, status=404)

    form = EjecucionViajeForm(data=body)
    if not form.is_valid():
        errores = {
            campo: [item.get("message", "") for item in mensajes]
            for campo, mensajes in form.errors.get_json_data().items()
        }
        return JsonResponse(
            {
                "ok": False,
                "error": "Revise los campos de ejecución.",
                "errors": errores,
            },
            status=400,
        )

    cleaned = form.cleaned_data
    ejecucion = getattr(viaje, "ejecucion_viaje", None)
    if ejecucion is None:
        ejecucion = EjecucionViaje(viaje=viaje)

    fecha_salida = cleaned.get("fecha_salida")
    kilometraje_salida = cleaned.get("kilometraje_salida")
    precio_litro_salida = cleaned.get("precio_litro_salida")
    litros_cargados_salida = cleaned.get("litros_cargados_salida")
    total_combustible = cleaned.get("total_combustible")
    observaciones_salida = (cleaned.get("observaciones_salida") or "").strip() or None
    fecha_retorno = cleaned.get("fecha_retorno")
    kilometraje_retorno = cleaned.get("kilometraje_retorno")
    observaciones_retorno = (cleaned.get("observaciones_retorno") or "").strip() or None

    if modo == "salida":
        if ejecucion.fecha_salida:
            return JsonResponse({"ok": False, "error": "La salida ya fue registrada."}, status=400)
        required_values = {
            "fecha_salida": fecha_salida,
            "kilometraje_salida": kilometraje_salida,
            "precio_litro_salida": precio_litro_salida,
            "litros_cargados_salida": litros_cargados_salida,
        }
        faltantes = [campo for campo, valor in required_values.items() if valor in (None, "")]
        if faltantes:
            return JsonResponse({"ok": False, "error": "Complete los campos obligatorios de salida."}, status=400)
        ejecucion.fecha_salida = fecha_salida
        ejecucion.kilometraje_salida = kilometraje_salida
        ejecucion.precio_litro_salida = precio_litro_salida
        ejecucion.litros_cargados_salida = litros_cargados_salida
        ejecucion.total_combustible = total_combustible
        ejecucion.combustible_salida = total_combustible
        ejecucion.observaciones_salida = observaciones_salida
        if not ejecucion.observaciones:
            ejecucion.observaciones = observaciones_salida
        viaje.estado = "EN_EJECUCION"

    elif modo == "entrada":
        if not ejecucion.fecha_salida:
            return JsonResponse({"ok": False, "error": "Primero debe registrar la salida del viaje."}, status=400)
        if ejecucion.fecha_retorno:
            return JsonResponse({"ok": False, "error": "La entrada ya fue registrada."}, status=400)
        if fecha_retorno is None or kilometraje_retorno is None:
            return JsonResponse({"ok": False, "error": "Complete los campos obligatorios de entrada."}, status=400)
        if ejecucion.kilometraje_salida is not None and kilometraje_retorno < ejecucion.kilometraje_salida:
            return JsonResponse({"ok": False, "error": "El kilometraje de retorno debe ser mayor o igual al de salida."}, status=400)
        ejecucion.fecha_retorno = fecha_retorno
        ejecucion.kilometraje_retorno = kilometraje_retorno
        ejecucion.observaciones_retorno = observaciones_retorno
        if not ejecucion.observaciones and observaciones_retorno:
            ejecucion.observaciones = observaciones_retorno
        viaje.estado = "FINALIZADO"

    else:
        if fecha_salida is not None:
            ejecucion.fecha_salida = fecha_salida
            viaje.estado = "EN_EJECUCION" if not ejecucion.fecha_retorno else "FINALIZADO"
        if kilometraje_salida is not None:
            ejecucion.kilometraje_salida = kilometraje_salida
        if precio_litro_salida is not None:
            ejecucion.precio_litro_salida = precio_litro_salida
        if litros_cargados_salida is not None:
            ejecucion.litros_cargados_salida = litros_cargados_salida
        if total_combustible is not None:
            ejecucion.total_combustible = total_combustible
            ejecucion.combustible_salida = total_combustible
        if observaciones_salida is not None:
            ejecucion.observaciones_salida = observaciones_salida
        if fecha_retorno is not None:
            ejecucion.fecha_retorno = fecha_retorno
            viaje.estado = "FINALIZADO" if ejecucion.fecha_salida else viaje.estado
        if kilometraje_retorno is not None:
            if ejecucion.kilometraje_salida is not None and kilometraje_retorno < ejecucion.kilometraje_salida:
                return JsonResponse({"ok": False, "error": "El kilometraje de retorno debe ser mayor o igual al de salida."}, status=400)
            ejecucion.kilometraje_retorno = kilometraje_retorno
        if observaciones_retorno is not None:
            ejecucion.observaciones_retorno = observaciones_retorno

    with transaction.atomic():
        viaje.save(update_fields=["estado", "updated_at"])
        ejecucion.save()

    estado_ejecucion = _estado_ejecucion_viaje(viaje)
    return JsonResponse(
        {
            "ok": True,
            "data": {
                "viaje_id": viaje.id,
                "numero_viaje": viaje.numero_viaje,
                "estado_ejecucion": estado_ejecucion,
                "etapa": {
                    "PROGRAMADO": "Programado",
                    "EN_EJECUCION": "En ejecución",
                    "FINALIZADO": "Finalizado",
                }.get(estado_ejecucion, "Programado"),
            },
        }
    )


class SolicitudListView(LoginRequiredMixin, ListView):
    model = Solicitud
    template_name = "sg_transporte_hospitalario/solicitud_list.html"
    context_object_name = "solicitudes_activas"

    def dispatch(self, request, *args, **kwargs):
        if not puede_ver_modulo(request.user):
            return redirect("acceso_denegado")
        return redirect(f"{reverse('sg_transporte_hospitalario_dashboard')}?tab=solicitud")

    def get_queryset(self):
        return _obtener_qs_solicitudes_propias(self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(construir_contexto_dashboard(self.request.user))
        context["tabs"] = TAB_SECTIONS
        context["active_tab"] = "solicitud"
        context["usuario_empleado_id"] = getattr(getattr(self.request.user, "empleado", None), "id", None)
        context["estados"] = Solicitud.Estado.choices
        return context


class SolicitudFormView(LoginRequiredMixin, TemplateView):
    template_name = "sg_transporte_hospitalario/solicitud_form.html"
    modo = "crear"

    def dispatch(self, request, *args, **kwargs):
        if not puede_ver_modulo(request.user):
            return redirect("acceso_denegado")
        return super().dispatch(request, *args, **kwargs)

    def _get_solicitud(self, solo_propias=None):
        pk = self.kwargs.get("pk")
        if not pk:
            return None
        if solo_propias is None:
            solo_propias = self.modo == "editar"
        qs = _obtener_qs_solicitudes_propias(self.request.user) if solo_propias else _obtener_qs_solicitudes_activas(self.request.user)
        return get_object_or_404(qs, pk=pk)

    def get(self, request, *args, **kwargs):
        solicitud = self._get_solicitud()
        form = SolicitudForm(instance=solicitud)
        if self.modo == "ver":
            for field in form.fields.values():
                field.disabled = True

        pacientes_payload = []
        personal_payload = []
        if solicitud:
            solicitud_pacientes = solicitud.solicitud_pacientes.all()
            pacientes_payload = [
                {
                    "paciente_id": item.paciente_id,
                    "ingreso_id": item.ingreso_id,
                    "nombre": " ".join(
                        x for x in [
                            getattr(item.paciente, "primer_nombre", None),
                            getattr(item.paciente, "segundo_nombre", None),
                            getattr(item.paciente, "primer_apellido", None),
                            getattr(item.paciente, "segundo_apellido", None),
                        ] if x
                    ) or "-",
                    "expediente": str(getattr(item.paciente, "expediente_numero", "-") or "-"),
                    "ingreso": f"ING-{item.ingreso_id}" if item.ingreso_id else "-",
                    "sala": item.ingreso.sala.nombre_sala if item.ingreso and item.ingreso.sala else "-",
                }
                for item in solicitud_pacientes
            ]

            solicitud_personal = solicitud.solicitud_personal.all()
            personal_payload = [
                {
                    "empleado_id": item.empleado_id,
                    "nombre": _nombre_empleado(item.empleado),
                    "cargo": _cargo_empleado(item.empleado),
                    "observacion": item.observacion or "",
                }
                for item in solicitud_personal
            ]

        context = self.get_context_data(
            form=form,
            solicitud=solicitud,
            pacientes_payload=json.dumps(pacientes_payload),
            personal_payload=json.dumps(personal_payload),
        )
        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        if not puede_gestionar_modulo(request.user):
            return JsonResponse({"ok": False, "error": "Acceso denegado."}, status=403)

        solicitud = self._get_solicitud()
        if solicitud and not solicitud.puede_editar:
            mostrar_mensaje(request, "warning", "La solicitud ya está asociada a una autorización y no puede editarse.")
            return redirect("sg_transporte_hospitalario_solicitud_list")

        form = SolicitudForm(request.POST, instance=solicitud)
        pacientes_payload = _parse_payload_json(request.POST.get("pacientes_json"))
        personal_payload = _parse_payload_json(request.POST.get("empleados_json"))

        empleado_usuario = getattr(request.user, "empleado", None)
        if empleado_usuario is None:
            form.add_error(None, "El usuario actual no tiene empleado asociado.")

        if form.is_valid() and empleado_usuario is not None:
            pacientes_unicos = _validar_reglas_pacientes(form, pacientes_payload, solicitud.id if solicitud else None)

        if form.is_valid() and empleado_usuario is not None:
            with transaction.atomic():
                solicitud_guardada = form.save(commit=False)
                if not solicitud_guardada.pk:
                    solicitud_guardada.numero_solicitud = _generar_numero_solicitud()
                    solicitud_guardada.solicitante_empleado = empleado_usuario
                    solicitud_guardada.estado = Solicitud.Estado.PENDIENTE
                    solicitud_guardada.activo = True
                solicitud_guardada.save()

                SolicitudPaciente.objects.filter(solicitud=solicitud_guardada).delete()
                SolicitudPersonal.objects.filter(solicitud=solicitud_guardada).delete()

                for item in pacientes_unicos:
                    paciente_id = item.get("paciente_id")
                    ingreso_id = item.get("ingreso_id")
                    SolicitudPaciente.objects.create(
                        solicitud=solicitud_guardada,
                        paciente_id=paciente_id or None,
                        ingreso_id=ingreso_id or None,
                    )

                empleados_ids_guardados = set()
                for item in personal_payload:
                    empleado_id = item.get("empleado_id")
                    if not empleado_id or empleado_id in empleados_ids_guardados:
                        continue
                    empleados_ids_guardados.add(empleado_id)
                    SolicitudPersonal.objects.create(
                        solicitud=solicitud_guardada,
                        empleado_id=empleado_id,
                        observacion=item.get("observacion") or None,
                    )

            accion = "actualizada" if solicitud else "creada"
            mostrar_mensaje(request, "success", f"Solicitud {solicitud_guardada.numero_solicitud} {accion} correctamente.")
            return redirect("sg_transporte_hospitalario_solicitud_list")

        context = self.get_context_data(
            form=form,
            solicitud=solicitud,
            pacientes_payload=json.dumps(pacientes_payload),
            personal_payload=json.dumps(personal_payload),
        )
        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(construir_contexto_dashboard(self.request.user))
        context["tabs"] = TAB_SECTIONS
        context["active_tab"] = "solicitud"
        context["modo"] = self.modo
        context["titulo_form"] = {
            "crear": "Nueva solicitud",
            "editar": "Editar solicitud",
            "ver": "Detalle de solicitud",
        }.get(self.modo, "Solicitud")
        context["form_action"] = self.request.path
        context["solicitud"] = kwargs.get("solicitud")
        context["form"] = kwargs.get("form")
        context["pacientes_payload"] = kwargs.get("pacientes_payload", "[]")
        context["personal_payload"] = kwargs.get("personal_payload", "[]")
        context["tipo_pacientes_id"] = (
            TipoSolicitud.objects
            .filter(codigo__iexact="PACIENTES", activo=True)
            .values_list("id", flat=True)
            .first()
            or ""
        )
        context.update(_metadata_opciones_formulario(context["form"], context["solicitud"], self.request.user))
        return context


class SolicitudCreateView(SolicitudFormView):
    modo = "crear"

    def post(self, request, *args, **kwargs):
        if not puede_gestionar_modulo(request.user):
            return JsonResponse({"ok": False, "error": "Acceso denegado."}, status=403)

        form = SolicitudForm(request.POST)
        pacientes_payload = _parse_payload_json(request.POST.get("pacientes_json"))
        personal_payload = _parse_payload_json(request.POST.get("empleados_json"))

        empleado_usuario = getattr(request.user, "empleado", None)
        if empleado_usuario is None:
            form.add_error(None, "El usuario actual no tiene empleado asociado.")

        if form.is_valid() and empleado_usuario is not None:
            pacientes_unicos = _validar_reglas_pacientes(form, pacientes_payload)

        if form.is_valid() and empleado_usuario is not None:
            with transaction.atomic():
                solicitud = form.save(commit=False)
                solicitud.numero_solicitud = _generar_numero_solicitud()
                solicitud.solicitante_empleado = empleado_usuario
                solicitud.estado = Solicitud.Estado.PENDIENTE
                solicitud.activo = True
                solicitud.save()

                for item in pacientes_unicos:
                    paciente_id = item.get("paciente_id")
                    ingreso_id = item.get("ingreso_id")

                    SolicitudPaciente.objects.create(
                        solicitud=solicitud,
                        paciente_id=paciente_id or None,
                        ingreso_id=ingreso_id or None,
                    )

                empleados_ids_guardados = set()
                for item in personal_payload:
                    empleado_id = item.get("empleado_id")
                    if not empleado_id or empleado_id in empleados_ids_guardados:
                        continue
                    empleados_ids_guardados.add(empleado_id)
                    SolicitudPersonal.objects.create(
                        solicitud=solicitud,
                        empleado_id=empleado_id,
                        observacion=item.get("observacion") or None,
                    )

            mostrar_mensaje(request, "success", f"Solicitud {solicitud.numero_solicitud} creada correctamente.")
            return redirect("sg_transporte_hospitalario_solicitud_list")

        context = self.get_context_data(
            form=form,
            solicitud=None,
            pacientes_payload=json.dumps(pacientes_payload),
            personal_payload=json.dumps(personal_payload),
        )
        return self.render_to_response(context)


class SolicitudUpdateView(SolicitudFormView):
    modo = "editar"

    def post(self, request, *args, **kwargs):
        if not puede_gestionar_modulo(request.user):
            return JsonResponse({"ok": False, "error": "Acceso denegado."}, status=403)

        solicitud = self._get_solicitud()
        if solicitud is None:
            return JsonResponse({"ok": False, "error": "Solicitud no encontrada."}, status=404)

        if not solicitud.puede_editar:
            mostrar_mensaje(request, "warning", "La solicitud ya está asociada a una autorización y no puede editarse.")
            return redirect("sg_transporte_hospitalario_solicitud_list")

        form = SolicitudForm(request.POST, instance=solicitud)
        pacientes_payload = _parse_payload_json(request.POST.get("pacientes_json"))
        personal_payload = _parse_payload_json(request.POST.get("empleados_json"))

        if form.is_valid():
            pacientes_unicos = _validar_reglas_pacientes(form, pacientes_payload, solicitud.id if solicitud else None)

        if form.is_valid():
            with transaction.atomic():
                solicitud_actualizada = form.save()

                SolicitudPaciente.objects.filter(solicitud=solicitud_actualizada).delete()
                SolicitudPersonal.objects.filter(solicitud=solicitud_actualizada).delete()

                for item in pacientes_unicos:
                    paciente_id = item.get("paciente_id")
                    ingreso_id = item.get("ingreso_id")

                    SolicitudPaciente.objects.create(
                        solicitud=solicitud_actualizada,
                        paciente_id=paciente_id or None,
                        ingreso_id=ingreso_id or None,
                    )

                empleados_ids_guardados = set()
                for item in personal_payload:
                    empleado_id = item.get("empleado_id")
                    if not empleado_id or empleado_id in empleados_ids_guardados:
                        continue
                    empleados_ids_guardados.add(empleado_id)
                    SolicitudPersonal.objects.create(
                        solicitud=solicitud_actualizada,
                        empleado_id=empleado_id,
                        observacion=item.get("observacion") or None,
                    )

            mostrar_mensaje(request, "success", f"Solicitud {solicitud_actualizada.numero_solicitud} actualizada correctamente.")
            return redirect("sg_transporte_hospitalario_solicitud_list")

        context = self.get_context_data(
            form=form,
            solicitud=solicitud,
            pacientes_payload=json.dumps(pacientes_payload),
            personal_payload=json.dumps(personal_payload),
        )
        return self.render_to_response(context)


class SolicitudDeleteView(LoginRequiredMixin, View):
    http_method_names = ["post"]

    def post(self, request, *args, **kwargs):
        if not puede_gestionar_modulo(request.user):
            return JsonResponse({"ok": False, "error": "Acceso denegado."}, status=403)

        pk = kwargs.get("pk")
        if not pk:
            return JsonResponse({"ok": False, "error": "Solicitud no encontrada."}, status=404)

        solicitud = get_object_or_404(_obtener_qs_solicitudes_propias(request.user), pk=pk)
        if not solicitud.puede_editar:
            mostrar_mensaje(request, "warning", "La solicitud ya fue autorizada o no puede eliminarse.")
            return redirect(f"{reverse('sg_transporte_hospitalario_dashboard')}?tab=solicitud")

        numero_solicitud = solicitud.numero_solicitud
        with transaction.atomic():
            SolicitudPaciente.objects.filter(solicitud=solicitud).delete()
            SolicitudPersonal.objects.filter(solicitud=solicitud).delete()
            solicitud.delete()

        mostrar_mensaje(request, "success", f"Solicitud {numero_solicitud} eliminada correctamente.")
        return redirect(f"{reverse('sg_transporte_hospitalario_dashboard')}?tab=solicitud")


@method_decorator(xframe_options_exempt, name="dispatch")
class SolicitudDetailView(SolicitudFormView):
    modo = "ver"

    def post(self, request, *args, **kwargs):
        return redirect("sg_transporte_hospitalario_solicitud_ver", pk=kwargs.get("pk"))
