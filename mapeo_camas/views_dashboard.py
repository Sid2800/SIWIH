# 2026-05-29: extraído de mapeo_camas/views.py en refactor E (split)
"""Dashboard operativo de KPIs y monitoreo en tiempo real."""

from datetime import timedelta
from collections import defaultdict

from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import OuterRef, Subquery
from django.http import HttpResponse
from django.shortcuts import redirect
from django.utils import timezone
from django.views.decorators.http import require_GET
from django.views.generic import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from core.utils.utilidades_fechas import hora_local_iso
from core.services.reporte.EXCEL.reporte_service_excel import ServiceExcel
from servicio.models import Cama

from mapeo_camas.models import (
    AsignacionCamaPaciente,
    DetalleMapeoCama,
    HistorialEstadoCama,
    MovimientoCama,
)

from ._helpers import (
    _dashboard_error,
    _dashboard_granularidad,
    _dashboard_parse_range,
    _dashboard_response,
    _nombre_cama,
    _nombre_usuario,
    _paciente_payload,
    _rango_meta,
    _snapshot_estado_camas,
)
from ._permisos import _tiene_permiso_dashboard
from ._sesion import _obtener_sesion_mapeo_activa


__all__ = [
    "DashboardMapeoCamasView",
    "dashboard_kpis",
    "dashboard_ocupacion_hora",
    "dashboard_export_ocupacion_excel",
    "dashboard_ultimos_ingresos",
    "dashboard_ultimos_movimientos",
]


# [2026-06-26] En el reporte Excel de ocupacion se contabiliza solo estado OCUPADA
# para alinear con lectura operativa del historial por cama.
ESTADOS_OCUPADOS_REPORTE = {"OCUPADA"}


ESTADOS_KPI_DASHBOARD = {
    "ocupadas": ("OCUPADA", "PRE_ALTA"),
    "disponibles": ("VACIA", "LIBRE"),
    "fuera_servicio": ("FUERA_SERVICIO", "MANTENIMIENTO"),
    "consulta_externa": ("CONSULTA_EXTERNA",),
    "pre_altas": ("PRE_ALTA",),
}


def _iterar_dias_intervalo(inicio, fin):
    """2026-06-16: Devuelve días tocados por el intervalo [inicio, fin)."""
    if not inicio or not fin or fin <= inicio:
        return
    # [2026-06-26] Normaliza a hora local para evitar desfases de día por UTC en el conteo.
    inicio_local = timezone.localtime(inicio)
    fin_local = timezone.localtime(fin)
    cursor = inicio_local.date()
    ultimo = (fin_local - timedelta(seconds=1)).date()
    while cursor <= ultimo:
        yield cursor
        cursor = cursor + timedelta(days=1)


def _snapshot_estado_por_cama_desde_historial(corte):
    """2026-06-26: Snapshot historico estricto por cama usando solo HistorialEstadoCama."""
    return {
        r["cama_id"]: (r["estado_nuevo__codigo"] or "VACIA")
        for r in HistorialEstadoCama.objects
        .filter(id=Subquery(
            HistorialEstadoCama.objects
            .filter(cama_id=OuterRef("cama_id"), fecha_hora__lte=corte)
            .order_by("-fecha_hora", "-id")
            .values("id")[:1]
        ))
        .values("cama_id", "estado_nuevo__codigo")
    }


def _calcular_metricas_ocupacion_por_cama(desde, hasta):
    """2026-06-16: Calcula horas/días ocupados e ingresos por cama para el rango."""
    # [2026-06-26] Para reporte historico de ocupacion, usar solo historial evita arrastre
    # de estado actual desde AsignacionCamaPaciente y reduce inflado de horas/dias.
    snapshot_desde = _snapshot_estado_por_cama_desde_historial(desde)

    eventos = list(
        HistorialEstadoCama.objects
        .filter(fecha_hora__gte=desde, fecha_hora__lte=hasta)
        .values("cama_id", "fecha_hora", "estado_nuevo__codigo")
        .order_by("cama_id", "fecha_hora", "id")
    )

    # [2026-06-26] Para mantener coherencia con horas/dias del reporte, ingresos/pacientes
    # también se calculan desde historial (evento de ocupación) en lugar de asignaciones.
    ingresos_rows = list(
        HistorialEstadoCama.objects
        .filter(
            fecha_hora__gte=desde,
            fecha_hora__lte=hasta,
            ingreso_id__isnull=False,
            estado_nuevo__codigo="OCUPADA",
        )
        .values("cama_id", "ingreso_id")
    )

    camas_ids = set(snapshot_desde.keys())
    camas_ids.update(e["cama_id"] for e in eventos)
    camas_ids.update(r["cama_id"] for r in ingresos_rows)

    eventos_por_cama = defaultdict(list)
    for ev in eventos:
        eventos_por_cama[ev["cama_id"]].append(ev)

    ingresos_por_cama = defaultdict(set)
    for row in ingresos_rows:
        ingresos_por_cama[row["cama_id"]].add(row["ingreso_id"])

    metricas = {}
    for cama_id in camas_ids:
        estado_inicial = snapshot_desde.get(cama_id)
        ocupada = estado_inicial in ESTADOS_OCUPADOS_REPORTE
        ultimo_punto = desde
        segundos_ocupada = 0.0
        dias_ocupada = set()

        for ev in eventos_por_cama.get(cama_id, []):
            fecha_evento = ev["fecha_hora"]
            if ocupada and fecha_evento > ultimo_punto:
                segundos_ocupada += (fecha_evento - ultimo_punto).total_seconds()
                for d in _iterar_dias_intervalo(ultimo_punto, fecha_evento):
                    dias_ocupada.add(d)

            ocupada = (ev.get("estado_nuevo__codigo") in ESTADOS_OCUPADOS_REPORTE)
            ultimo_punto = fecha_evento

        if ocupada and hasta > ultimo_punto:
            segundos_ocupada += (hasta - ultimo_punto).total_seconds()
            for d in _iterar_dias_intervalo(ultimo_punto, hasta):
                dias_ocupada.add(d)

        metricas[cama_id] = {
            "horas_ocupada": round(segundos_ocupada / 3600.0, 2),
            "dias_ocupada": len(dias_ocupada),
            "ingresos_total": len(ingresos_por_cama.get(cama_id, set())),
        }

    return metricas


def _formatear_kpi_horas_porcentaje(horas_activas, porcentaje):
    return f"{horas_activas:.1f} h · {porcentaje:.1f}%"


def _calcular_horas_por_estado_dashboard(desde, hasta):
    """2026-07-10: Calcula horas activas por estado para las cards KPI del dashboard."""
    total_horas_ventana = max((hasta - desde).total_seconds() / 3600.0, 0.0)
    snapshot_desde = _snapshot_estado_por_cama_desde_historial(desde)

    eventos = list(
        HistorialEstadoCama.objects
        .filter(fecha_hora__gte=desde, fecha_hora__lte=hasta)
        .values("cama_id", "fecha_hora", "estado_nuevo__codigo")
        .order_by("cama_id", "fecha_hora", "id")
    )

    camas_ids = set(snapshot_desde.keys())
    camas_ids.update(ev["cama_id"] for ev in eventos)
    total_camas_periodo = len(camas_ids)
    total_horas_ventana_camas = total_horas_ventana * total_camas_periodo

    eventos_por_cama = defaultdict(list)
    for ev in eventos:
        eventos_por_cama[ev["cama_id"]].append(ev)

    segundos_por_estado = defaultdict(float)
    detalle_por_cama = {}

    for cama_id in camas_ids:
        # [2026-07-10] Si no existe estado previo en historial, se asume VACIA para no inflar horas.
        estado_actual = snapshot_desde.get(cama_id) or "VACIA"
        ultimo_punto = desde
        segundos_por_estado_cama = defaultdict(float)

        for ev in eventos_por_cama.get(cama_id, []):
            fecha_evento = ev["fecha_hora"]
            if fecha_evento > ultimo_punto:
                duracion = (fecha_evento - ultimo_punto).total_seconds()
                segundos_por_estado[estado_actual] += duracion
                segundos_por_estado_cama[estado_actual] += duracion

            estado_actual = ev.get("estado_nuevo__codigo") or "VACIA"
            ultimo_punto = fecha_evento

        if hasta > ultimo_punto:
            duracion = (hasta - ultimo_punto).total_seconds()
            segundos_por_estado[estado_actual] += duracion
            segundos_por_estado_cama[estado_actual] += duracion

        detalle_por_cama[cama_id] = {
            codigo: round(segundos / 3600.0, 2)
            for codigo, segundos in segundos_por_estado_cama.items()
        }

    horas_por_estado = {
        codigo: round(segundos / 3600.0, 2)
        for codigo, segundos in segundos_por_estado.items()
    }

    porcentajes_por_kpi = {}
    horas_por_kpi = {}
    textos_por_kpi = {}
    for campo, codigos in ESTADOS_KPI_DASHBOARD.items():
        horas_activas = round(sum(horas_por_estado.get(codigo, 0.0) for codigo in codigos), 2)
        porcentaje = 0.0
        if total_horas_ventana_camas > 0:
            porcentaje = round((horas_activas / total_horas_ventana_camas) * 100.0, 1)
        porcentaje = max(0.0, min(porcentaje, 100.0))
        horas_por_kpi[campo] = horas_activas
        porcentajes_por_kpi[campo] = porcentaje
        textos_por_kpi[campo] = _formatear_kpi_horas_porcentaje(horas_activas, porcentaje)

    return {
        "horas_ventana": round(total_horas_ventana, 2),
        "total_camas_periodo": total_camas_periodo,
        "horas_ventana_camas": round(total_horas_ventana_camas, 2),
        "horas_por_estado": horas_por_estado,
        "horas_por_kpi": horas_por_kpi,
        "porcentajes_por_kpi": porcentajes_por_kpi,
        "textos_por_kpi": textos_por_kpi,
        "detalle_por_cama": detalle_por_cama,
    }


def _construir_filas_reporte_ocupacion(desde, hasta):
    """2026-06-16: Agrupa métricas por servicio/sala/cama para exportación Excel."""
    metricas_por_cama = _calcular_metricas_ocupacion_por_cama(desde, hasta)
    if not metricas_por_cama:
        return []

    # [2026-07-10] Reutiliza el detalle por cama del KPI para alinear estados del Excel.
    detalle_horas_estado_por_cama = _calcular_horas_por_estado_dashboard(desde, hasta).get("detalle_por_cama", {})
    codigos_pre_alta = ESTADOS_KPI_DASHBOARD.get("pre_altas", ("PRE_ALTA",))
    codigos_vacia = ESTADOS_KPI_DASHBOARD.get("disponibles", ("VACIA",))
    codigos_fuera_servicio = ESTADOS_KPI_DASHBOARD.get("fuera_servicio", ("FUERA_SERVICIO",))
    codigos_consulta_externa = ESTADOS_KPI_DASHBOARD.get("consulta_externa", ("CONSULTA_EXTERNA",))

    total_dias_ventana = len(list(_iterar_dias_intervalo(desde, hasta)))
    total_dias_ventana = max(total_dias_ventana, 1)
    total_horas_ventana = max((hasta - desde).total_seconds() / 3600.0, 0.0)
    camas = list(
        Cama.objects
        .filter(pk__in=metricas_por_cama.keys())
        .select_related("sala__servicio")
        .order_by("sala__servicio__nombre_servicio", "sala__nombre_sala", "numero_cama")
    )

    agrupado = defaultdict(lambda: defaultdict(list))
    for cama in camas:
        servicio_nombre = getattr(getattr(cama, "sala", None), "servicio", None)
        servicio_nombre = getattr(servicio_nombre, "nombre_servicio", "SIN_SERVICIO") or "SIN_SERVICIO"
        sala_nombre = getattr(getattr(cama, "sala", None), "nombre_sala", "SIN_SALA") or "SIN_SALA"
        agrupado[servicio_nombre][sala_nombre].append(cama)

    filas = []
    for servicio_nombre in sorted(agrupado.keys()):
        servicio_horas = 0.0
        servicio_horas_pre_alta = 0.0
        servicio_horas_vacia = 0.0
        servicio_horas_fuera_servicio = 0.0
        servicio_horas_consulta_externa = 0.0
        servicio_dias = 0
        servicio_ingresos = 0
        servicio_camas = 0

        for sala_nombre in sorted(agrupado[servicio_nombre].keys()):
            camas_sala = agrupado[servicio_nombre][sala_nombre]
            sala_horas = 0.0
            sala_horas_pre_alta = 0.0
            sala_horas_vacia = 0.0
            sala_horas_fuera_servicio = 0.0
            sala_horas_consulta_externa = 0.0
            sala_dias = 0
            sala_ingresos = 0

            for cama in camas_sala:
                met = metricas_por_cama.get(cama.pk, {})
                detalle_estado = detalle_horas_estado_por_cama.get(cama.pk, {})
                horas = float(met.get("horas_ocupada") or 0.0)
                horas_pre_alta = round(sum(float(detalle_estado.get(codigo) or 0.0) for codigo in codigos_pre_alta), 2)
                horas_vacia = round(sum(float(detalle_estado.get(codigo) or 0.0) for codigo in codigos_vacia), 2)
                horas_fuera_servicio = round(sum(float(detalle_estado.get(codigo) or 0.0) for codigo in codigos_fuera_servicio), 2)
                horas_consulta_externa = round(sum(float(detalle_estado.get(codigo) or 0.0) for codigo in codigos_consulta_externa), 2)
                dias = int(met.get("dias_ocupada") or 0)
                ingresos = int(met.get("ingresos_total") or 0)
                pct = round((horas / total_horas_ventana) * 100, 2) if total_horas_ventana else 0.0

                filas.append({
                    "nivel": "CAMA",
                    "servicio": servicio_nombre,
                    "sala": sala_nombre,
                    "cama": str(cama.numero_cama),
                    "n_camas": 1,
                    "horas_ocupada": horas,
                    "horas_pre_alta": horas_pre_alta,
                    "horas_vacia": horas_vacia,
                    "horas_fuera_servicio": horas_fuera_servicio,
                    "horas_consulta_externa": horas_consulta_externa,
                    "horas_ventana": round(total_horas_ventana, 2),
                    "ocupacion_pct": pct,
                    "dias_ocupada": dias,
                    "dias_ventana": total_dias_ventana,
                    "dias_ventana_pct": round((dias / total_dias_ventana) * 100, 2) if total_dias_ventana else 0.0,
                    "ingresos_total": ingresos,
                })

                sala_horas += horas
                sala_horas_pre_alta += horas_pre_alta
                sala_horas_vacia += horas_vacia
                sala_horas_fuera_servicio += horas_fuera_servicio
                sala_horas_consulta_externa += horas_consulta_externa
                sala_dias += dias
                sala_ingresos += ingresos

            sala_camas = len(camas_sala)
            sala_horas_teoricas = total_horas_ventana * sala_camas
            sala_pct = round((sala_horas / sala_horas_teoricas) * 100, 2) if sala_horas_teoricas else 0.0
            filas.append({
                "nivel": "TOTAL_SALA",
                "servicio": servicio_nombre,
                "sala": sala_nombre,
                "cama": "TOTAL SALA",
                "n_camas": sala_camas,
                "horas_ocupada": round(sala_horas, 2),
                "horas_pre_alta": round(sala_horas_pre_alta, 2),
                "horas_vacia": round(sala_horas_vacia, 2),
                "horas_fuera_servicio": round(sala_horas_fuera_servicio, 2),
                "horas_consulta_externa": round(sala_horas_consulta_externa, 2),
                "horas_ventana": round(sala_horas_teoricas, 2),
                "ocupacion_pct": sala_pct,
                "dias_ocupada": sala_dias,
                "dias_ventana": total_dias_ventana * sala_camas,
                "dias_ventana_pct": round((sala_dias / (total_dias_ventana * sala_camas)) * 100, 2) if sala_camas else 0.0,
                "ingresos_total": sala_ingresos,
            })

            servicio_horas += sala_horas
            servicio_horas_pre_alta += sala_horas_pre_alta
            servicio_horas_vacia += sala_horas_vacia
            servicio_horas_fuera_servicio += sala_horas_fuera_servicio
            servicio_horas_consulta_externa += sala_horas_consulta_externa
            servicio_dias += sala_dias
            servicio_ingresos += sala_ingresos
            servicio_camas += sala_camas

        servicio_horas_teoricas = total_horas_ventana * servicio_camas
        servicio_pct = round((servicio_horas / servicio_horas_teoricas) * 100, 2) if servicio_horas_teoricas else 0.0
        filas.append({
            "nivel": "TOTAL_SERVICIO",
            "servicio": servicio_nombre,
            "sala": "",
            "cama": "TOTAL SERVICIO",
            "n_camas": servicio_camas,
            "horas_ocupada": round(servicio_horas, 2),
            "horas_pre_alta": round(servicio_horas_pre_alta, 2),
            "horas_vacia": round(servicio_horas_vacia, 2),
            "horas_fuera_servicio": round(servicio_horas_fuera_servicio, 2),
            "horas_consulta_externa": round(servicio_horas_consulta_externa, 2),
            "horas_ventana": round(servicio_horas_teoricas, 2),
            "ocupacion_pct": servicio_pct,
            "dias_ocupada": servicio_dias,
            "dias_ventana": total_dias_ventana * servicio_camas,
            "dias_ventana_pct": round((servicio_dias / (total_dias_ventana * servicio_camas)) * 100, 2) if servicio_camas else 0.0,
            "ingresos_total": servicio_ingresos,
        })

    if filas:
        total_horas = round(sum(float(f.get("horas_ocupada") or 0.0) for f in filas if f.get("nivel") == "TOTAL_SERVICIO"), 2)
        total_horas_pre_alta = round(sum(float(f.get("horas_pre_alta") or 0.0) for f in filas if f.get("nivel") == "TOTAL_SERVICIO"), 2)
        total_horas_vacia = round(sum(float(f.get("horas_vacia") or 0.0) for f in filas if f.get("nivel") == "TOTAL_SERVICIO"), 2)
        total_horas_fuera_servicio = round(sum(float(f.get("horas_fuera_servicio") or 0.0) for f in filas if f.get("nivel") == "TOTAL_SERVICIO"), 2)
        total_horas_consulta_externa = round(sum(float(f.get("horas_consulta_externa") or 0.0) for f in filas if f.get("nivel") == "TOTAL_SERVICIO"), 2)
        total_camas = int(sum(int(f.get("n_camas") or 0) for f in filas if f.get("nivel") == "TOTAL_SERVICIO"))
        total_ingresos = int(sum(int(f.get("ingresos_total") or 0) for f in filas if f.get("nivel") == "TOTAL_SERVICIO"))
        total_dias = int(sum(int(f.get("dias_ocupada") or 0) for f in filas if f.get("nivel") == "TOTAL_SERVICIO"))
        horas_ventana_total = total_horas_ventana * total_camas
        filas.append({
            "nivel": "SUPER_TOTAL",
            "servicio": "SUPER TOTAL",
            "sala": "",
            "cama": "SUPER TOTAL",
            "n_camas": total_camas,
            "horas_ocupada": total_horas,
            "horas_pre_alta": total_horas_pre_alta,
            "horas_vacia": total_horas_vacia,
            "horas_fuera_servicio": total_horas_fuera_servicio,
            "horas_consulta_externa": total_horas_consulta_externa,
            "horas_ventana": round(horas_ventana_total, 2),
            "ocupacion_pct": round((total_horas / horas_ventana_total) * 100, 2) if horas_ventana_total else 0.0,
            "dias_ocupada": total_dias,
            "dias_ventana": total_dias_ventana * total_camas,
            "dias_ventana_pct": round((total_dias / (total_dias_ventana * total_camas)) * 100, 2) if total_camas else 0.0,
            "ingresos_total": total_ingresos,
        })

    return filas


def _generar_excel_ocupacion(desde, hasta, filas, username):
    """2026-06-16: Genera Excel reutilizando plantilla visual base de ServiceExcel (core)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ocupacion"

    total_dias_ventana = len(list(_iterar_dias_intervalo(desde, hasta)))
    total_dias_ventana = max(total_dias_ventana, 1)

    columnas = [
        "Nivel",
        "Servicio",
        "Sala",
        "Cama",
        "N camas",
        "Horas ocupada",
        "Horas pre-alta",
        "Horas vacia",
        "Horas fuera servicio",
        "Horas consulta externa",
        "Horas ventana",
        "% ocupacion",
        "Dias ocupada",
        "Dias ventana",
        "% dia ocupacion",
        "Ingresos total",
    ]
    anchos = [16, 30, 26, 18, 12, 16, 16, 16, 20, 20, 16, 14, 14, 14, 14, 16]
    for idx, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = ancho

    col_fin = get_column_letter(len(columnas))
    ServiceExcel.dibujar_encabezado_excel(ws, col_fin=col_fin)
    ws.append([])
    ws.append([])
    ws.merge_cells(f"A{ws.max_row}:{col_fin}{ws.max_row}")
    titulo = ws.cell(row=ws.max_row, column=1)
    titulo.value = "Dashboard Mapeo de Camas - Ocupacion por servicio/sala/cama"
    titulo.font = Font(name="Helvetica", bold=True, size=12)
    titulo.alignment = Alignment(horizontal="center", vertical="center")

    ws.append([])
    ws.merge_cells(f"A{ws.max_row}:{col_fin}{ws.max_row}")
    subtitulo = ws.cell(row=ws.max_row, column=1)
    subtitulo.value = f"Rango: {hora_local_iso(desde)} -> {hora_local_iso(hasta)}"
    subtitulo.font = Font(name="Helvetica", size=10, bold=True)
    subtitulo.alignment = Alignment(horizontal="center", vertical="center")

    ws.append(columnas)
    fila_header = ws.max_row
    for col in range(1, len(columnas) + 1):
        c = ws.cell(row=fila_header, column=col)
        c.font = ServiceExcel.FONT_TITULO_TABLA
        c.fill = ServiceExcel.FILL_TITULO_TABLA
        c.alignment = ServiceExcel.ALIGN_TITULO
        c.border = ServiceExcel.BORDER_GENERAL

    for item in filas:
        ws.append([
            item["nivel"],
            item["servicio"],
            item["sala"],
            item["cama"],
            item.get("n_camas", 0),
            item["horas_ocupada"],
            item.get("horas_pre_alta", 0.0),
            item.get("horas_vacia", 0.0),
            item.get("horas_fuera_servicio", 0.0),
            item.get("horas_consulta_externa", 0.0),
            item["horas_ventana"],
            item["ocupacion_pct"] / 100.0,
            item["dias_ocupada"],
            item.get("dias_ventana", 0),
            item.get("dias_ventana_pct", item["dias_ocupada"] / total_dias_ventana if total_dias_ventana else 0.0) / 100.0,
            item["ingresos_total"],
        ])
        f = ws.max_row
        es_total = item["nivel"].startswith("TOTAL_")
        es_super_total = item["nivel"] == "SUPER_TOTAL"
        for col in range(1, len(columnas) + 1):
            c = ws.cell(row=f, column=col)
            c.border = ServiceExcel.BORDER_GENERAL
            c.alignment = Alignment(horizontal="center", vertical="center")
            if col in (2, 3, 4):
                c.alignment = Alignment(horizontal="left", vertical="center")
            if es_super_total:
                c.font = ServiceExcel.FONT_TOTAL
                c.fill = ServiceExcel.FILL_TOTAL
            elif es_total:
                c.font = ServiceExcel.FONT_SUBTITULO
                c.fill = ServiceExcel.FILL_SUBTITULO
            else:
                c.font = ServiceExcel.FONT_GENERAL

        for i in range(1, len(columnas) + 1):
            if i == 12:
                ws.cell(row=f, column=i).number_format = "0.00%"
            if i == 15:
                ws.cell(row=f, column=i).number_format = "0.00%"

    ServiceExcel.dibujar_pie_excel(ws, hora_local_iso(timezone.now()), username or "", username or "")

    nombre = (
        f"dashboard_ocupacion_mapeo_{desde.strftime('%Y%m%d')}_{hasta.strftime('%Y%m%d')}.xlsx"
    )
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre}"'
    wb.save(response)
    return response


# =============================================================================
# [2026-05-28] Dashboard de KPIs hospitalarios en tiempo real
# =============================================================================
class DashboardMapeoCamasView(LoginRequiredMixin, TemplateView):
    """[2026-05-28] Dashboard operativo de KPIs y gráficas en tiempo real."""
    template_name = "mapeo_camas/dashboard/dashboard.html"

    def dispatch(self, request, *args, **kwargs):
        if not _tiene_permiso_dashboard(request.user):
            return redirect("acceso_denegado")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["titulo"] = "Dashboard · Mapeo de Camas"
        context["subtitulo"] = "Indicadores operativos y monitoreo continuo"
        return context


@login_required
@require_GET
def dashboard_kpis(request):
    if not _tiene_permiso_dashboard(request.user):
        return _dashboard_error("Acceso denegado.", status=403)
    try:
        desde, hasta = _dashboard_parse_range(request)
        # 2026-06-16: KPIs del dashboard alimentados solo con tablas de mapeo_camas.
        conteo_por_estado = _snapshot_estado_camas(hasta)
        total_camas = sum(conteo_por_estado.values())
        # [2026-07-10] Las cards de estados pasan a lectura por horas activas dentro del rango.
        metricas_horas = _calcular_horas_por_estado_dashboard(desde, hasta)
        pct = metricas_horas["porcentajes_por_kpi"].get("ocupadas", 0.0)

        return _dashboard_response({
            "estados": conteo_por_estado,
            "total_camas": total_camas,
            "ocupadas": metricas_horas["textos_por_kpi"].get("ocupadas", "0.0 h · 0.0%"),
            "disponibles": metricas_horas["textos_por_kpi"].get("disponibles", "0.0 h · 0.0%"),
            "fuera_servicio": metricas_horas["textos_por_kpi"].get("fuera_servicio", "0.0 h · 0.0%"),
            "porcentaje_ocupacion": pct,
            "consulta_externa": metricas_horas["textos_por_kpi"].get("consulta_externa", "0.0 h · 0.0%"),
            "pre_altas": metricas_horas["textos_por_kpi"].get("pre_altas", "0.0 h · 0.0%"),
            "horas_ventana": metricas_horas["horas_ventana"],
            "horas_ventana_camas": metricas_horas["horas_ventana_camas"],
            "kpis_detalle": {
                campo: {
                    "horas_activas": metricas_horas["horas_por_kpi"].get(campo, 0.0),
                    "porcentaje": metricas_horas["porcentajes_por_kpi"].get(campo, 0.0),
                }
                for campo in ESTADOS_KPI_DASHBOARD.keys()
            },
        }, meta=_rango_meta(desde, hasta))
    except Exception as exc:
        return _dashboard_error(exc)
@login_required
@require_GET
def dashboard_ocupacion_hora(request):
    if not _tiene_permiso_dashboard(request.user):
        return _dashboard_error("Acceso denegado.", status=403)
    try:
        desde, hasta = _dashboard_parse_range(request)
        granularidad = _dashboard_granularidad(desde, hasta)
        # [2026-07-02] Fase 2: serie temporal por cama con estado efectivo usando solo historial.
        # Regla de negocio: solo OCUPADA cuenta como ocupación para esta métrica.
        estado_inicial_por_cama = _snapshot_estado_por_cama_desde_historial(desde)

        eventos = list(
            HistorialEstadoCama.objects
            .filter(fecha_hora__gte=desde, fecha_hora__lte=hasta)
            .values("cama_id", "fecha_hora", "estado_nuevo__codigo", "id")
            .order_by("fecha_hora", "id")
        )

        def _bin_key(dt):
            local = timezone.localtime(dt)
            if granularidad == "hora":
                return local.replace(minute=0, second=0, microsecond=0)
            if granularidad == "dia":
                return local.replace(hour=0, minute=0, second=0, microsecond=0)
            return local.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

        def _bin_label(dt):
            if granularidad == "hora":
                return dt.strftime("%d/%m %H:00")
            if granularidad == "dia":
                return dt.strftime("%d/%m")
            return dt.strftime("%m/%Y")

        def _bin_next(dt):
            if granularidad == "hora":
                return dt + timedelta(hours=1)
            if granularidad == "dia":
                return dt + timedelta(days=1)
            year = dt.year + (1 if dt.month == 12 else 0)
            month = 1 if dt.month == 12 else dt.month + 1
            return dt.replace(year=year, month=month)

        bins = []
        cursor = _bin_key(desde)
        fin = timezone.localtime(hasta)
        while cursor <= fin:
            bins.append(cursor)
            cursor = _bin_next(cursor)
        bins_set = set(bins)

        eventos_por_bin = defaultdict(list)
        camas_ids = set(estado_inicial_por_cama.keys())
        for ev in eventos:
            camas_ids.add(ev["cama_id"])
            b = _bin_key(ev["fecha_hora"])
            if b in bins_set:
                eventos_por_bin[b].append(ev)

        total_camas = len(camas_ids)
        estado_ocupada_por_cama = {
            cama_id: (estado_codigo == "OCUPADA")
            for cama_id, estado_codigo in estado_inicial_por_cama.items()
        }

        items = []
        for b in bins:
            # [2026-07-02] Aplica todos los cambios del bin y toma estado final por cama (0/1).
            for ev in eventos_por_bin.get(b, []):
                estado_ocupada_por_cama[ev["cama_id"]] = (ev.get("estado_nuevo__codigo") == "OCUPADA")

            ocupadas = sum(
                1
                for cama_id in camas_ids
                if estado_ocupada_por_cama.get(cama_id, False)
            )
            ocupadas = max(0, min(ocupadas, total_camas))
            porcentaje = 0.0
            if total_camas > 0:
                porcentaje = (ocupadas / total_camas) * 100.0
            porcentaje = round(max(0.0, min(porcentaje, 100.0)), 1)

            items.append({
                "hora": _bin_label(b),
                "porcentaje": porcentaje,
            })
        return _dashboard_response(
            {"items": items, "granularidad": granularidad},
            meta=_rango_meta(desde, hasta),
        )
    except Exception as exc:
        return _dashboard_error(exc)
@login_required
@require_GET
def dashboard_ultimos_movimientos(request):
    if not _tiene_permiso_dashboard(request.user):
        return _dashboard_error("Acceso denegado.", status=403)
    try:
        desde, hasta = _dashboard_parse_range(request)
        try:
            limit = int(request.GET.get("limit") or 30)
        except (TypeError, ValueError):
            limit = 30
        limit = max(1, min(limit, 500))

        movimientos = (
            MovimientoCama.objects
            .filter(fecha_hora__gte=desde, fecha_hora__lte=hasta)
            .select_related(
                "cama_origen__sala__servicio",
                "cama_destino__sala__servicio",
                "ingreso__paciente",
                "usuario",
            )
            .order_by("-fecha_hora")[:limit]
        )
        items = []
        for mov in movimientos:
            paciente = _paciente_payload(
                mov.ingreso.paciente if mov.ingreso_id else None,
                ingreso_id=mov.ingreso_id,
            )
            servicio_destino = getattr(getattr(mov.cama_destino, "sala", None), "servicio", None)
            servicio_nombre = getattr(servicio_destino, "nombre_servicio", "") or ""
            items.append({
                # 2026-06-16: tabla basada en MovimientoCama para reflejar MOVIMIENTO/TRASLADO real.
                "fecha": hora_local_iso(mov.fecha_hora),
                "tipo": "MOVIMIENTO",
                "cama_origen": _nombre_cama(mov.cama_origen),
                "cama_destino": _nombre_cama(mov.cama_destino),
                "paciente": paciente["nombre"] if paciente else "",
                "servicio": servicio_nombre,
                "usuario": _nombre_usuario(mov.usuario),
            })
        return _dashboard_response({"items": items}, meta=_rango_meta(desde, hasta))
    except Exception as exc:
        return _dashboard_error(exc)


@login_required
@require_GET
def dashboard_ultimos_ingresos(request):
    if not _tiene_permiso_dashboard(request.user):
        return _dashboard_error("Acceso denegado.", status=403)
    try:
        desde, hasta = _dashboard_parse_range(request)
        try:
            limit = int(request.GET.get("limit") or 30)
        except (TypeError, ValueError):
            limit = 30
        limit = max(1, min(limit, 500))

        historiales = (
            HistorialEstadoCama.objects
            .filter(
                fecha_hora__gte=desde,
                fecha_hora__lte=hasta,
                ingreso__isnull=False,
                estado_nuevo__codigo__in=("OCUPADA", "PRE_ALTA"),
            )
            .select_related(
                "cama__sala__servicio",
                "ingreso__paciente",
                "usuario",
                "estado_nuevo",
            )
            .order_by("-fecha_hora")[:limit]
        )

        items = []
        for hist in historiales:
            paciente = _paciente_payload(
                hist.ingreso.paciente if hist.ingreso_id else None,
                ingreso_id=hist.ingreso_id,
            )
            servicio_destino = getattr(getattr(hist.cama, "sala", None), "servicio", None)
            servicio_nombre = getattr(servicio_destino, "nombre_servicio", "") or ""
            items.append({
                # 2026-06-16: ingresos de mapa desde historial real de estado, no solo asignaciones.
                "fecha": hora_local_iso(hist.fecha_hora),
                "tipo": getattr(hist.estado_nuevo, "codigo", "OCUPADA") or "OCUPADA",
                "cama_destino": _nombre_cama(hist.cama),
                "paciente": paciente["nombre"] if paciente else "",
                "servicio": servicio_nombre,
                "usuario": _nombre_usuario(hist.usuario),
            })
        return _dashboard_response({"items": items}, meta=_rango_meta(desde, hasta))
    except Exception as exc:
        return _dashboard_error(exc)


@login_required
@require_GET
def dashboard_export_ocupacion_excel(request):
    if not _tiene_permiso_dashboard(request.user):
        return _dashboard_error("Acceso denegado.", status=403)
    try:
        desde, hasta = _dashboard_parse_range(request)
        filas = _construir_filas_reporte_ocupacion(desde, hasta)
        return _generar_excel_ocupacion(desde, hasta, filas, getattr(request.user, "username", ""))
    except Exception as exc:
        return _dashboard_error(exc)
