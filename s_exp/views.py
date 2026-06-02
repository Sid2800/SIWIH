import json
import logging
from datetime import timedelta, datetime
from io import BytesIO

from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.views.generic import TemplateView
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.csrf import csrf_protect, csrf_exempt

from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Count, Q, F
from django.shortcuts import redirect

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    Workbook = None

from reportlab.lib.pagesizes import LETTER, landscape
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.platypus import (
    BaseDocTemplate, PageTemplate, Frame, Paragraph, Table, TableStyle, Spacer
)

from .models import (
    MotivoSolicitud,
    ExpedientePrestamo,
    SolicitudPrestamo,
    SolicitudExpedienteDetalle,
    Prestamo,
    Devolucion,
    LogHistorico,
    ExpedienteEstadoLog,
)
from usuario.models import PerfilUnidad
from expediente.models import Expediente, PacienteAsignacion
from paciente.models import Paciente

logger = logging.getLogger("s_exp")


# ============================================
# UTILIDAD: Formateo de fecha/hora en zona local
# ============================================
def _fmt_local(dt, formato="%d/%m/%Y %H:%M"):
    """
    Formatea un datetime CONVIRTIÉNDOLO a la zona horaria local del sistema
    (TIME_ZONE = America/Tegucigalpa, UTC-6), en formato de 24 horas.

    Compatibilidad:
      - La BD guarda en UTC (USE_TZ=True), igual que TODOS los módulos.
        NO se cambia el almacenamiento para no romper compatibilidad.
      - La conversión a hora local se hace SOLO al mostrar, con
        timezone.localtime(), idéntico a core/utils/utilidades_fechas.py.

    Formato 24h: "%H:%M" → ej. "09:09" / "20:30".

    Args:
        dt: datetime aware (o None).
        formato: patrón strftime (por defecto 24h).

    Returns:
        str: fecha/hora local formateada, o '' si dt es None.
    """
    if not dt:
        return ''
    # Si es aware, convertir a la zona local; si es naive, asumir que ya es local
    if timezone.is_aware(dt):
        dt = timezone.localtime(dt)
    return dt.strftime(formato).strip()


# ============================================
# UTILIDAD: Registrar Log en BD
# ============================================
def _registrar_log(usuario, accion, descripcion, objeto_tipo=None, objeto_id=None):
    """
    Registra un evento en la bitácora de auditoría del sistema (ExpedienteLog).
    
    Args:
        usuario: Instancia de User que realiza la acción.
        accion: Código de la acción (ej. 'SOLICITUD_CREADA').
        descripcion: Texto explicativo del evento.
        objeto_tipo: Nombre del modelo afectado (opcional).
        objeto_id: ID del registro afectado (opcional).
    """
    LogHistorico.objects.create(
        accion=accion,
        usuario=usuario,
        detalle=descripcion,
        objeto_tipo=objeto_tipo,
        objeto_id=objeto_id,
    )


# ============================================
# UTILIDAD: Verificar permisos basados en PerfilUnidad
# ============================================
def _es_exp_admin(user):
    """
    Verifica si un usuario tiene permisos administrativos sobre el módulo de expedientes.
    Requisitos:
    - Debe estar registrado en RRHH (rrhh_empleado)
    - Debe tener uno de estos roles: Administrador, Digitador, Directivo
    """
    # Validación global: usuario debe estar en RRHH
    if not _es_usuario_valido_rrhh(user):
        return False

    if user.is_superuser or user.is_staff:
        return True
    if user.groups.filter(name='administradores').exists():
        return True
    # Permitir acceso a usuarios con roles: admin, digitador, directivo
    return PerfilUnidad.objects.filter(
        usuario=user,
        rol__in=['admin', 'digitador', 'directivo']
    ).exists()


def _es_usuario_valido_rrhh(user):
    """
    Verifica si el usuario está completamente registrado en la cadena RRHH.
    Requisito previo para cualquier acceso al módulo s_exp.

    Valida la cadena completa:
    1. Existe en rrhh_empleado (usuario_id = user.id)
    2. Existe en rrhh_personalnoclinico O rrhh_personalsalud (empleado_id)
    3. Tiene servicio_unidad_id asignado en uno de los dos

    Retorna: True si está completamente registrado, False en caso contrario
    """
    from rrhh.models import Empleado, PersonalNoClinico, PersonalSalud

    try:
        # Paso 1: Verificar que existe rrhh_empleado
        empleado = Empleado.objects.filter(usuario_id=user.id).first()
        if not empleado:
            logger.warning(f"Usuario {user.username} (id={user.id}) no existe en rrhh_empleado")
            return False

        # Paso 2 & 3: Verificar que está en PersonalNoClinico O PersonalSalud CON servicio_unidad
        personal_no_clinico = PersonalNoClinico.objects.filter(
            empleado_id=empleado.id,
            servicio_unidad_id__isnull=False  # Verificar que servicio_unidad_id está asignado
        ).exists()

        if personal_no_clinico:
            logger.info(f"Usuario {user.username}: Validado - PersonalNoClinico con servicio_unidad")
            return True

        personal_salud = PersonalSalud.objects.filter(
            empleado_id=empleado.id,
            servicio_unidad_id__isnull=False  # Verificar que servicio_unidad_id está asignado
        ).exists()

        if personal_salud:
            logger.info(f"Usuario {user.username}: Validado - PersonalSalud con servicio_unidad")
            return True

        # No está en PersonalNoClinico ni PersonalSalud, o no tiene servicio_unidad
        logger.warning(f"Usuario {user.username}: Empleado registrado pero sin PersonalNoClinico/PersonalSalud o sin servicio_unidad asignado")
        return False

    except Exception as e:
        logger.error(f"Error al validar RRHH para usuario {user.username}: {e}", exc_info=True)
        return False


def _es_exp_solicitante(user):
    """True si el usuario puede acceder al módulo (solicitar expedientes).
    Condiciones:
    - Debe estar registrado en RRHH (rrhh_empleado)
    - Debe tener rol 'exp_solicitante' o 'admin' en PerfilUnidad, o ser admin del módulo
    """
    # Validación global: usuario debe estar en RRHH
    if not _es_usuario_valido_rrhh(user):
        return False

    if _es_exp_admin(user):
        return True
    return PerfilUnidad.objects.filter(usuario=user, rol='exp_solicitante').exists()


def _get_unidad_usuario(user):
    """
    Obtiene el NOMBRE de la unidad del usuario con cascada de resolución:
      1. PerfilUnidad (sistema de roles del módulo s_exp)
      2. Cadena RRHH: Empleado → PersonalNoClinico/PersonalSalud → servicio_unidad

    Esto cubre tanto solicitantes (que suelen tener PerfilUnidad) como
    admins/digitadores (que suelen estar solo en RRHH). Devuelve '' si no
    se encuentra en ninguno.
    """
    # 1) PerfilUnidad
    perfil = PerfilUnidad.objects.filter(usuario=user).select_related('servicio_unidad').first()
    if perfil and perfil.servicio_unidad:
        return perfil.servicio_unidad.nombre_unidad

    # 2) Cadena RRHH (reutilizamos el servicio de datos de solicitud)
    try:
        from s_exp.services.datos_solicitud import UbicacionUsuario
        unidad = UbicacionUsuario.resolver(user)
        if unidad:
            return unidad.nombre_unidad
    except Exception:
        pass

    return ''


def _get_servicio_unidad_from_rrhh(user):
    """
    Obtiene la Unidad de Servicio del usuario mediante la cadena de relaciones RRHH.

    Verifica explícitamente la cadena de relaciones:
    1. auth_user.id
    2. rrhh_empleado donde usuario_id = auth_user.id
    3. rrhh_personalnoclinico donde empleado_id = rrhh_empleado.id
    4. servicio_unidad_id en rrhh_personalnoclinico

    Retorna:
        - Tuple: (servicio_unidad_obj, es_valido)
        - servicio_unidad_obj: La instancia de ServicioUnidad o None si no existe
        - es_valido: True si el usuario está correctamente registrado en RRHH, False en caso contrario
    """
    from rrhh.models import Empleado, PersonalNoClinico, PersonalSalud

    try:
        # Paso 1: Verificar que existe rrhh_empleado donde usuario_id = user.id
        empleado = Empleado.objects.filter(usuario_id=user.id).first()
        if not empleado:
            logger.warning(f"Usuario {user.username} (id={user.id}) no tiene registro en rrhh_empleado")
            return None, False

        # Paso 2: Intentar obtener PersonalNoClinico donde empleado_id = empleado.id
        personal_no_clinico = PersonalNoClinico.objects.filter(
            empleado_id=empleado.id
        ).select_related('servicio_unidad').first()

        if personal_no_clinico:
            # Paso 3: Verificar que servicio_unidad_id está asignado
            if personal_no_clinico.servicio_unidad_id:
                logger.info(f"Usuario {user.username}: ServicioUnidad {personal_no_clinico.servicio_unidad_id} desde PersonalNoClinico")
                return personal_no_clinico.servicio_unidad, True
            else:
                logger.warning(f"Usuario {user.username}: PersonalNoClinico sin servicio_unidad asignado")
                return None, True  # Registrado en RRHH pero sin unidad

        # Paso 2 (alternativo): Intentar obtener PersonalSalud si no tiene PersonalNoClinico
        personal_salud = PersonalSalud.objects.filter(
            empleado_id=empleado.id
        ).select_related('servicio_unidad').first()

        if personal_salud:
            # Paso 3: Verificar que servicio_unidad_id está asignado
            if personal_salud.servicio_unidad_id:
                logger.info(f"Usuario {user.username}: ServicioUnidad {personal_salud.servicio_unidad_id} desde PersonalSalud")
                return personal_salud.servicio_unidad, True
            else:
                logger.warning(f"Usuario {user.username}: PersonalSalud sin servicio_unidad asignado")
                return None, True  # Registrado en RRHH pero sin unidad

        # Si llegamos aquí: empleado existe pero sin PersonalNoClinico ni PersonalSalud
        logger.warning(f"Usuario {user.username}: Empleado registrado pero sin PersonalNoClinico ni PersonalSalud")
        return None, True

    except Exception as e:
        logger.error(f"Error al verificar RRHH para usuario {user.username}: {e}", exc_info=True)
        return None, False


def _resolver_ubicacion_expediente(expediente, info_exp=None):
    """
    Resuelve la ubicación ACTUAL del expediente (texto legible).

    Prioridad (obtención híbrida durante la transición a expediente_ubicacion):
    1. NUEVO catálogo: ExpedientePrestamo.ubicacion (FK a ExpedienteUbicacion)
       → es la fuente más precisa para préstamos del módulo s_exp.
    2. LEGACY: expediente.localizacion.descripcion_localizacion
       → lo que usa el módulo expediente (atenciones/ingresos).
    3. LEGACY: info_exp.ubicacion_fisica (texto libre antiguo).
    4. "ADMISION" como último fallback (ARCHIVO quedó deprecado en s_exp).

    Args:
        expediente: instancia de Expediente
        info_exp: instancia opcional de ExpedientePrestamo

    Returns:
        str: descripción de la ubicación actual
    """
    # 1) Nuevo catálogo (FK relacional) — solo si el expediente está prestado/movido
    try:
        if info_exp and getattr(info_exp, 'ubicacion_id', None):
            desc = info_exp.ubicacion.descripcion
            if desc:
                return desc
    except Exception:
        pass

    # 2) Legacy: localizacion del expediente (atenciones/ingresos)
    try:
        if expediente.localizacion and expediente.localizacion.descripcion_localizacion:
            return expediente.localizacion.descripcion_localizacion
    except Exception:
        pass

    # 3) Legacy: texto libre antiguo
    if info_exp and getattr(info_exp, 'ubicacion_fisica', None):
        return info_exp.ubicacion_fisica

    return "ADMISION"


def _set_localizacion_por_solicitud(expediente, solicitud, usuario_admin):
    """
    Actualiza expediente.localizacion al entregar un préstamo.

    La nueva ubicación es la unidad del SOLICITANTE, obtenida desde:
    SolicitudPrestamo.servicio_unidad (capturada via RRHH al crear la solicitud)
    o, si no existe, se intenta resolver desde la cadena RRHH del usuario.

    Args:
        expediente: instancia de Expediente
        solicitud: instancia de SolicitudPrestamo
        usuario_admin: usuario que aprueba/entrega el préstamo

    Returns:
        str: descripción de la nueva ubicación asignada
    """
    from expediente.models import Localizacion

    nombre_ubicacion = None

    # 1. Intentar tomar de SolicitudPrestamo.servicio_unidad
    try:
        if solicitud.servicio_unidad and solicitud.servicio_unidad.nombre_unidad:
            nombre_ubicacion = solicitud.servicio_unidad.nombre_unidad
    except Exception:
        pass

    # 2. Fallback: resolver via cadena RRHH del usuario solicitante
    if not nombre_ubicacion:
        try:
            servicio_unidad, _ok = _get_servicio_unidad_from_rrhh(solicitud.usuario)
            if servicio_unidad and servicio_unidad.nombre_unidad:
                nombre_ubicacion = servicio_unidad.nombre_unidad
        except Exception:
            pass

    # 3. Fallback final: el nombre de la unidad si está, o el texto genérico "PRESTADO".
    if not nombre_ubicacion:
        if solicitud.servicio_unidad_id and solicitud.servicio_unidad:
            nombre_ubicacion = (solicitud.servicio_unidad.nombre_unidad or '').strip()
        if not nombre_ubicacion:
            nombre_ubicacion = 'PRESTADO'

    nombre_ubicacion = nombre_ubicacion.upper()

    try:
        loc_obj, _ = Localizacion.objects.get_or_create(
            descripcion_localizacion=nombre_ubicacion,
            defaults={'estado': True}
        )
        expediente.localizacion = loc_obj
        expediente.modificado_por = usuario_admin
        expediente.save(update_fields=['localizacion', 'modificado_por', 'fecha_modificado'])
        return nombre_ubicacion
    except Exception as e:
        logger.warning(f"No se pudo actualizar localizacion del expediente #{expediente.numero}: {e}")
        return nombre_ubicacion


def _set_localizacion_admision(expediente, usuario_admin):
    """
    Devuelve expediente.localizacion (LEGACY) a 'ADMISION' tras una devolución.

    El módulo de Solicitud de Expedientes ya NO usa 'ARCHIVO': cuando un
    expediente se devuelve, regresa a ADMISION. Esta función mantiene
    sincronizado el campo legacy expediente.localizacion (texto) mientras dura
    la transición; la fuente principal es ExpedientePrestamo.ubicacion (FK al
    catálogo expediente_ubicacion).

    Args:
        expediente: instancia de Expediente
        usuario_admin: usuario que recibe la devolución

    Returns:
        str: "ADMISION"
    """
    from expediente.models import Localizacion

    try:
        loc_obj = Localizacion.objects.filter(
            descripcion_localizacion__iexact='ADMISION'
        ).first()
        if not loc_obj:
            loc_obj, _ = Localizacion.objects.get_or_create(
                descripcion_localizacion='ADMISION',
                defaults={'estado': True}
            )
        expediente.localizacion = loc_obj
        expediente.modificado_por = usuario_admin
        expediente.save(update_fields=['localizacion', 'modificado_por', 'fecha_modificado'])
    except Exception as e:
        logger.warning(f"No se pudo regresar a ADMISION el expediente #{expediente.numero}: {e}")

    return 'ADMISION'


# ============================================
# MIXIN: Acceso basado en Groups
# ============================================
class SExpAdminMixin(LoginRequiredMixin):
    """Acceso solo para administradores del módulo."""
    def dispatch(self, request, *args, **kwargs):
        if not _es_exp_admin(request.user):
            return redirect('acceso_denegado')
        return super().dispatch(request, *args, **kwargs)


class SExpUsuarioMixin(LoginRequiredMixin):
    """Acceso para cualquier usuario con acceso al módulo (Solicitantes + Admin)."""
    def dispatch(self, request, *args, **kwargs):
        if not _es_exp_solicitante(request.user):
            return redirect('acceso_denegado')
        return super().dispatch(request, *args, **kwargs)


# ============================================
# VISTAS ADMIN (Templates)
# ============================================

class DashboardAdminView(SExpAdminMixin, TemplateView):
    """Redirige a Gestión de Solicitudes como landing del admin."""
    def get(self, request, *args, **kwargs):
        return redirect('s_exp_solicitudes')


class GestionSolicitudesView(SExpAdminMixin, TemplateView):
    template_name = 's_exp/gestion_solicitudes.html'


class MonitoreoPrestamosView(SExpAdminMixin, TemplateView):
    template_name = 's_exp/monitoreo_prestamos.html'


class ControlDevolucionesView(SExpAdminMixin, TemplateView):
    template_name = 's_exp/control_devoluciones.html'


class ReportesView(SExpAdminMixin, TemplateView):
    template_name = 's_exp/reportes.html'


# ============================================
# VISTAS USUARIO (Templates)
# ============================================

class BuscadorExpedientesView(SExpUsuarioMixin, TemplateView):
    template_name = 's_exp/buscador_expedientes.html'


class SeguimientoView(SExpUsuarioMixin, TemplateView):
    template_name = 's_exp/seguimiento_usuario.html'


# ============================================
# APIs ADMIN - Dashboard Stats
# ============================================

@require_GET
def dashboard_stats_api(request):
    """Retorna estadísticas para el dashboard del admin."""
    if not _es_exp_admin(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)

    try:
        from expediente.models import Expediente
        total = Expediente.objects.count()

        # Expedientes con préstamo activo
        prestados = ExpedientePrestamo.objects.filter(estado_id='EXP_PRESTADO').count()
        disponibles = total - prestados

        solicitudes_pendientes = SolicitudPrestamo.objects.filter(estado_flujo_id='SOL_PENDIENTE').count()

        ahora = timezone.now()
        prestamos_activos = Prestamo.objects.filter(estado='Entregado').count()
        prestamos_vencidos = Prestamo.objects.filter(
            estado='Entregado',
            fecha_limite__lt=ahora
        ).count()

        # Próximos a vencer (más del 90% de tiempo usado)
        proximos_vencer = 0
        for p in Prestamo.objects.filter(estado='Entregado', fecha_limite__gte=ahora):
            if p.porcentaje_tiempo_usado >= 90:
                proximos_vencer += 1

        devoluciones_parciales = Prestamo.objects.filter(estado='DevolucionParcial').count()

        return JsonResponse({
            "total_expedientes": total,
            "disponibles": disponibles,
            "prestados": prestados,
            "solicitudes_pendientes": solicitudes_pendientes,
            "prestamos_activos": prestamos_activos,
            "prestamos_vencidos": prestamos_vencidos,
            "proximos_vencer": proximos_vencer,
            "devoluciones_parciales": devoluciones_parciales,
        })
    except Exception as e:
        logger.error(f"Error en dashboard_stats_api: {e}", exc_info=True)
        return JsonResponse({"error": "Error interno del servidor"}, status=500)


# ============================================
# APIs ADMIN - Gestión de Solicitudes
# ============================================

@require_GET
def listar_solicitudes_api(request):
    """
    API para alimentar el DataTable de gestión de solicitudes (Admin).
    Soporta filtrado por estado y búsqueda server-side.
    """
    """Lista solicitudes para el admin (DataTables server-side)."""
    if not _es_exp_admin(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)

    try:
        draw = int(request.GET.get('draw', 0))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 20))
        search_value = request.GET.get('search[value]', '').strip()
        estado_filtro = request.GET.get('estado', '')

        # Optimizamos joins para reducir queries N+1: traemos en una sola consulta
        # todas las relaciones necesarias para construir la respuesta.
        qs = SolicitudPrestamo.objects.select_related(
            'usuario', 'servicio_unidad', 'motivo', 'estado_flujo'
        ).annotate(cant_expedientes=Count('detalles'))

        if estado_filtro:
            qs = qs.filter(estado_flujo_id=estado_filtro)

        if search_value:
            qs = qs.filter(
                Q(usuario__username__icontains=search_value) |
                Q(motivo__nombre__icontains=search_value) |
                Q(id__icontains=search_value)
            )

        total_records = SolicitudPrestamo.objects.count()
        filtered_records = qs.count()

        solicitudes = qs.order_by('-fecha_creacion')[start:start + length]

        # Importamos los servicios de acceso a datos.
        # Toda la lectura de DNI/nombre/etc. pasa por aquí, NO se accede a
        # campos snapshot deprecados directamente.
        from s_exp.services.datos_solicitud import DatosDetalleSolicitud, DatosSolicitud

        data = []
        for s in solicitudes:
            # Cada detalle se enriquece con datos vivos del paciente/expediente
            detalles_info = []
            for d in s.detalles.select_related(
                'expediente_prestamo__expediente', 'paciente'
            ):
                info = DatosDetalleSolicitud.enriquecer(d)
                detalles_info.append(info)

            prestamo_id = None
            try:
                prestamo_id = s.prestamo.id
            except Exception:
                prestamo_id = None

            # Construimos el dict de respuesta usando los servicios
            data.append({
                "id": s.id,
                "prestamo_id": prestamo_id,
                "usuario": DatosSolicitud.usuario_username(s),
                "usuario_nombre": DatosSolicitud.usuario_nombre_completo(s),
                "fecha_creacion": _fmt_local(s.fecha_creacion),
                "estado_flujo": DatosSolicitud.estado_codigo(s),
                "estado_flujo_nombre": DatosSolicitud.estado_nombre(s),
                "motivo": DatosSolicitud.motivo_nombre(s),
                "observaciones": s.observaciones or "",
                # 'unidad' reemplaza tanto area_destino como servicio_unidad antiguos
                "unidad": DatosSolicitud.unidad_nombre(s),
                "unidad_id": DatosSolicitud.unidad_id(s),
                # Mantenemos 'area_destino' como alias para compatibilidad con
                # frontend existente que aún lee ese key (se removerá luego).
                "area_destino": DatosSolicitud.unidad_nombre(s),
                "cant_expedientes": s.cant_expedientes,
                "expedientes": detalles_info,
                "tiempo_sugerido_horas": s.tiempo_sugerido_horas,
            })

        return JsonResponse({
            "draw": draw,
            "recordsTotal": total_records,
            "recordsFiltered": filtered_records,
            "data": data,
        })
    except Exception as e:
        logger.error(f"Error en listar_solicitudes_api: {e}", exc_info=True)
        return JsonResponse({"error": "Error interno del servidor"}, status=500)


@csrf_protect
@require_POST
def aprobar_solicitud_api(request):
    """
    Aprueba una solicitud y crea el préstamo.
    Soporta decisiones individuales por expediente (aprobado/rechazado).
    Si todos los expedientes son rechazados, la solicitud pasa a SOL_RECHAZADA.
    """
    if not _es_exp_admin(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Datos inválidos"}, status=400)

    solicitud_id = body.get('solicitud_id')
    tiempo_limite = body.get('tiempo_limite_horas', 24)
    es_minutos = body.get('es_minutos', False)
    expedientes_decisiones = body.get('expedientes_decisiones', [])

    if int(tiempo_limite) < 1:
        return JsonResponse({"error": "El tiempo debe ser mayor a 0"}, status=400)

    # Validar tope de 72 horas cuando no es modo minutos (el frontend ya convierte días a horas)
    if not es_minutos and int(tiempo_limite) > 72:
        return JsonResponse({"error": "El tiempo máximo de préstamo es 72 horas (3 días)"}, status=400)

    # Mapa de decisiones: {detalle_id: {aprobado, observaciones}}
    mapa_decisiones = {}
    for d in expedientes_decisiones:
        det_id = d.get('detalle_id')
        if det_id is None:
            continue
        aprobado = d.get('aprobado', True)
        mapa_decisiones[det_id] = {
            'aprobado': aprobado,
            'observaciones': (d.get('observaciones') or '').strip(),
        }

    try:
        solicitud = SolicitudPrestamo.objects.get(id=solicitud_id, estado_flujo_id='SOL_PENDIENTE')
    except SolicitudPrestamo.DoesNotExist:
        return JsonResponse({"error": "Solicitud no encontrada o ya procesada"}, status=404)

    try:
        detalles = list(solicitud.detalles.select_related('expediente_prestamo__expediente'))

        # Verificar que los expedientes aprobados estén disponibles
        for d in detalles:
            info = mapa_decisiones.get(d.id, {'aprobado': True, 'observaciones': ''})
            if info['aprobado'] and d.expediente_prestamo.estado_id == 'EXP_PRESTADO':
                return JsonResponse({
                    "error": f"El expediente #{d.expediente_prestamo.expediente.numero} ya no está disponible"
                }, status=400)

        # Aplicar decisiones por expediente
        aprobados = []
        rechazados = []
        for d in detalles:
            info = mapa_decisiones.get(d.id, {'aprobado': True, 'observaciones': ''})
            d.aprobado = info['aprobado']
            # Guardar SIEMPRE las observaciones (tanto aprobados como rechazados pueden tenerlas)
            d.motivo_rechazo_individual = info['observaciones'] or None
            if info['aprobado']:
                aprobados.append(d)
            else:
                rechazados.append(d)
            d.save()

        # Texto motivo general: usamos las primeras observaciones de rechazados (para Prestamo.motivo_rechazo si aplica)
        motivo_rechazo_general = " | ".join(
            f"#{d.expediente_prestamo.expediente.numero}: {d.motivo_rechazo_individual}"
            for d in rechazados if d.motivo_rechazo_individual
        )

        todos_rechazados = len(aprobados) == 0

        if todos_rechazados:
            # Rechazar toda la solicitud
            solicitud.estado_flujo_id = 'SOL_RECHAZADA'
            solicitud.save()

            for d in rechazados:
                ep = d.expediente_prestamo
                estado_ant = ep.estado
                ep.estado_id = 'EXP_DISPONIBLE'
                ep.save()
                ExpedienteEstadoLog.objects.create(
                    expediente=ep.expediente,
                    estado_anterior=estado_ant,
                    estado_nuevo_id='EXP_DISPONIBLE',
                    usuario=request.user,
                    solicitud=solicitud,
                    observacion=f"Liberado: todos los expedientes rechazados. Motivo: {motivo_rechazo_general}"
                )

            Prestamo.objects.create(
                solicitud=solicitud,
                admin_aprobador=request.user,
                motivo_rechazo=motivo_rechazo_general,
                estado='Cerrado'
            )

            _registrar_log(
                request.user, 'SOLICITUD_RECHAZADA',
                f'Solicitud #{solicitud.id} rechazada (todos los expedientes rechazados individualmente). Motivo: {motivo_rechazo_general}',
                'SolicitudPrestamo', solicitud.id
            )
            return JsonResponse({"success": True, "todos_rechazados": True})

        # Al menos un expediente aprobado: continuar con la solicitud
        solicitud.estado_flujo_id = 'SOL_APROBADA_ORGANIZANDO'
        solicitud.save()

        # Aprobados → EXP_APARTADO
        for d in aprobados:
            if d.expediente_prestamo.estado_id != 'EXP_APARTADO':
                estado_ant = d.expediente_prestamo.estado
                d.expediente_prestamo.estado_id = 'EXP_APARTADO'
                d.expediente_prestamo.save()
                ExpedienteEstadoLog.objects.create(
                    expediente=d.expediente_prestamo.expediente,
                    estado_anterior=estado_ant,
                    estado_nuevo_id='EXP_APARTADO',
                    usuario=request.user,
                    solicitud=solicitud,
                    observacion="Apartado al aprobar solicitud"
                )

        # Rechazados → EXP_DISPONIBLE
        for d in rechazados:
            ep = d.expediente_prestamo
            estado_ant = ep.estado
            ep.estado_id = 'EXP_DISPONIBLE'
            ep.save()
            ExpedienteEstadoLog.objects.create(
                expediente=ep.expediente,
                estado_anterior=estado_ant,
                estado_nuevo_id='EXP_DISPONIBLE',
                usuario=request.user,
                solicitud=solicitud,
                observacion=f"No se prestará en esta solicitud. Motivo: {motivo_rechazo_general}"
            )

        prestamo = Prestamo.objects.create(
            solicitud=solicitud,
            admin_aprobador=request.user,
            tiempo_limite_horas=int(tiempo_limite),
            es_minutos=es_minutos,
            estado='Activo'
        )

        detalle_rechazo = f" ({len(rechazados)} expediente(s) rechazado(s))" if rechazados else ""
        _registrar_log(
            request.user, 'SOLICITUD_APROBADA',
            f'Solicitud #{solicitud.id} aprobada{detalle_rechazo}. En proceso de organización.',
            'Prestamo', prestamo.id
        )

        logger.info(f"Solicitud #{solicitud.id} aprobada por {request.user.username}")
        return JsonResponse({"success": True, "todos_rechazados": False, "prestamo_id": prestamo.id})

    except Exception as e:
        logger.error(f"Error en aprobar_solicitud_api: {e}", exc_info=True)
        return JsonResponse({"error": "Error interno del servidor"}, status=500)


@require_GET
def expedientes_revision_api(request, solicitud_id):
    """Retorna los expedientes APROBADOS de una solicitud en revisión (estado SOL_APROBADA_ORGANIZANDO)."""
    if not _es_exp_admin(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)

    try:
        solicitud = SolicitudPrestamo.objects.get(
            id=solicitud_id, estado_flujo_id='SOL_APROBADA_ORGANIZANDO'
        )
    except SolicitudPrestamo.DoesNotExist:
        return JsonResponse({"error": "Solicitud no encontrada o no está en revisión"}, status=404)

    from s_exp.services.datos_solicitud import DatosDetalleSolicitud

    expedientes = []
    for d in solicitud.detalles.select_related(
        'expediente_prestamo__expediente', 'paciente'
    ).filter(aprobado=True):
        expedientes.append({
            "detalle_id": d.id,
            "numero": DatosDetalleSolicitud.numero_expediente(d),
            "paciente_id": DatosDetalleSolicitud.paciente_id(d),
            "paciente_nombre": DatosDetalleSolicitud.paciente_nombre_completo(d),
            "paciente_identidad": DatosDetalleSolicitud.paciente_dni(d),
        })
    return JsonResponse({"expedientes": expedientes})


@require_GET
def expedientes_solicitud_api(request, solicitud_id):
    """Retorna los expedientes de una solicitud pendiente para el modal de aprobación."""
    if not _es_exp_admin(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)

    try:
        solicitud = SolicitudPrestamo.objects.get(id=solicitud_id, estado_flujo_id='SOL_PENDIENTE')
    except SolicitudPrestamo.DoesNotExist:
        return JsonResponse({"error": "Solicitud no encontrada o ya procesada"}, status=404)

    try:
        # Importamos los servicios — sin acceso directo a snapshots
        from s_exp.services.datos_solicitud import DatosDetalleSolicitud

        expedientes = []
        for d in solicitud.detalles.select_related(
            'expediente_prestamo__expediente', 'paciente'
        ):
            expedientes.append({
                "detalle_id": d.id,
                "numero": DatosDetalleSolicitud.numero_expediente(d),
                "paciente_id": DatosDetalleSolicitud.paciente_id(d),
                "paciente_nombre": DatosDetalleSolicitud.paciente_nombre_completo(d),
                "paciente_identidad": DatosDetalleSolicitud.paciente_dni(d),
                "estado_fisico": d.expediente_prestamo.estado_id,
            })
        return JsonResponse({
            "expedientes": expedientes,
            "tiempo_sugerido_horas": solicitud.tiempo_sugerido_horas,
            "motivo": solicitud.motivo.nombre if solicitud.motivo_id else "",
        })
    except Exception as e:
        logger.error(f"Error en expedientes_solicitud_api: {e}", exc_info=True)
        return JsonResponse({"error": "Error interno del servidor"}, status=500)


@require_GET
def imprimir_solicitud_pdf(request, solicitud_id):
    """Genera y descarga el PDF de una solicitud (aprobada/organizando/listo/prestamo/devolucion/finalizada)."""
    if not _es_exp_admin(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)

    estados_permitidos = {
        'SOL_APROBADA_ORGANIZANDO', 'SOL_LISTO_RECOGER',
        'SOL_EN_PRESTAMO', 'SOL_EN_DEVOLUCION',
        'SOL_FINALIZADA', 'SOL_INCOMPLETA',
    }
    try:
        solicitud = SolicitudPrestamo.objects.select_related(
            'usuario', 'motivo', 'prestamo', 'prestamo__admin_aprobador'
        ).get(id=solicitud_id)
    except SolicitudPrestamo.DoesNotExist:
        return JsonResponse({"error": "Solicitud no encontrada"}, status=404)

    if solicitud.estado_flujo_id not in estados_permitidos:
        return JsonResponse({"error": "La solicitud no está en un estado imprimible"}, status=400)

    try:
        from s_exp.services.pdf_solicitud_service import generar_pdf_solicitud
        pdf_bytes = generar_pdf_solicitud(solicitud)
    except Exception as e:
        logger.error(f"Error generando PDF solicitud {solicitud_id}: {e}", exc_info=True)
        return JsonResponse({"error": "Error al generar el PDF"}, status=500)

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    tz = timezone.get_current_timezone()
    ts = timezone.now().astimezone(tz).strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'inline; filename="solicitud_{solicitud.id}_{ts}.pdf"'
    return response


@csrf_protect
@csrf_protect
@require_POST
def revisar_entrega_api(request):
    """
    Revisión de Entrega — el admin verifica físicamente cada expediente antes de marcar listo.
    Permite desmarcar expedientes que no se encontraron físicamente y registrar comentario por expediente.
    Los desmarcados pasan a EXP_DISPONIBLE y quedan con aprobado=False + motivo_rechazo_individual.
    No cambia el estado de la solicitud (sigue en SOL_APROBADA_ORGANIZANDO).
    """
    if not _es_exp_admin(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Datos inválidos"}, status=400)

    solicitud_id = body.get('solicitud_id')
    decisiones = body.get('decisiones', [])

    try:
        solicitud = SolicitudPrestamo.objects.get(
            id=solicitud_id, estado_flujo_id='SOL_APROBADA_ORGANIZANDO'
        )
    except SolicitudPrestamo.DoesNotExist:
        return JsonResponse({"error": "Solicitud no encontrada o no está en revisión"}, status=404)

    try:
        mapa = {d.get('detalle_id'): d for d in decisiones if d.get('detalle_id') is not None}
        cambios = 0
        for d in solicitud.detalles.select_related('expediente_prestamo__expediente'):
            info = mapa.get(d.id)
            if info is None:
                continue
            encontrado = bool(info.get('encontrado', True))
            comentario = (info.get('comentario') or '').strip()

            if not encontrado and d.aprobado:
                # Marcado como no encontrado físicamente
                d.aprobado = False
                d.motivo_rechazo_individual = comentario or 'No encontrado físicamente'
                d.save()

                ep = d.expediente_prestamo
                estado_ant = ep.estado
                ep.estado_id = 'EXP_DISPONIBLE'
                ep.save()
                ExpedienteEstadoLog.objects.create(
                    expediente=ep.expediente,
                    estado_anterior=estado_ant,
                    estado_nuevo_id='EXP_DISPONIBLE',
                    usuario=request.user,
                    solicitud=solicitud,
                    observacion=f"Revisión de entrega: {d.motivo_rechazo_individual}"
                )
                cambios += 1
            elif encontrado and comentario and comentario != (d.motivo_rechazo_individual or ''):
                # Sólo actualizar comentario sin cambiar aprobación
                d.motivo_rechazo_individual = comentario
                d.save()
                cambios += 1

        # Si todos los expedientes quedaron rechazados, cerrar la solicitud
        aprobados_restantes = solicitud.detalles.filter(aprobado=True).count()
        if aprobados_restantes == 0:
            solicitud.estado_flujo_id = 'SOL_RECHAZADA'
            solicitud.save()
            try:
                p = solicitud.prestamo
                p.estado = 'Cerrado'
                p.save()
            except Exception:
                pass

        _registrar_log(
            request.user, 'REVISION_ENTREGA',
            f'Revisión de entrega para solicitud #{solicitud.id}: {cambios} cambio(s).',
            'SolicitudPrestamo', solicitud.id
        )
        return JsonResponse({
            "success": True,
            "cambios": cambios,
            "todos_rechazados": aprobados_restantes == 0,
        })
    except Exception as e:
        logger.error(f"Error en revisar_entrega_api: {e}", exc_info=True)
        return JsonResponse({"error": "Error interno del servidor"}, status=500)


@csrf_protect
@require_POST
def marcar_listo_recojer_api(request):
    """Admin marca que los expedientes ya están organizados físicamente y listos en ventanilla."""
    if not _es_exp_admin(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)

    try:
        body = json.loads(request.body)
        solicitud_id = body.get('solicitud_id')

        solicitud = SolicitudPrestamo.objects.get(id=solicitud_id, estado_flujo_id='SOL_APROBADA_ORGANIZANDO')
        # Validar que al menos un expediente siga aprobado
        if solicitud.detalles.filter(aprobado=True).count() == 0:
            return JsonResponse({"error": "No hay expedientes aprobados para entregar"}, status=400)
        solicitud.estado_flujo_id = 'SOL_LISTO_RECOGER'
        solicitud.notificado_listo = False  # Reset para que el sistema dispare la alerta al usuario
        solicitud.save()

        _registrar_log(
            request.user, 'SOLICITUD_LISTA',
            f'Solicitud #{solicitud.id} marcada como lista para recoger.',
            'SolicitudPrestamo', solicitud.id
        )

        return JsonResponse({"success": True})
    except SolicitudPrestamo.DoesNotExist:
        return JsonResponse({"error": "Solicitud no encontrada o no está en proceso de organización"}, status=404)
    except Exception as e:
        logger.error(f"Error en marcar_listo_recojer_api: {e}", exc_info=True)
        return JsonResponse({"error": "Error interno del servidor"}, status=500)


@csrf_protect
@require_POST
def rechazar_solicitud_api(request):
    """
    Rechaza una solicitud de préstamo pendiente.
    Libera automáticamente los expedientes que estaban apartados (EXP_APARTADO -> EXP_DISPONIBLE).
    
    Body JSON:
        solicitud_id (int): ID de la solicitud a rechazar.
        motivo (str): Razón del rechazo.
    """
    """Rechaza una solicitud con motivo obligatorio."""
    if not _es_exp_admin(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Datos inválidos"}, status=400)

    solicitud_id = body.get('solicitud_id')
    motivo_rechazo = body.get('motivo_rechazo', '').strip()

    if not motivo_rechazo:
        return JsonResponse({"error": "El motivo de rechazo es obligatorio"}, status=400)

    try:
        solicitud = SolicitudPrestamo.objects.get(id=solicitud_id, estado_flujo_id='SOL_PENDIENTE')
    except SolicitudPrestamo.DoesNotExist:
        return JsonResponse({"error": "Solicitud no encontrada o ya procesada"}, status=404)

    try:
        solicitud.estado_flujo_id = 'SOL_RECHAZADA'
        solicitud.save()

        # Liberar expedientes: volver a ponerlos disponibles
        for detalle in solicitud.detalles.select_related('expediente_prestamo'):
            ep = detalle.expediente_prestamo
            estado_anterior = ep.estado
            ep.estado_id = 'EXP_DISPONIBLE'
            ep.save()

            ExpedienteEstadoLog.objects.create(
                expediente=ep.expediente,
                estado_anterior=estado_anterior,
                estado_nuevo_id='EXP_DISPONIBLE',
                usuario=request.user,
                solicitud=solicitud,
                observacion=f"Expediente liberado por rechazo de solicitud. Motivo: {motivo_rechazo}"
            )

        Prestamo.objects.create(
            solicitud=solicitud,
            admin_aprobador=request.user,
            motivo_rechazo=motivo_rechazo,
            estado='Cerrado'
        )

        _registrar_log(
            request.user, 'SOLICITUD_RECHAZADA',
            f'Solicitud #{solicitud.id} rechazada. Motivo: {motivo_rechazo}',
            'SolicitudPrestamo', solicitud.id
        )

        logger.info(f"Solicitud #{solicitud.id} rechazada por {request.user.username}")
        return JsonResponse({"success": True})

    except Exception as e:
        logger.error(f"Error en rechazar_solicitud_api: {e}", exc_info=True)
        return JsonResponse({"error": "Error interno del servidor"}, status=500)


# ============================================
# APIs ADMIN - Monitoreo de Préstamos
# ============================================

@require_GET
def prestamos_activos_api(request):
    """Lista préstamos activos/entregados para monitoreo con DataTables server-side."""
    if not _es_exp_admin(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)

    try:
        draw = int(request.GET.get('draw', 0))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '').strip()
        estado_filtro = request.GET.get('estado', '')

        qs = Prestamo.objects.select_related(
            'solicitud__usuario', 'solicitud__motivo', 'solicitud__servicio_unidad'
        ).filter(
            estado__in=['Activo', 'Entregado', 'Vencido', 'DevolucionParcial', 'DevueltoVencido']
        )

        if estado_filtro:
            qs = qs.filter(estado=estado_filtro)

        if search_value:
            qs = qs.filter(
                Q(solicitud__usuario__username__icontains=search_value) |
                Q(id__icontains=search_value) |
                Q(solicitud__usuario__first_name__icontains=search_value) |
                Q(solicitud__usuario__last_name__icontains=search_value)
            )

        total_records = Prestamo.objects.filter(
            estado__in=['Activo', 'Entregado', 'Vencido', 'DevolucionParcial', 'DevueltoVencido']
        ).count()
        filtered_records = qs.count()

        prestamos = qs.order_by('-fecha_aprobacion')[start:start + length]

        data = []
        for p in prestamos:
            numeros = list(
                p.solicitud.detalles.select_related('expediente_prestamo__expediente')
                .filter(aprobado=True)
                .values_list('expediente_prestamo__expediente__numero', flat=True)
            )

            # Usamos los servicios para evitar acceso directo a snapshots
            from s_exp.services.datos_solicitud import DatosSolicitud
            data.append({
                "id": p.id,
                "solicitud_id": p.solicitud.id,
                "usuario": DatosSolicitud.usuario_username(p.solicitud),
                "usuario_nombre": DatosSolicitud.usuario_nombre_completo(p.solicitud),
                # 'area_destino' es alias retrocompat — el valor viene de la unidad FK
                "area_destino": DatosSolicitud.unidad_nombre(p.solicitud),
                "unidad": DatosSolicitud.unidad_nombre(p.solicitud),
                "unidad_id": DatosSolicitud.unidad_id(p.solicitud),
                "motivo": DatosSolicitud.motivo_nombre(p.solicitud),
                "estado": p.estado,
                "fecha_aprobacion": _fmt_local(p.fecha_aprobacion),
                "fecha_entrega": _fmt_local(p.fecha_entrega) or None,
                "fecha_limite": p.fecha_limite.isoformat() if p.fecha_limite else None,
                "tiempo_limite_horas": p.tiempo_limite_horas,
                "tiempo_restante_segundos": p.tiempo_restante_segundos,
                "porcentaje_tiempo_usado": p.porcentaje_tiempo_usado,
                "esta_vencido": p.esta_vencido,
                "expedientes": numeros,
                "cant_expedientes": len(numeros),
                "solicitud_estado_flujo": p.solicitud.estado_flujo_id,
            })

        return JsonResponse({
            "draw": draw,
            "recordsTotal": total_records,
            "recordsFiltered": filtered_records,
            "data": data
        })

    except Exception as e:
        logger.error(f"Error en prestamos_activos_api: {e}", exc_info=True)
        return JsonResponse({"error": "Error interno del servidor"}, status=500)


@csrf_protect
@require_POST
def marcar_entregado_api(request):
    """Marca un préstamo como entregado e inicia el cronómetro."""
    if not _es_exp_admin(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Datos inválidos"}, status=400)

    prestamo_id = body.get('prestamo_id')

    try:
        prestamo = Prestamo.objects.get(id=prestamo_id, estado='Activo')
        if prestamo.solicitud.estado_flujo_id != 'SOL_LISTO_RECOGER':
             return JsonResponse({"error": "La solicitud debe estar marcada como 'Listo para recoger' antes de entregar."}, status=400)
    except Prestamo.DoesNotExist:
        return JsonResponse({"error": "Préstamo no encontrado o no está en estado Activo"}, status=404)

    try:
        ahora = timezone.now()
        prestamo.fecha_entrega = ahora

        # Lógica de vencimiento flexible (Pruebas vs Producción)
        if prestamo.es_minutos:
            # Si el préstamo se configuró en minutos (para pruebas)
            prestamo.fecha_limite = ahora + timedelta(minutes=prestamo.tiempo_limite_horas)
        else:
            # Configuración estándar en horas
            prestamo.fecha_limite = ahora + timedelta(hours=prestamo.tiempo_limite_horas)

        prestamo.estado = 'Entregado'
        prestamo.save()

        prestamo.solicitud.estado_flujo_id = 'SOL_EN_PRESTAMO'
        prestamo.solicitud.save()

        # Resolver UNA sola vez la ubicación destino del préstamo (Opción A):
        # el expediente se mueve a la unidad del SOLICITANTE (catálogo nuevo).
        from expediente.services.ubicaciones import CatalogoUbicaciones
        ubicacion_destino = None
        try:
            ubicacion_destino = CatalogoUbicaciones.ubicacion_del_solicitante(prestamo.solicitud)
        except Exception as _e:
            logger.warning(f"No se pudo resolver ubicacion del solicitante: {_e}")

        # Solo marcar como prestados los expedientes aprobados
        for d in prestamo.solicitud.detalles.select_related('expediente_prestamo__expediente').filter(aprobado=True):
            estado_anterior = d.expediente_prestamo.estado
            d.expediente_prestamo.estado_id = 'EXP_PRESTADO'

            # NUEVO: registrar la ubicación actual via FK al catálogo unificado.
            if ubicacion_destino is not None:
                d.expediente_prestamo.ubicacion = ubicacion_destino

            d.expediente_prestamo.save()

            # LEGACY: mantener sincronizado expediente.localizacion (texto) mientras
            # dura la transición — atenciones/ingresos aún leen de ahí.
            try:
                _set_localizacion_por_solicitud(
                    d.expediente_prestamo.expediente,
                    prestamo.solicitud,
                    request.user,
                )
            except Exception as _e:
                logger.warning(f"No se pudo actualizar localizacion legacy al entregar: {_e}")

            ExpedienteEstadoLog.objects.create(
                expediente=d.expediente_prestamo.expediente,
                estado_anterior=estado_anterior,
                estado_nuevo_id='EXP_PRESTADO',
                usuario=request.user,
                solicitud=prestamo.solicitud
            )

        _registrar_log(
            request.user, 'PRESTAMO_ENTREGADO',
            f'Préstamo #{prestamo.id} entregado. Cronómetro iniciado: {prestamo.tiempo_limite_horas}h.',
            'Prestamo', prestamo.id
        )

        logger.info(f"Préstamo #{prestamo.id} entregado por {request.user.username}")
        return JsonResponse({
            "success": True,
            "fecha_entrega": _fmt_local(ahora),  # 12h local
            "fecha_limite": prestamo.fecha_limite.isoformat(),
        })

    except Exception as e:
        logger.error(f"Error en marcar_entregado_api: {e}", exc_info=True)
        return JsonResponse({"error": "Error interno del servidor"}, status=500)


# ============================================
# APIs ADMIN - Control de Devoluciones
# ============================================

@require_GET
def prestamos_para_devolucion_api(request):
    """Lista préstamos que están pendientes de devolución."""
    if not _es_exp_admin(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)

    try:
        # Mostrar solicitudes marcadas para devolución O con devoluciones incompletas pendientes
        qs = Prestamo.objects.select_related('solicitud__usuario').filter(
            solicitud__estado_flujo_id='SOL_EN_DEVOLUCION',
            estado__in=['Entregado', 'Vencido', 'DevolucionParcial']
        ).order_by('fecha_limite')

        # Importamos los servicios — acceso unificado a datos del detalle
        from s_exp.services.datos_solicitud import DatosDetalleSolicitud, DatosSolicitud

        data = []
        for p in qs:
            detalles = []
            # Mostrar TODOS los expedientes aprobados (devueltos y pendientes)
            for d in p.solicitud.detalles.select_related(
                'expediente_prestamo__expediente', 'expediente_prestamo__estado', 'paciente'
            ).filter(aprobado=True):
                estado_fisico_id = d.expediente_prestamo.estado_id or ''
                # Estado de devolución para el front:
                #   - 'pendiente'  : aún no devuelto
                #   - 'devuelto'   : devuelto correctamente (EXP_DISPONIBLE)
                #   - 'perdido'    : marcado como perdido (EXP_PERDIDO)
                if not d.devuelto:
                    estado_devolucion = 'pendiente'
                elif estado_fisico_id == 'EXP_PERDIDO':
                    estado_devolucion = 'perdido'
                else:
                    estado_devolucion = 'devuelto'

                detalles.append({
                    "id": d.id,
                    "numero": DatosDetalleSolicitud.numero_expediente(d),
                    "estado_fisico": d.expediente_prestamo.estado.nombre,
                    "estado_devolucion": estado_devolucion,
                    "devuelto": bool(d.devuelto),
                    "paciente_id": DatosDetalleSolicitud.paciente_id(d),
                    "paciente_identidad": DatosDetalleSolicitud.paciente_dni(d),
                    "paciente_nombre": DatosDetalleSolicitud.paciente_nombre_completo(d),
                    "comentario_devolucion": d.comentario_devolucion or '',
                })

            data.append({
                "id": p.id,
                "solicitud_id": p.solicitud.id,
                "usuario": DatosSolicitud.usuario_username(p.solicitud),
                "usuario_nombre": DatosSolicitud.usuario_nombre_completo(p.solicitud),
                "unidad": DatosSolicitud.unidad_nombre(p.solicitud),
                "estado": p.estado,
                "detalles_expedientes": detalles,
                "cant_expedientes": p.solicitud.detalles.filter(aprobado=True).count(),
                "cant_devueltos": p.solicitud.detalles.filter(aprobado=True, devuelto=True).count(),
                "fecha_limite": p.fecha_limite.isoformat() if p.fecha_limite else None,
                "esta_vencido": p.esta_vencido,
            })

        return JsonResponse({"data": data})

    except Exception as e:
        logger.error(f"Error en prestamos_para_devolucion_api: {e}", exc_info=True)
        return JsonResponse({"error": "Error interno del servidor"}, status=500)


@csrf_protect
@require_POST
def solicitar_devolucion_api(request):
    """El personal (usuario) marca que ya no usará los expedientes y los devuelve al archivo."""
    if not request.user.is_authenticated:
         return JsonResponse({"error": "No autenticado"}, status=401)
    
    try:
        body = json.loads(request.body)
        solicitud_id = body.get('solicitud_id')
        
        # El usuario puede devolver si está en préstamo o si fue incompleta (quedan pendientes)
        solicitud = SolicitudPrestamo.objects.get(
            id=solicitud_id, 
            usuario=request.user, 
            estado_flujo_id__in=['SOL_EN_PRESTAMO', 'SOL_INCOMPLETA']
        )
        
        solicitud.estado_flujo_id = 'SOL_EN_DEVOLUCION'
        solicitud.save()

        _registrar_log(
            request.user, 'SOLICITUD_DEVOLUCION_INICIADA',
            f'Usuario marcó solicitud #{solicitud.id} para devolución.',
            'SolicitudPrestamo', solicitud.id
        )

        return JsonResponse({"success": True})
    except SolicitudPrestamo.DoesNotExist:
        return JsonResponse({"error": "Solicitud no encontrada o no está en préstamo"}, status=404)
    except Exception as e:
        logger.error(f"Error en solicitar_devolucion_api: {e}", exc_info=True)
        return JsonResponse({"error": "Error interno del servidor"}, status=500)


@csrf_protect
@require_POST
def procesar_devolucion_api(request):
    """Admin audita los expedientes recibidos. Marca cuáles llegaron y cuáles se perdieron."""
    if not _es_exp_admin(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)

    try:
        body = json.loads(request.body)
        prestamo_id = body.get('prestamo_id')
        detalles_recibidos = body.get('detalles_recibidos', [])
        detalles_perdidos = body.get('detalles_perdidos', [])
        detalles_no_recibidos = body.get('detalles_no_recibidos', [])
        # Comentarios por expediente: { detalle_id: "comentario" }
        comentarios_por_detalle = body.get('comentarios_por_detalle', {}) or {}
        notas = body.get('notas', '')

        def _comentario(det_id):
            v = comentarios_por_detalle.get(str(det_id)) or comentarios_por_detalle.get(det_id)
            return (v or '').strip() or None

        prestamo = Prestamo.objects.get(id=prestamo_id)
        solicitud = prestamo.solicitud

        # 1. Procesar los que llegaron (Disponibles)
        esta_vencido = prestamo.esta_vencido
        unidad_usuario = _get_unidad_usuario(request.user)

        # Resolver UNA sola vez la ubicación ADMISION (catálogo nuevo).
        # Al devolver, el expediente regresa a ADMISION.
        from expediente.services.ubicaciones import CatalogoUbicaciones
        ubicacion_admision = None
        try:
            ubicacion_admision = CatalogoUbicaciones.ubicacion_admision()
        except Exception as _e:
            logger.warning(f"No se pudo resolver ubicacion ADMISION: {_e}")

        for det_id in detalles_recibidos:
            detalle = SolicitudExpedienteDetalle.objects.get(id=det_id, solicitud=solicitud)
            if not detalle.devuelto:
                detalle.devuelto = True
                if esta_vencido:
                    detalle.fuera_de_tiempo = True
                detalle.comentario_devolucion = _comentario(det_id)
                detalle.save()

                ep = detalle.expediente_prestamo
                estado_ant = ep.estado
                ep.estado_id = 'EXP_DISPONIBLE'
                ep.ubicacion_fisica = unidad_usuario
                # NUEVO: regresar al catálogo unificado → ADMISION
                if ubicacion_admision is not None:
                    ep.ubicacion = ubicacion_admision
                ep.save()

                # LEGACY: mantener sincronizado expediente.localizacion (texto) → ADMISION
                try:
                    _set_localizacion_admision(ep.expediente, request.user)
                except Exception as _e:
                    logger.warning(f"No se pudo regresar a ADMISION (legacy) al devolver: {_e}")

                ExpedienteEstadoLog.objects.create(
                    expediente=ep.expediente,
                    estado_anterior=estado_ant,
                    estado_nuevo_id='EXP_DISPONIBLE',
                    usuario=request.user,
                    solicitud=solicitud,
                    observacion="Devuelto correctamente" + (" (Fuera de tiempo)" if esta_vencido else "")
                )

        # 2. Procesar los perdidos (Cuentan como procesados/cerrados para la solicitud)
        for det_id in detalles_perdidos:
            detalle = SolicitudExpedienteDetalle.objects.get(id=det_id, solicitud=solicitud)
            if not detalle.devuelto:
                detalle.devuelto = True  # Se marca como procesado
                detalle.comentario_devolucion = _comentario(det_id) or 'Marcado como perdido'
                detalle.save()
                
                ep = detalle.expediente_prestamo
                estado_ant = ep.estado
                ep.estado_id = 'EXP_PERDIDO'
                ep.save()
                
                ExpedienteEstadoLog.objects.create(
                    expediente=ep.expediente,
                    estado_anterior=estado_ant,
                    estado_nuevo_id='EXP_PERDIDO',
                    usuario=request.user,
                    solicitud=solicitud,
                    observacion="Marcado como perdido durante auditoría"
                )

        # 3. Procesar los NO recibidos (Siguen pendientes)
        for det_id in detalles_no_recibidos:
            detalle = SolicitudExpedienteDetalle.objects.get(id=det_id, solicitud=solicitud)
            ep = detalle.expediente_prestamo
            ExpedienteEstadoLog.objects.create(
                expediente=ep.expediente,
                estado_anterior=ep.estado,
                estado_nuevo_id='EXP_PRESTADO',
                usuario=request.user,
                solicitud=solicitud,
                observacion="Auditado como NO RECIBIDO (Sigue en préstamo)"
            )

        # 4. Determinar estado final de la solicitud (solo expedientes aprobados cuentan)
        total_exp = solicitud.detalles.filter(aprobado=True).count()
        devueltos_ahora = solicitud.detalles.filter(aprobado=True, devuelto=True).count()
        hay_no_recibidos = len(detalles_no_recibidos) > 0
        
        if devueltos_ahora >= total_exp and not hay_no_recibidos:
            # Todo procesado: ver si fue vencido
            if prestamo.esta_vencido:
                solicitud.estado_flujo_id = 'SOL_FINALIZADA'
                prestamo.estado = 'DevueltoVencido'
            else:
                solicitud.estado_flujo_id = 'SOL_FINALIZADA'
                prestamo.estado = 'Cerrado'
            prestamo.fecha_devolucion_real = timezone.now()
        else:
            # Faltan expedientes (no_recibidos pendientes)
            solicitud.estado_flujo_id = 'SOL_INCOMPLETA'
            prestamo.estado = 'DevolucionParcial'
            
        solicitud.save()
        prestamo.save()

        # Registrar devolución parcial
        Devolucion.objects.create(
            prestamo=prestamo,
            cantidad_esperada=total_exp,
            cantidad_recibida=devueltos_ahora,
            estado='Completa' if devueltos_ahora >= total_exp and not hay_no_recibidos else 'Incompleta',
            notas_admin=notas
        )

        # Registrar log de auditoría — esto también dispara el changes-check
        # para que las pantallas conectadas (Mis Solicitudes, etc.) detecten
        # el cambio y se actualicen automáticamente.
        descripcion_log = (
            f"Auditoría de devolución del préstamo #{prestamo.id}: "
            f"{devueltos_ahora}/{total_exp} expedientes devueltos. "
            f"Estado solicitud: {solicitud.estado_flujo_id}."
        )
        _registrar_log(
            request.user,
            'DEVOLUCION_PROCESADA',
            descripcion_log,
            'Prestamo', prestamo.id
        )

        return JsonResponse({"success": True, "estado": solicitud.estado_flujo_id})

    except Exception as e:
        logger.error(f"Error en procesar_devolucion_api: {e}", exc_info=True)
        return JsonResponse({"error": "Error interno del servidor"}, status=500)

@require_GET
def buscar_expedientes_api(request):
    """
    Buscador principal de expedientes con filtrado de disponibilidad en tiempo real.
    
    Query Params:
        q (str): Término de búsqueda (número, identidad o nombre).
        tipo (str): 'expediente', 'identidad' o 'nombre'.
        
    Returns:
        JSON con los resultados y flags de 'disponible' (True/False).
    """
    """
    Busca pacientes en la base SIWI por identidad, N° expediente o nombre.
    Todos los expedientes están disponibles excepto aquellos con préstamo activo.
    """
    if not _es_exp_solicitante(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)

    try:
        query = request.GET.get('q', '').strip()
        tipo = request.GET.get('tipo', 'expediente')  # expediente, identidad, nombre

        if not query:
            return JsonResponse({"data": []})

        # IDs de expedientes no disponibles (Cualquier estado que no sea disponible)
        expedientes_no_disponibles = set(
            ExpedientePrestamo.objects.exclude(estado_id='EXP_DISPONIBLE')
            .values_list('expediente_id', flat=True)
        )
        # También las solicitudes activas que podrían no haber actualizado el estado físico aún.
        # IMPORTANTE: solo cuentan los detalles APROBADOS — los rechazados ya no apartan al expediente.
        en_proceso = set(
            SolicitudExpedienteDetalle.objects.filter(
                solicitud__estado_flujo_id__in=['SOL_PENDIENTE', 'SOL_APROBADA_ORGANIZANDO', 'SOL_LISTO_RECOGER', 'SOL_EN_PRESTAMO', 'SOL_EN_DEVOLUCION', 'SOL_INCOMPLETA'],
                devuelto=False,
                aprobado=True,
            ).values_list('expediente_prestamo__expediente_id', flat=True)
        )
        expedientes_prestados_ids = expedientes_no_disponibles | en_proceso

        # Usamos el servicio DatosPaciente para formatear nombres de manera consistente.
        # IMPORTANTE: aquí SÍ devolvemos paciente_id porque el frontend lo guarda
        # en el carrito y lo envía al backend al crear la solicitud (ya no captura
        # texto, solo IDs).
        from s_exp.services.datos_solicitud import DatosPaciente

        # Ubicaciones que cuentan como "en el archivo / disponible para prestar".
        # Regla: un expediente solo puede prestarse si está físicamente en ADMISION.
        # Durante la transición híbrida también aceptamos ARCHIVO (legacy).
        # A futuro, dejar únicamente 'ADMISION'.
        UBICACIONES_DISPONIBLES = {'ADMISION', 'ARCHIVO'}

        def _construir_resultado(exp, paciente):
            """Helper interno: dado un Expediente y un Paciente, arma el dict de respuesta."""
            # Precargamos la ubicación (catálogo nuevo) y sus relaciones para
            # que _resolver_ubicacion_expediente no dispare queries extra.
            info_exp = ExpedientePrestamo.objects.select_related(
                'ubicacion__unidad_clinica__area_atencion__servicio',
                'ubicacion__unidad_clinica__sala',
                'ubicacion__unidad_clinica__servicio_aux',
                'ubicacion__unidad_no_clinica',
            ).filter(expediente=exp).first()

            ubicacion_texto = _resolver_ubicacion_expediente(exp, info_exp)

            # DISPONIBLE requiere DOS condiciones:
            #   1. No estar prestado ni en un proceso de solicitud activo.
            #   2. Estar físicamente en ADMISION (o ARCHIVO durante el híbrido).
            #      Si está en un área clínica (atención/ingreso), NO se presta.
            ubic_upper = (ubicacion_texto or '').strip().upper()
            en_ubicacion_disponible = ubic_upper in UBICACIONES_DISPONIBLES
            disponible = (exp.id not in expedientes_prestados_ids) and en_ubicacion_disponible

            return {
                "expediente_id": exp.id,
                "numero_expediente": exp.numero,
                "paciente_id": paciente.id if paciente else None,
                "paciente_nombre": DatosPaciente.nombre_completo(paciente),
                "paciente_dni": DatosPaciente.dni(paciente),
                "disponible": disponible,
                "ubicacion_fisica": ubicacion_texto,
            }

        resultados = []

        if tipo == 'expediente':
            # Buscar por número de expediente — Top 5 resultados
            expedientes_encontrados = set()
            try:
                numero_int = int(query.lstrip("0") or "0")
                for exp in Expediente.objects.filter(numero=numero_int)[:5]:
                    expedientes_encontrados.add(exp.id)
            except ValueError:
                pass

            # También permite buscar via Paciente.expediente_numero (vínculo viejo)
            pacientes_por_exp = Paciente.objects.filter(expediente_numero__icontains=query)[:5]
            for pac in pacientes_por_exp:
                asig = PacienteAsignacion.objects.filter(paciente=pac).select_related('expediente').first()
                if asig:
                    expedientes_encontrados.add(asig.expediente.id)

            for exp in Expediente.objects.filter(id__in=expedientes_encontrados)[:5]:
                asignacion = PacienteAsignacion.objects.filter(
                    expediente=exp
                ).select_related('paciente').order_by('-estado').first()
                paciente = asignacion.paciente if asignacion else None
                resultados.append(_construir_resultado(exp, paciente))

        elif tipo == 'identidad':
            query_limpio = query.replace("-", "").replace(" ", "")
            pacientes = Paciente.objects.filter(dni__icontains=query_limpio, estado='A')[:5]
            for pac in pacientes:
                asignaciones = PacienteAsignacion.objects.filter(
                    paciente=pac, estado='1'
                ).select_related('expediente')
                for asig in asignaciones:
                    resultados.append(_construir_resultado(asig.expediente, pac))
                    if len(resultados) >= 5:
                        break
                if len(resultados) >= 5:
                    break

        elif tipo == 'nombre':
            palabras = query.split()
            filtro = Q(estado='A')
            for palabra in palabras:
                filtro &= (
                    Q(primer_nombre__icontains=palabra) |
                    Q(segundo_nombre__icontains=palabra) |
                    Q(primer_apellido__icontains=palabra) |
                    Q(segundo_apellido__icontains=palabra)
                )
            pacientes = Paciente.objects.filter(filtro)[:5]
            for pac in pacientes:
                asignaciones = PacienteAsignacion.objects.filter(
                    paciente=pac, estado='1'
                ).select_related('expediente')
                for asig in asignaciones:
                    resultados.append(_construir_resultado(asig.expediente, pac))
                    if len(resultados) >= 5:
                        break
                if len(resultados) >= 5:
                    break

        return JsonResponse({"data": resultados})

    except Exception as e:
        logger.error(f"Error en buscar_expedientes_api: {e}", exc_info=True)
        return JsonResponse({"error": "Error interno del servidor"}, status=500)


@csrf_protect
@require_POST
def crear_solicitud_api(request):
    """
    Crea una nueva solicitud de préstamo iniciada por un usuario del sistema.
    Verifica la disponibilidad física de los expedientes antes de permitir la creación.
    Asigna automáticamente la unidad de servicio del usuario desde su registro en RRHH.
    """
    if not _es_exp_solicitante(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Datos inválidos"}, status=400)

    expediente_ids = body.get('expedientes', [])  # lista de expediente IDs (de tabla Expediente)
    motivo_id = body.get('motivo_id')
    observaciones = body.get('observaciones', '').strip()
    tiempo_sugerido_horas = body.get('tiempo_sugerido_horas')

    if not expediente_ids:
        return JsonResponse({"error": "Debe seleccionar al menos un expediente"}, status=400)
    if not motivo_id:
        return JsonResponse({"error": "El motivo es obligatorio"}, status=400)

    # Validar tiempo sugerido (opcional). Mismo día: max horas hasta 4 PM. Días posteriores: max 72h.
    if tiempo_sugerido_horas is not None:
        try:
            tiempo_sugerido_horas = int(tiempo_sugerido_horas)
        except (TypeError, ValueError):
            return JsonResponse({"error": "Tiempo sugerido inválido"}, status=400)
        if tiempo_sugerido_horas < 1:
            return JsonResponse({"error": "El tiempo sugerido debe ser mayor a 0"}, status=400)
        if tiempo_sugerido_horas > 72:
            return JsonResponse({"error": "El tiempo sugerido no puede superar 72 horas"}, status=400)

    # Validar motivo
    try:
        motivo = MotivoSolicitud.objects.get(id=motivo_id, activo=True)
    except MotivoSolicitud.DoesNotExist:
        return JsonResponse({"error": "Motivo no válido"}, status=400)

    # Obtener unidad de servicio desde RRHH (fuente principal)
    servicio_unidad, es_registrado_rrhh = _get_servicio_unidad_from_rrhh(request.user)
    if not es_registrado_rrhh:
        return JsonResponse({
            "error": "El usuario no está registrado en el sistema RRHH (Recursos Humanos). Contacte al administrador."
        }, status=403)

    # Nota: ya no calculamos area_destino como texto. La unidad queda referenciada
    # por FK en `servicio_unidad`. Para mostrar el nombre, los servicios consultan
    # en vivo (ver s_exp/services/datos_solicitud.py → DatosSolicitud.unidad_nombre).

    try:

        # Verificar que existan y no estén prestados o en proceso
        prestados = set(
            ExpedientePrestamo.objects.filter(estado_id='EXP_PRESTADO')
            .values_list('expediente_id', flat=True)
        )
        en_proceso = set(
            SolicitudExpedienteDetalle.objects.filter(
                solicitud__estado_flujo_id__in=['SOL_PENDIENTE', 'SOL_APROBADA_ORGANIZANDO'],
                aprobado=True,
            ).values_list('expediente_prestamo__expediente_id', flat=True)
        )
        expedientes_prestados_ids = prestados | en_proceso

        expedientes = Expediente.objects.filter(id__in=expediente_ids)
        if expedientes.count() != len(expediente_ids):
            return JsonResponse({"error": "Algunos expedientes no fueron encontrados"}, status=400)

        for exp in expedientes:
            if exp.id in expedientes_prestados_ids:
                return JsonResponse({
                    "error": f"El expediente #{exp.numero} ya no está disponible"
                }, status=400)

        # Crear solicitud (sin snapshots de texto; solo referencia FK)
        solicitud = SolicitudPrestamo.objects.create(
            usuario=request.user,
            motivo=motivo,
            observaciones=observaciones or None,
            servicio_unidad=servicio_unidad,  # ubicación del solicitante (FK a servicio.Unidad)
            tiempo_sugerido_horas=tiempo_sugerido_horas,
        )

        # Crear detalles guardando SOLO el paciente_id (FK).
        # Los datos del paciente (DNI/nombre) se consultan en vivo cuando se
        # muestran, usando DatosDetalleSolicitud.
        for exp in expedientes:
            # Obtener o crear ExpedientePrestamo (estado físico actual)
            ep, created_ep = ExpedientePrestamo.objects.get_or_create(
                expediente=exp,
                defaults={'estado_id': 'EXP_APARTADO'}
            )
            if not created_ep:
                estado_anterior = ep.estado
                ep.estado_id = 'EXP_APARTADO'
                ep.save()
                ExpedienteEstadoLog.objects.create(
                    expediente=exp,
                    estado_anterior=estado_anterior,
                    estado_nuevo_id='EXP_APARTADO',
                    usuario=request.user,
                    solicitud=solicitud,
                    observacion="Expediente apartado por solicitud"
                )
            else:
                ExpedienteEstadoLog.objects.create(
                    expediente=exp,
                    estado_anterior=None,
                    estado_nuevo_id='EXP_APARTADO',
                    usuario=request.user,
                    solicitud=solicitud
                )

            # Buscar el paciente asignado AL MOMENTO de la solicitud.
            # Si después el expediente se reasigna a otro paciente, esta
            # solicitud conserva el paciente original via FK.
            asig = PacienteAsignacion.objects.filter(
                expediente=exp, estado='1'
            ).select_related('paciente').first()
            paciente_actual = asig.paciente if asig else None

            SolicitudExpedienteDetalle.objects.create(
                solicitud=solicitud,
                expediente_prestamo=ep,
                paciente=paciente_actual,  # FK al paciente (no snapshot)
            )

        _registrar_log(
            request.user, 'SOLICITUD_CREADA',
            f'Solicitud #{solicitud.id} creada con {expedientes.count()} expedientes.',
            'SolicitudPrestamo', solicitud.id
        )

        logger.info(f"Solicitud #{solicitud.id} creada por {request.user.username}")
        return JsonResponse({
            "success": True,
            "solicitud_id": solicitud.id,
            "mensaje": f"Solicitud #{solicitud.id} creada exitosamente con {expedientes.count()} expediente(s)."
        })

    except Exception as e:
        logger.error(f"Error en crear_solicitud_api: {e}", exc_info=True)
        return JsonResponse({"error": "Error interno del servidor"}, status=500)


# ============================================
# APIs USUARIO - Seguimiento
# ============================================

@require_GET
def mis_solicitudes_api(request):
    """
    Lista las solicitudes del usuario actual con filtros opcionales de fecha.
    
    Query Params:
        filtro (str): 'hoy', 'semana', 'mes', 'rango' o '' para todas.
        fecha_inicio (str): Fecha inicio en formato YYYY-MM-DD (aplica con filtro='rango').
        fecha_fin (str): Fecha fin en formato YYYY-MM-DD (aplica con filtro='rango').
    """
    if not _es_exp_solicitante(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)

    try:
        qs = SolicitudPrestamo.objects.filter(
            usuario=request.user
        ).select_related('servicio_unidad').order_by('-fecha_creacion')

        # --- Aplicar filtros de fecha (mismo patrón que reportes del módulo) ---
        filtro = request.GET.get('filtro', '').strip()
        from datetime import date as date_type
        hoy = date_type.today()

        if filtro == 'hoy':
            qs = qs.filter(
                fecha_creacion__gte=str(hoy),
                fecha_creacion__lte=str(hoy) + ' 23:59:59'
            )
        elif filtro == 'semana':
            inicio_semana = hoy - timedelta(days=hoy.weekday())  # Lunes
            fin_semana = inicio_semana + timedelta(days=6)        # Domingo
            qs = qs.filter(
                fecha_creacion__gte=str(inicio_semana),
                fecha_creacion__lte=str(fin_semana) + ' 23:59:59'
            )
        elif filtro == 'mes':
            import calendar
            ultimo_dia = calendar.monthrange(hoy.year, hoy.month)[1]
            inicio_mes = str(hoy.replace(day=1))
            fin_mes = str(hoy.replace(day=ultimo_dia))
            qs = qs.filter(
                fecha_creacion__gte=inicio_mes,
                fecha_creacion__lte=fin_mes + ' 23:59:59'
            )
        elif filtro == 'rango':
            fecha_inicio_str = request.GET.get('fecha_inicio', '').strip()
            fecha_fin_str = request.GET.get('fecha_fin', '').strip()
            if fecha_inicio_str:
                qs = qs.filter(fecha_creacion__gte=fecha_inicio_str)
            if fecha_fin_str:
                qs = qs.filter(fecha_creacion__lte=fecha_fin_str + ' 23:59:59')
        # Si filtro está vacío retorna todas las solicitudes
        data = []
        for s in qs:
            # Importamos los servicios para acceso unificado a datos
            from s_exp.services.datos_solicitud import DatosDetalleSolicitud, DatosSolicitud

            # Cada detalle se enriquece desde el FK paciente (no desde snapshots)
            detalles_info = []
            for d in s.detalles.select_related(
                'expediente_prestamo__expediente', 'paciente'
            ):
                detalles_info.append(DatosDetalleSolicitud.enriquecer(d))

            prestamo_info = None
            try:
                p = s.prestamo
                prestamo_info = {
                    "id": p.id,
                    "estado": p.estado,
                    "fecha_entrega": _fmt_local(p.fecha_entrega) or None,
                    "fecha_limite": p.fecha_limite.isoformat() if p.fecha_limite else None,
                    "tiempo_restante_segundos": p.tiempo_restante_segundos,
                    "porcentaje_tiempo_usado": p.porcentaje_tiempo_usado,
                    "esta_vencido": p.esta_vencido,
                    "motivo_rechazo": p.motivo_rechazo or "",
                    "comentarios": p.comentarios or "",
                }
            except Prestamo.DoesNotExist:
                pass

            data.append({
                "id": s.id,
                "fecha_creacion": _fmt_local(s.fecha_creacion),
                "estado_flujo": DatosSolicitud.estado_codigo(s),
                "estado_flujo_nombre": DatosSolicitud.estado_nombre(s),
                "motivo": DatosSolicitud.motivo_nombre(s),
                "observaciones": s.observaciones or "",
                "unidad": DatosSolicitud.unidad_nombre(s),
                "unidad_id": DatosSolicitud.unidad_id(s),
                # Alias para retrocompatibilidad con frontend actual
                "area_destino": DatosSolicitud.unidad_nombre(s),
                "expedientes": detalles_info,
                "cant_expedientes": len(detalles_info),
                "prestamo": prestamo_info,
            })

        return JsonResponse({"data": data})

    except Exception as e:
        logger.error(f"Error en mis_solicitudes_api: {e}", exc_info=True)
        return JsonResponse({"error": "Error interno del servidor"}, status=500)


# ============================================
# APIs - Alertas
# ============================================

@require_GET
def changes_check_api(request):
    """
    Endpoint ULTRA LIGERO usado por el polling inteligente del frontend.

    Devuelve los timestamps del último cambio en cada sección del módulo s_exp.
    El frontend compara estos timestamps con los últimos vistos y solo
    recarga las tablas si hubo un cambio real (preserva el estado de UI
    como tarjetas expandidas, scroll, etc).

    Es deliberadamente ligero: solo hace agregaciones MAX(timestamp) sin
    devolver datos grandes. Se llama cada 3-5s pero su carga es mínima.
    """
    # Validación de seguridad: solo usuarios autenticados (sin sesión = sin polling)
    if not request.user.is_authenticated:
        return JsonResponse({"error": "No autenticado"}, status=401)

    from django.db.models import Max
    from .models import LogHistorico, SolicitudPrestamo, Prestamo, Devolucion

    # GRANULARIDAD POR TIPO DE ACCIÓN:
    # Cada sección (pantalla del admin) solo se notifica de los eventos que
    # le incumben directamente. Esto evita que el admin reciba un banner en
    # "Gestión Solicitudes" cuando lo que cambió es una devolución, etc.
    user = request.user

    # ----- Eventos por sección (vía LogHistorico) -----
    # Gestión Solicitudes (admin) → nuevas solicitudes creadas por usuarios
    gestion_acciones = ['SOLICITUD_CREADA']

    # Control de Devoluciones (admin) → usuario pide devolver expedientes
    devoluciones_acciones = ['SOLICITUD_DEVOLUCION_INICIADA']

    # Monitoreo Préstamos (admin) → préstamos entregados/devueltos por OTROS admins
    # (no tiene un tipo de log "de usuario", solo cambios de estado por admins)
    monitoreo_acciones = ['PRESTAMO_ENTREGADO', 'DEVOLUCION_PROCESADA']

    # Mis Solicitudes (usuario) → cualquier cambio en sus propias solicitudes
    # hecho por un admin (aprobada, lista, etc).
    mis_solic_acciones = [
        'SOLICITUD_APROBADA', 'SOLICITUD_RECHAZADA', 'SOLICITUD_LISTA',
        'PRESTAMO_ENTREGADO', 'DEVOLUCION_PROCESADA',
    ]

    def _max_log(acciones, excluir_self=True):
        """MAX(timestamp) de logs filtrados por tipo. Opcional: excluir al user actual."""
        qs = LogHistorico.objects.filter(accion__in=acciones)
        if excluir_self:
            qs = qs.exclude(usuario=user)
        return qs.aggregate(ts=Max('timestamp'))['ts']

    # Gestión y Devoluciones: NO excluir al usuario (acciones de "usuario"
    # siempre deben notificar al admin, incluso si admin = solicitante en pruebas).
    gestion_ts = _max_log(gestion_acciones, excluir_self=False)
    devoluciones_ts = _max_log(devoluciones_acciones, excluir_self=False)

    # Monitoreo: SÍ excluir al usuario (acciones admin no deben auto-notificarse)
    monitoreo_ts = _max_log(monitoreo_acciones, excluir_self=True)

    # Mis Solicitudes: estos son cambios sobre las solicitudes del usuario.
    # No excluir nada (necesita ver TODO cambio que afecte sus solicitudes).
    mis_solic_ts = _max_log(mis_solic_acciones, excluir_self=False)

    # Global: cualquier log (excluyendo los del propio usuario para no spam)
    global_ts = LogHistorico.objects.exclude(usuario=user).aggregate(ts=Max('timestamp'))['ts']

    def _iso(dt):
        return dt.isoformat() if dt else ''

    return JsonResponse({
        # Fallback general (para Dashboard / Reportes)
        'global': _iso(global_ts),

        # Cada sección solo recibe sus eventos específicos:
        'solicitudes': _iso(gestion_ts),       # Gestión Solicitudes (admin)
        'mis_solicitudes': _iso(mis_solic_ts), # Mis Solicitudes (usuario)
        'prestamos': _iso(monitoreo_ts),       # Monitoreo
        'devoluciones': _iso(devoluciones_ts), # Control de Devoluciones
    })


@require_GET
def alertas_usuario_api(request):
    """Retorna alertas para el usuario actual."""
    if not request.user.is_authenticated:
        return JsonResponse({"alertas": []})

    try:
        alertas = []

        # Alertas para solicitantes: préstamos a punto de vencer
        prestamos_usuario = Prestamo.objects.filter(
            solicitud__usuario=request.user,
            estado='Entregado'
        )

        for p in prestamos_usuario:
            if p.esta_vencido:
                alertas.append({
                    "tipo": "danger",
                    "titulo": "Préstamo Vencido",
                    "mensaje": f"El préstamo #{p.id} ha superado el límite de tiempo. Devuelva los expedientes de inmediato.",
                    "prestamo_id": p.id,
                })
                continue

            # Minutos restantes para alertas de 10 / 5 min
            min_restantes = None
            if p.fecha_limite:
                min_restantes = int((p.fecha_limite - timezone.now()).total_seconds() // 60)

            if min_restantes is not None and 0 < min_restantes <= 5:
                alertas.append({
                    "tipo": "danger",
                    "titulo": "¡5 minutos para vencer!",
                    "mensaje": f"El préstamo #{p.id} vence en {min_restantes} minuto(s). Devuelva los expedientes ahora.",
                    "prestamo_id": p.id,
                    "sticky": True,
                })
            elif min_restantes is not None and 5 < min_restantes <= 10:
                alertas.append({
                    "tipo": "warning",
                    "titulo": "10 minutos para vencer",
                    "mensaje": f"El préstamo #{p.id} vence en {min_restantes} minuto(s). Prepare la devolución.",
                    "prestamo_id": p.id,
                })
            elif p.porcentaje_tiempo_usado >= 90:
                alertas.append({
                    "tipo": "warning",
                    "titulo": "Préstamo por Vencer",
                    "mensaje": f"El préstamo #{p.id} está próximo a vencer. Considere devolver los expedientes.",
                    "prestamo_id": p.id,
                })

        # Alertas de Vencimiento Recurrentes (Sticky cada 5 min)
        prestamos_vencidos = Prestamo.objects.filter(
            solicitud__usuario=request.user,
            estado='Vencido'
        )
        ahora = timezone.now()
        for p in prestamos_vencidos:
            reaparecer = False
            if not p.alerta_vencimiento_leida_at:
                reaparecer = True
            else:
                diferencia = ahora - p.alerta_vencimiento_leida_at
                if diferencia.total_seconds() > 300:  # 5 min
                    reaparecer = True
            
            if reaparecer:
                alertas.append({
                    "tipo": "danger",
                    "titulo": "¡PRÉSTAMO VENCIDO!",
                    "mensaje": f"El préstamo #{p.id} está vencido. Por favor devuelva los expedientes.",
                    "prestamo_id": p.id,
                    "sticky": True,
                    "tipo_alerta": "vencimiento"
                })

        # Solicitudes aprobadas listas para retirar (Persistentes hasta que el usuario las acepte)
        solicitudes_aprobadas = SolicitudPrestamo.objects.filter(
            usuario=request.user,
            estado_flujo_id='SOL_LISTO_RECOGER',
            notificado_listo=False
        )
        for s in solicitudes_aprobadas:
            alertas.append({
                "tipo": "success",
                "titulo": "¡Listo para recoger!",
                "mensaje": "Sus expedientes ya estan listos para recoger.",
                "solicitud_id": s.id,
                "sticky": True
            })

        # Solicitudes rechazadas recientes
        solicitudes_rechazadas = SolicitudPrestamo.objects.filter(

            usuario=request.user,
            estado_flujo_id='SOL_RECHAZADA'
        ).order_by('-fecha_creacion')[:5]
        for s in solicitudes_rechazadas:
            try:
                motivo = s.prestamo.motivo_rechazo or "Sin motivo especificado"
            except Prestamo.DoesNotExist:
                motivo = "Sin motivo especificado"
            alertas.append({
                "tipo": "danger",
                "titulo": "Solicitud Rechazada",
                "mensaje": f"Su solicitud #{s.id} fue rechazada. Motivo: {motivo}",
                "solicitud_id": s.id,
            })

        return JsonResponse({"alertas": alertas})

    except Exception as e:
        logger.error(f"Error en alertas_usuario_api: {e}", exc_info=True)
        return JsonResponse({"alertas": []})


@csrf_exempt
@require_POST
def marcar_notificacion_leida_api(request):
    """Marca una notificación de 'Listo para recoger' como leída por el usuario."""
    if not request.user.is_authenticated:
        return JsonResponse({"error": "No autenticado"}, status=401)

    try:
        import json
        body = json.loads(request.body)
        solicitud_id = body.get('solicitud_id')

        if not solicitud_id:
            return JsonResponse({"error": "Falta ID de solicitud"}, status=400)

        solicitud = SolicitudPrestamo.objects.get(id=solicitud_id, usuario=request.user)
        solicitud.notificado_listo = True
        solicitud.save()

        return JsonResponse({"success": True})

    except SolicitudPrestamo.DoesNotExist:
        return JsonResponse({"error": "Solicitud no encontrada"}, status=404)
    except Exception as e:
        logger.error(f"Error en marcar_notificacion_leida_api: {e}", exc_info=True)
        return JsonResponse({"error": "Error interno"}, status=500)


@csrf_exempt
@require_POST
def marcar_vencimiento_leido_api(request):
    """Marca una alerta de vencimiento como aceptada temporalmente (5 min)."""
    if not request.user.is_authenticated:
        return JsonResponse({"error": "No autenticado"}, status=401)

    try:
        import json
        body = json.loads(request.body)
        prestamo_id = body.get('prestamo_id')

        if not prestamo_id:
            return JsonResponse({"error": "Falta ID de préstamo"}, status=400)

        prestamo = Prestamo.objects.get(id=prestamo_id, solicitud__usuario=request.user)
        prestamo.alerta_vencimiento_leida_at = timezone.now()
        prestamo.save()

        return JsonResponse({"success": True})

    except Prestamo.DoesNotExist:
        return JsonResponse({"error": "Préstamo no encontrado"}, status=404)
    except Exception as e:
        logger.error(f"Error en marcar_vencimiento_leido_api: {e}", exc_info=True)
        return JsonResponse({"error": "Error interno"}, status=500)


# ============================================
# APIs - Reportes

@require_GET
def reportes_data_api(request):
    """Retorna datos completos para los reportes con filtros de fecha.
    Cuenta solicitudes reales (SolicitudPrestamo) en el período seleccionado.
    """
    if not _es_exp_admin(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)

    try:
        fecha_inicio = request.GET.get('fecha_inicio', '')
        fecha_fin = request.GET.get('fecha_fin', '')

        # Convertir a datetimes tz-aware para evitar RuntimeWarning
        from datetime import datetime, time as _dtime
        dt_ini = dt_fin = None
        if fecha_inicio:
            try:
                dt_ini = timezone.make_aware(
                    datetime.combine(datetime.strptime(fecha_inicio, '%Y-%m-%d').date(), _dtime.min)
                )
            except (ValueError, TypeError):
                dt_ini = None
        if fecha_fin:
            try:
                dt_fin = timezone.make_aware(
                    datetime.combine(datetime.strptime(fecha_fin, '%Y-%m-%d').date(), _dtime.max)
                )
            except (ValueError, TypeError):
                dt_fin = None

        # Filtros base sobre SolicitudPrestamo.fecha_creacion
        sol_filtros = {}
        if dt_ini:
            sol_filtros['fecha_creacion__gte'] = dt_ini
        if dt_fin:
            sol_filtros['fecha_creacion__lte'] = dt_fin

        qs_solicitudes = SolicitudPrestamo.objects.filter(**sol_filtros)

        # --- RESUMEN GENERAL ---
        total_solicitudes = qs_solicitudes.count()
        total_expedientes_solicitados = SolicitudExpedienteDetalle.objects.filter(
            solicitud__in=qs_solicitudes
        ).count()
        total_aprobadas = qs_solicitudes.filter(
            estado_flujo_id__in=['SOL_APROBADA_ORGANIZANDO', 'SOL_LISTO_RECOGER',
                                 'SOL_EN_PRESTAMO', 'SOL_EN_DEVOLUCION',
                                 'SOL_FINALIZADA', 'SOL_INCOMPLETA']
        ).count()
        total_rechazadas = qs_solicitudes.filter(
            estado_flujo_id='SOL_RECHAZADA'
        ).count()
        total_pendientes = qs_solicitudes.filter(
            estado_flujo_id='SOL_PENDIENTE'
        ).count()

        # --- DEMANDA POR ÁREA ---
        # La unidad ahora es relacional (servicio_unidad FK). Agrupamos por el
        # nombre de la unidad consultado en vivo, en lugar del antiguo texto
        # area_destino (eliminado en el refactor relacional).
        demanda_area = list(
            qs_solicitudes.values(
                area_destino=F('servicio_unidad__nombre_unidad')
            ).annotate(
                total=Count('id')
            ).order_by('-total')
        )

        # --- MOTIVOS DE USO ---
        motivos = list(
            qs_solicitudes.values(nombre=F('motivo__nombre')).annotate(
                total=Count('id')
            ).order_by('-total')[:10]
        )

        # --- EXPEDIENTE MÁS SOLICITADO ---
        expedientes_top = list(
            SolicitudExpedienteDetalle.objects.filter(
                solicitud__in=qs_solicitudes
            ).values(
                numero=F('expediente_prestamo__expediente__numero')
            ).annotate(
                total=Count('id')
            ).order_by('-total')[:10]
        )

        # --- USUARIOS CON MÁS SOLICITUDES ---
        usuarios_top = list(
            qs_solicitudes.values(
                username=F('usuario__username'),
                nombre_completo=F('usuario__first_name'),
            ).annotate(
                total=Count('id')
            ).order_by('-total')[:10]
        )
        # Construir nombre completo
        for u in usuarios_top:
            u['nombre'] = u.pop('nombre_completo', '') or u['username']

        # --- RECHAZOS CON DETALLE ---
        rechazos_qs = qs_solicitudes.filter(
            estado_flujo_id='SOL_RECHAZADA'
        ).select_related('usuario')
        rechazos = []
        for s in rechazos_qs:
            try:
                motivo_r = s.prestamo.motivo_rechazo or ""
            except Prestamo.DoesNotExist:
                motivo_r = ""
            rechazos.append({
                "solicitud_id": s.id,
                "usuario": s.usuario.username,
                "fecha": _fmt_local(s.fecha_creacion),
                "motivo_rechazo": motivo_r,
            })

        # --- MOROSIDAD (préstamos vencidos activos) ---
        ahora = timezone.now()
        filtros_prestamo = {}
        if dt_ini:
            filtros_prestamo['fecha_aprobacion__gte'] = dt_ini
        if dt_fin:
            filtros_prestamo['fecha_aprobacion__lte'] = dt_fin

        morosos = Prestamo.objects.filter(
            estado__in=['Entregado', 'Vencido'],
            fecha_limite__lt=ahora,
            **filtros_prestamo
        ).select_related('solicitud__usuario', 'solicitud__servicio_unidad')

        from s_exp.services.datos_solicitud import DatosSolicitud
        morosidad = []
        for p in morosos:
            morosidad.append({
                "prestamo_id": p.id,
                "usuario": DatosSolicitud.usuario_username(p.solicitud),
                # 'area' viene de la FK unidad (deprecado area_destino texto)
                "area": DatosSolicitud.unidad_nombre(p.solicitud),
                "fecha_limite": _fmt_local(p.fecha_limite),
                "dias_vencido": (ahora - p.fecha_limite).days if p.fecha_limite else 0,
            })

        # --- INCONSISTENCIAS (devoluciones parciales) ---
        parciales = Prestamo.objects.filter(
            estado='DevolucionParcial',
            **filtros_prestamo
        ).select_related('solicitud__usuario')

        inconsistencias = []
        for p in parciales:
            total_exp = p.solicitud.detalles.count()
            devueltos = sum(d.cantidad_recibida for d in p.devoluciones.all())
            inconsistencias.append({
                "prestamo_id": p.id,
                "usuario": p.solicitud.usuario.username,
                "total_expedientes": total_exp,
                "devueltos": devueltos,
                "faltantes": total_exp - devueltos,
            })

        return JsonResponse({
            "resumen": {
                "total_solicitudes": total_solicitudes,
                "total_expedientes": total_expedientes_solicitados,
                "aprobadas": total_aprobadas,
                "rechazadas": total_rechazadas,
                "pendientes": total_pendientes,
            },
            "demanda_area": demanda_area,
            "motivos": motivos,
            "expedientes_top": expedientes_top,
            "usuarios_top": usuarios_top,
            "rechazos": rechazos,
            "morosidad": morosidad,
            "inconsistencias": inconsistencias,
        })

    except Exception as e:
        logger.error(f"Error en reportes_data_api: {e}", exc_info=True)
        return JsonResponse({"error": "Error interno del servidor"}, status=500)



# ============================================
# API: Catálogo de Motivos
# ============================================

@require_GET
def motivos_api(request):
    """Retorna la lista de motivos activos para el dropdown."""
    if not request.user.is_authenticated:
        return JsonResponse({"error": "No autenticado"}, status=401)

    motivos = MotivoSolicitud.objects.filter(activo=True).order_by('nombre')
    data = [{"id": m.id, "nombre": m.nombre} for m in motivos]
    return JsonResponse({"data": data})


# ============================================
# API: Info del usuario (unidad)
# ============================================

@require_GET
def info_usuario_api(request):
    """Retorna información del usuario para el formulario de solicitud."""
    if not request.user.is_authenticated:
        return JsonResponse({"error": "No autenticado"}, status=401)

    unidad = _get_unidad_usuario(request.user)
    return JsonResponse({
        "unidad": unidad,
        "es_admin": _es_exp_admin(request.user),
    })


# ============================================
# API: Historial de préstamos por paciente
# ============================================

@require_GET
def historial_prestamos_paciente_api(request, paciente_id):
    """Retorna el historial de préstamos asociados a un paciente."""
    if not request.user.is_authenticated:
        return JsonResponse({"error": "No autenticado"}, status=401)

    try:
        from expediente.models import PacienteAsignacion

        # Obtener expedientes del paciente
        asignaciones = PacienteAsignacion.objects.filter(
            paciente_id=paciente_id
        ).select_related('expediente')

        expediente_ids = [a.expediente_id for a in asignaciones]

        if not expediente_ids:
            return JsonResponse({"data": [], "en_prestamo": False})

        # Buscar detalles de solicitud que involucren esos expedientes
        detalles = SolicitudExpedienteDetalle.objects.filter(
            expediente_prestamo__expediente_id__in=expediente_ids
        ).select_related(
            'solicitud', 'solicitud__usuario', 'solicitud__motivo',
            'solicitud__estado_flujo', 'solicitud__servicio_unidad',
            'expediente_prestamo', 'expediente_prestamo__expediente', 'paciente'
        ).order_by('-solicitud__fecha_creacion')

        # Servicios para acceso unificado (sin snapshots)
        from s_exp.services.datos_solicitud import DatosDetalleSolicitud, DatosSolicitud

        data = []
        en_prestamo_actual = False

        for d in detalles:
            s = d.solicitud
            estado = s.estado_flujo_id
            if estado in ('SOL_EN_PRESTAMO', 'SOL_APROBADA_ORGANIZANDO') and not d.devuelto:
                en_prestamo_actual = True

            data.append({
                "numero_expediente": DatosDetalleSolicitud.numero_expediente(d),
                "fecha_solicitud": _fmt_local(s.fecha_creacion),
                "motivo": DatosSolicitud.motivo_nombre(s),
                "solicitante": DatosSolicitud.usuario_nombre_completo(s),
                "estado": DatosSolicitud.estado_nombre(s),
                "devuelto": d.devuelto,
                # alias de retrocompat (frontend espera 'area_destino')
                "area_destino": DatosSolicitud.unidad_nombre(s),
            })

        return JsonResponse({"data": data, "en_prestamo": en_prestamo_actual})

    except Exception as e:
        logger.error(f"Error en historial_prestamos_paciente_api: {e}", exc_info=True)
        return JsonResponse({"error": "Error interno del servidor"}, status=500)


@require_GET
def historial_prestamos_expediente_api(request, expediente_id):
    """Retorna el historial de préstamos asociados a un expediente."""
    if not request.user.is_authenticated:
        return JsonResponse({"error": "No autenticado"}, status=401)

    try:
        # Buscar detalles de solicitud que involucren ese expediente
        detalles = SolicitudExpedienteDetalle.objects.filter(
            expediente_prestamo__expediente_id=expediente_id
        ).select_related(
            'solicitud', 'solicitud__usuario', 'solicitud__motivo',
            'solicitud__estado_flujo', 'solicitud__servicio_unidad',
            'expediente_prestamo', 'expediente_prestamo__expediente', 'paciente'
        ).order_by('-solicitud__fecha_creacion')

        from s_exp.services.datos_solicitud import DatosDetalleSolicitud, DatosSolicitud

        data = []
        en_prestamo_actual = False

        for d in detalles:
            s = d.solicitud
            estado = s.estado_flujo_id
            if estado in ('SOL_EN_PRESTAMO', 'SOL_APROBADA_ORGANIZANDO') and not d.devuelto:
                en_prestamo_actual = True

            data.append({
                "numero_expediente": DatosDetalleSolicitud.numero_expediente(d),
                "fecha_solicitud": _fmt_local(s.fecha_creacion),
                "motivo": DatosSolicitud.motivo_nombre(s),
                "solicitante": DatosSolicitud.usuario_nombre_completo(s),
                "estado": DatosSolicitud.estado_nombre(s),
                "devuelto": d.devuelto,
                "area_destino": DatosSolicitud.unidad_nombre(s),
            })

        return JsonResponse({"data": data, "en_prestamo": en_prestamo_actual})

    except Exception as e:
        logger.error(f"Error en historial_prestamos_expediente_api: {e}", exc_info=True)
        return JsonResponse({"error": "Error interno del servidor"}, status=500)


# ============================================
# HISTORIAL DE SOLICITUDES (Admin)
# ============================================

class HistorialSolicitudesView(SExpAdminMixin, TemplateView):
    template_name = 's_exp/historial_solicitudes.html'


@require_GET
def historial_solicitudes_api(request):
    """Lista todas las solicitudes (historico) con paginación server-side."""
    if not _es_exp_admin(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)

    try:
        draw = int(request.GET.get('draw', 0))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 25))
        search_value = request.GET.get('search[value]', '').strip()
        estado_filtro = request.GET.get('estado', '')

        qs = SolicitudPrestamo.objects.select_related(
            'usuario', 'estado_flujo', 'motivo'
        ).annotate(cant_exp=Count('detalles'))

        if estado_filtro:
            qs = qs.filter(estado_flujo_id=estado_filtro)

        if search_value:
            qs = qs.filter(
                Q(usuario__username__icontains=search_value) |
                Q(usuario__first_name__icontains=search_value) |
                Q(usuario__last_name__icontains=search_value) |
                Q(id__icontains=search_value) |
                Q(motivo__nombre__icontains=search_value)
            )

        total_records = SolicitudPrestamo.objects.count()
        filtered_records = qs.count()
        solicitudes = qs.order_by('-fecha_creacion')[start:start + length]

        data = []
        for s in solicitudes:
            numeros = list(
                s.detalles.values_list('expediente_prestamo__expediente__numero', flat=True)
            )
            # Eventos resumen (incompleta, devuelto fuera de tiempo)
            evento = None
            prestamo = s.prestamos.first()
            if s.estado_flujo_id == 'SOL_INCOMPLETA':
                faltantes = s.detalles.filter(devuelto=False).count()
                evento = f"⚠️ Incompleta: {faltantes} expediente(s) sin devolver"
            elif prestamo and prestamo.estado == 'DevueltoVencido':
                evento = "🕒 Devuelto fuera del tiempo acordado"
            elif s.estado_flujo_id == 'SOL_FINALIZADA':
                evento = "✅ Finalizada correctamente"

            from s_exp.services.datos_solicitud import DatosSolicitud
            data.append({
                "id": s.id,
                "usuario": DatosSolicitud.usuario_username(s),
                "usuario_nombre": DatosSolicitud.usuario_nombre_completo(s),
                "fecha_creacion": _fmt_local(s.fecha_creacion),
                "estado_flujo": DatosSolicitud.estado_codigo(s),
                "estado_flujo_nombre": DatosSolicitud.estado_nombre(s),
                "motivo": DatosSolicitud.motivo_nombre(s),
                "area_destino": DatosSolicitud.unidad_nombre(s),
                "expedientes": numeros,
                "evento_resumen": evento,
            })

        return JsonResponse({
            "draw": draw,
            "recordsTotal": total_records,
            "recordsFiltered": filtered_records,
            "data": data,
        })
    except Exception as e:
        logger.error(f"Error en historial_solicitudes_api: {e}", exc_info=True)
        return JsonResponse({"error": "Error interno del servidor"}, status=500)


@require_GET
def historial_solicitud_detalle_api(request, solicitud_id):
    """Retorna el detalle completo de una solicitud para el modal del historial."""
    if not _es_exp_admin(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)

    try:
        s = SolicitudPrestamo.objects.select_related(
            'usuario', 'estado_flujo', 'motivo', 'servicio_unidad'
        ).get(id=solicitud_id)

        from s_exp.services.datos_solicitud import DatosDetalleSolicitud, DatosSolicitud

        # Expedientes con estado físico actual (nombre del paciente vía FK)
        expedientes_data = []
        for d in s.detalles.select_related(
            'expediente_prestamo__expediente', 'expediente_prestamo__estado', 'paciente'
        ):
            ep = d.expediente_prestamo
            expedientes_data.append({
                "numero": DatosDetalleSolicitud.numero_expediente(d),
                "paciente": DatosDetalleSolicitud.paciente_nombre_completo(d),
                "estado_fisico": ep.estado.nombre if ep.estado else "—",
                "devuelto": d.devuelto,
            })

        # Logs de cambios de estado de expedientes en esta solicitud
        logs = ExpedienteEstadoLog.objects.filter(
            solicitud=s
        ).select_related('usuario', 'estado_anterior', 'estado_nuevo').order_by('fecha')

        logs_data = [{
            "fecha": _fmt_local(l.fecha),
            "accion": f"Exp #{l.expediente_id}: {l.estado_anterior.nombre if l.estado_anterior else '—'} → {l.estado_nuevo.nombre}",
            "usuario": l.usuario.username,
            "observacion": l.observacion or "",
        } for l in logs]

        prestamo = s.prestamos.first()
        return JsonResponse({"data": {
            "id": s.id,
            "usuario": DatosSolicitud.usuario_username(s),
            "usuario_nombre": DatosSolicitud.usuario_nombre_completo(s),
            "fecha_creacion": _fmt_local(s.fecha_creacion),
            "estado_flujo": DatosSolicitud.estado_codigo(s),
            "estado_flujo_nombre": DatosSolicitud.estado_nombre(s),
            "motivo": DatosSolicitud.motivo_nombre(s),
            "area_destino": DatosSolicitud.unidad_nombre(s),
            "expedientes": expedientes_data,
            "logs": logs_data,
            "prestamo": {"id": prestamo.id, "estado": prestamo.estado} if prestamo else None,
        }})
    except SolicitudPrestamo.DoesNotExist:
        return JsonResponse({"error": "Solicitud no encontrada"}, status=404)
    except Exception as e:
        logger.error(f"Error en historial_solicitud_detalle_api: {e}", exc_info=True)
        return JsonResponse({"error": "Error interno del servidor"}, status=500)


# ============================================
# EXPORTACIÓN DE REPORTES
# ============================================

def _obtener_datos_reporte_areas_motivos(fecha_inicio='', fecha_fin=''):
    """
    Construye una matriz de áreas (filas) x motivos (columnas) con conteos.
    Retorna: {
        'areas': ['Area1', 'Area2', ...],
        'motivos': ['Motivo1', 'Motivo2', ...],
        'datos': [[count, count, ...], ...],  // filas = áreas, columnas = motivos
        'totales_filas': [total_area1, total_area2, ...],
        'totales_columnas': [total_motivo1, total_motivo2, ...],
        'total_general': int
    }
    """
    # Filtrar solicitudes por rango de fechas (timezone-aware)
    from datetime import datetime, time as _dtime
    sol_filtros = {}
    if fecha_inicio:
        try:
            d_ini = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            dt_ini = timezone.make_aware(datetime.combine(d_ini, _dtime.min))
            sol_filtros['fecha_creacion__gte'] = dt_ini
        except (ValueError, TypeError):
            pass
    if fecha_fin:
        try:
            d_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            dt_fin = timezone.make_aware(datetime.combine(d_fin, _dtime.max))
            sol_filtros['fecha_creacion__lte'] = dt_fin
        except (ValueError, TypeError):
            pass

    qs_solicitudes = SolicitudPrestamo.objects.filter(**sol_filtros).select_related(
        'motivo', 'servicio_unidad'
    )

    # Obtener todas las áreas únicas. La unidad ahora es relacional
    # (servicio_unidad__nombre_unidad), ya no el texto area_destino eliminado.
    areas_raw = qs_solicitudes.values_list('servicio_unidad__nombre_unidad', flat=True).distinct()
    areas = sorted(set(a or 'Sin Área' for a in areas_raw))

    # Obtener todos los motivos únicos
    motivos_raw = qs_solicitudes.values_list('motivo__nombre', flat=True).distinct()
    motivos = sorted(set(m or 'Sin Motivo' for m in motivos_raw))

    # Construir matriz de conteos
    datos = {}
    for area in areas:
        datos[area] = {}
        for motivo in motivos:
            # Construir filtros de forma segura (vía relación servicio_unidad)
            filtros = {}
            if area == 'Sin Área':
                filtros['servicio_unidad__isnull'] = True
            else:
                filtros['servicio_unidad__nombre_unidad'] = area

            if motivo == 'Sin Motivo':
                filtros['motivo__isnull'] = True
            else:
                filtros['motivo__nombre'] = motivo

            count = qs_solicitudes.filter(**filtros).count()
            datos[area][motivo] = count

    # Construir filas de datos y calcular totales
    matriz_datos = []
    totales_filas = []
    for area in areas:
        fila = [datos[area].get(motivo, 0) for motivo in motivos]
        matriz_datos.append(fila)
        totales_filas.append(sum(fila))

    # Calcular totales por columna
    totales_columnas = []
    for i in range(len(motivos)):
        total = sum(fila[i] for fila in matriz_datos)
        totales_columnas.append(total)

    total_general = sum(totales_filas)

    return {
        'areas': areas,
        'motivos': motivos,
        'datos': matriz_datos,
        'totales_filas': totales_filas,
        'totales_columnas': totales_columnas,
        'total_general': total_general,
    }


def exportar_reporte_excel(request):
    """Exporta el reporte de áreas x motivos a Excel."""
    if not _es_exp_admin(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)

    if not Workbook:
        return JsonResponse({"error": "openpyxl no está instalado"}, status=400)

    try:
        fecha_inicio = request.GET.get('fecha_inicio', '')
        fecha_fin = request.GET.get('fecha_fin', '')

        # Obtener datos
        datos_reporte = _obtener_datos_reporte_areas_motivos(fecha_inicio, fecha_fin)

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Expedientes"

        # Estilos
        titulo_font = Font(name='Times New Roman', size=14, bold=True)
        encabezado_fill = PatternFill(start_color='008B8B', end_color='008B8B', fill_type='solid')
        encabezado_font = Font(bold=True, color='FFFFFF')
        total_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        total_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Título
        ws['A1'] = "REPORTE EXPEDIENTES PRESTADOS"
        ws['A1'].font = titulo_font
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = center_align

        # Rango de fechas
        fecha_texto = f"Período: {fecha_inicio or 'Inicio'} a {fecha_fin or 'Hoy'}"
        if not fecha_inicio and not fecha_fin:
            fecha_texto = "Período: Todos"
        ws['A2'] = fecha_texto
        ws.merge_cells('A2:F2')
        ws['A2'].alignment = center_align

        ws.append([])  # Espacio

        # Encabezados de tabla
        encabezados = ['Área'] + datos_reporte['motivos'] + ['TOTAL']
        ws.append(encabezados)

        header_row = ws.max_row
        for col in range(1, len(encabezados) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = encabezado_fill
            cell.font = encabezado_font
            cell.border = border
            cell.alignment = center_align

        # Datos
        for idx, area in enumerate(datos_reporte['areas']):
            fila = [area] + datos_reporte['datos'][idx] + [datos_reporte['totales_filas'][idx]]
            ws.append(fila)

            # Aplicar estilos a esta fila
            row_num = ws.max_row
            for col in range(1, len(fila) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.border = border
                cell.alignment = center_align if col > 1 else Alignment(horizontal='left', vertical='center')

        # Fila de totales
        totales_fila = ['TOTAL'] + datos_reporte['totales_columnas'] + [datos_reporte['total_general']]
        ws.append(totales_fila)

        totales_row = ws.max_row
        for col in range(1, len(totales_fila) + 1):
            cell = ws.cell(row=totales_row, column=col)
            cell.fill = total_fill
            cell.font = total_font
            cell.border = border
            cell.alignment = center_align if col > 1 else Alignment(horizontal='left', vertical='center')

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 20
        for i in range(2, len(encabezados) + 1):
            ws.column_dimensions[chr(64 + i)].width = 15

        # Guardar en memoria
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        # Retornar como descarga
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        tz = timezone.get_current_timezone()
        ts = timezone.now().astimezone(tz).strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="reporte_expedientes_prestados_{ts}.xlsx"'
        return response

    except Exception as e:
        logger.error(f"Error en exportar_reporte_excel: {e}", exc_info=True)
        return JsonResponse({"error": "Error al generar Excel"}, status=500)


def exportar_reporte_pdf(request):
    """Exporta el reporte de áreas x motivos a PDF con el mismo encabezado/pie del PDF de solicitudes."""
    if not _es_exp_admin(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)

    try:
        from reportlab.lib.units import inch
        from .services.pdf_solicitud_service import (
            IMG_GOB_SESAL, IMG_HEAC, IMG_FUNDAGES, IMG_SIWIH
        )

        fecha_inicio = request.GET.get('fecha_inicio', '')
        fecha_fin = request.GET.get('fecha_fin', '')

        # Obtener datos desde la BD
        datos_reporte = _obtener_datos_reporte_areas_motivos(fecha_inicio, fecha_fin)

        # Datos del usuario que genera el reporte.
        # _get_unidad_usuario ya hace la cascada PerfilUnidad → RRHH.
        user = request.user
        usuario_nombre = (f"{user.first_name} {user.last_name}".strip()) or user.username
        usuario_area = _get_unidad_usuario(user) or '—'

        # Tamaño de página: 8.5 x 13 pulgadas horizontal (13 ancho x 8.5 alto)
        page_size = (13 * inch, 8.5 * inch)
        margen_top = 3 * cm
        margen_bot = 2.5 * cm
        margen_lat = 1.5 * cm

        ahora = timezone.now()
        fecha_impresion = _fmt_local(ahora)  # 12h local

        buf = BytesIO()

        # Canvas personalizado para encabezado/pie con páginas numeradas
        class _PdfCanvas(rl_canvas.Canvas):
            def __init__(self, *args, draw_footer=None, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved_states = []
                self._draw_footer = draw_footer

            def showPage(self):
                self._saved_states.append(dict(self.__dict__))
                self._startPage()

            def save(self):
                total = len(self._saved_states)
                for state in self._saved_states:
                    self.__dict__.update(state)
                    if self._draw_footer:
                        self._draw_footer(self, total)
                    super().showPage()
                super().save()

        doc = BaseDocTemplate(
            buf,
            pagesize=page_size,
            leftMargin=margen_lat, rightMargin=margen_lat,
            topMargin=margen_top, bottomMargin=margen_bot,
            title='Reporte Expedientes Prestados',
        )

        frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='contenido')

        # HEADER - Mismo estilo que PDF de solicitudes (logos + título)
        def dibujar_header(canvas_obj, doc):
            canvas_obj.saveState()
            ancho, alto = doc.pagesize
            y_top = alto - 0.5 * cm

            # GOB_SESAL a la izquierda
            try:
                canvas_obj.drawImage(
                    IMG_GOB_SESAL, 0.5 * cm, y_top - 1.5 * cm,
                    width=6 * cm, height=1.5 * cm, preserveAspectRatio=True, mask='auto'
                )
            except Exception:
                pass

            # Texto centrado
            canvas_obj.setFont('Times-Bold', 11)
            canvas_obj.drawCentredString(
                ancho / 2, y_top - 0.75 * cm,
                'FUNDAGES - HOSPITAL DR. ENRIQUE AGUILAR CERRATO'
            )

            # Logos HEAC y FUNDAGES2 a la derecha
            try:
                canvas_obj.drawImage(
                    IMG_HEAC, ancho - 5 * cm, y_top - 2.0 * cm,
                    width=2.2 * cm, height=2.2 * cm, preserveAspectRatio=True, mask='auto'
                )
                canvas_obj.drawImage(
                    IMG_FUNDAGES, ancho - 2.5 * cm, y_top - 2.0 * cm,
                    width=2.2 * cm, height=2.2 * cm, preserveAspectRatio=True, mask='auto'
                )
            except Exception:
                pass

            canvas_obj.restoreState()

        # FOOTER - Mismo estilo que PDF de solicitudes
        def dibujar_footer(canvas_obj, total_pages):
            canvas_obj.saveState()
            ancho, alto = canvas_obj._pagesize
            y_bot = 1.2 * cm

            canvas_obj.setFont('Helvetica', 8)
            canvas_obj.setFillColor(colors.black)

            # Izquierda: fecha impresión
            canvas_obj.drawString(1.5 * cm, y_bot, f'Impreso: {fecha_impresion}')

            # Centro: página X de Y
            page_num = canvas_obj.getPageNumber()
            canvas_obj.drawCentredString(ancho / 2, y_bot, f'Página {page_num} de {total_pages}')

            # Derecha: SIWIH + logo
            try:
                canvas_obj.drawImage(
                    IMG_SIWIH, ancho - 3.3 * cm, y_bot - 0.1 * cm,
                    width=1.3 * cm, height=0.9 * cm, preserveAspectRatio=True, mask='auto'
                )
            except Exception:
                pass
            canvas_obj.setFont('Helvetica-Bold', 8)
            canvas_obj.drawRightString(ancho - 3.5 * cm, y_bot, 'SIWIH')

            canvas_obj.restoreState()

        doc.addPageTemplates([PageTemplate(id='main', frames=[frame], onPage=dibujar_header)])

        # Estilos tipográficos
        styles = getSampleStyleSheet()
        st_titulo = ParagraphStyle('titulo', parent=styles['Title'],
                                   fontName='Times-Bold', fontSize=16,
                                   alignment=TA_CENTER, spaceAfter=6,
                                   borderBottom=1, borderPadding=4)
        st_periodo = ParagraphStyle('periodo', parent=styles['Normal'],
                                    fontName='Helvetica', fontSize=11,
                                    alignment=TA_CENTER, spaceAfter=10)
        st_usuario_lbl = ParagraphStyle('usr_lbl', parent=styles['Normal'],
                                        fontName='Helvetica-Bold', fontSize=10,
                                        textColor=colors.HexColor('#006464'))
        st_usuario_val = ParagraphStyle('usr_val', parent=styles['Normal'],
                                        fontName='Helvetica', fontSize=10)
        st_tabla_head = ParagraphStyle('tabla_head', parent=styles['Normal'],
                                       fontName='Helvetica-Bold', fontSize=7,
                                       textColor=colors.white, alignment=TA_CENTER, leading=9)
        st_tabla_cell = ParagraphStyle('tabla_cell', parent=styles['Normal'],
                                       fontName='Helvetica', fontSize=10,
                                       alignment=TA_CENTER, leading=12)
        st_tabla_area = ParagraphStyle('tabla_area', parent=styles['Normal'],
                                       fontName='Helvetica-Bold', fontSize=8,
                                       alignment=TA_LEFT, leading=10)
        st_tabla_total = ParagraphStyle('tabla_total', parent=styles['Normal'],
                                        fontName='Helvetica-Bold', fontSize=10,
                                        alignment=TA_CENTER, leading=12)

        elementos = []

        # Título
        elementos.append(Paragraph('Reporte Expedientes Prestados', st_titulo))

        # Período
        fecha_texto = f"Período: del {fecha_inicio or 'inicio'} al {fecha_fin or 'hoy'}"
        elementos.append(Paragraph(fecha_texto, st_periodo))

        # Datos de usuario (quien genera el reporte)
        datos_usuario = [
            [Paragraph('Generado por:', st_usuario_lbl), Paragraph(usuario_nombre, st_usuario_val)],
            [Paragraph('Área:', st_usuario_lbl), Paragraph(usuario_area, st_usuario_val)],
        ]
        t_usuario = Table(datos_usuario, colWidths=[4 * cm, doc.width - 4 * cm])
        t_usuario.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
        ]))
        elementos.append(t_usuario)
        elementos.append(Spacer(1, 14))

        # Construir tabla: Áreas x Motivos (motivos en mayúsculas)
        encabezados = ['ÁREA'] + [str(m).upper() for m in datos_reporte['motivos']] + ['TOTAL']
        filas = [[Paragraph(str(h), st_tabla_head) for h in encabezados]]

        # Filas de datos
        for idx, area in enumerate(datos_reporte['areas']):
            fila = [Paragraph(str(area).upper(), st_tabla_area)]
            for col_idx in range(len(datos_reporte['motivos'])):
                count = datos_reporte['datos'][idx][col_idx]
                fila.append(Paragraph(str(count), st_tabla_cell))
            fila.append(Paragraph(str(datos_reporte['totales_filas'][idx]), st_tabla_total))
            filas.append(fila)

        # Fila de totales
        fila_total = [Paragraph('TOTAL', st_tabla_total)]
        for total_col in datos_reporte['totales_columnas']:
            fila_total.append(Paragraph(str(total_col), st_tabla_total))
        fila_total.append(Paragraph(str(datos_reporte['total_general']), st_tabla_total))
        filas.append(fila_total)

        # Anchos: Área 3cm, Total 2cm, motivos distribuyen el resto
        # Motivos con nombres largos (COMPLICACIONES..., INVESTIGACION) reciben
        # mayor peso para que la primera palabra quepa completa.
        num_motivos = len(datos_reporte['motivos'])
        area_w = 3 * cm
        total_w = 2 * cm
        disponible = doc.width - area_w - total_w

        def _peso_motivo(nombre):
            n = (nombre or '').upper()
            if 'COMPLICACION' in n or 'INVESTIGACI' in n:
                return 1.35
            return 1.0

        pesos = [_peso_motivo(m) for m in datos_reporte['motivos']]
        suma_pesos = sum(pesos) or 1
        motivo_widths = [disponible * (p / suma_pesos) for p in pesos]
        col_widths = [area_w] + motivo_widths + [total_w]

        tabla = Table(filas, colWidths=col_widths, repeatRows=1)

        tabla_styles = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#008b8b')),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#444444')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f1f5f5')]),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#d3d3d3')),
            ('BACKGROUND', (-1, 1), (-1, -2), colors.HexColor('#e8f4f4')),
        ]
        tabla.setStyle(TableStyle(tabla_styles))
        elementos.append(tabla)

        # Build PDF
        def make_canvas(*args, **kwargs):
            return _PdfCanvas(*args, draw_footer=dibujar_footer, **kwargs)

        doc.build(elementos, canvasmaker=make_canvas)

        pdf_bytes = buf.getvalue()
        buf.close()

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        tz = timezone.get_current_timezone()
        ts = timezone.now().astimezone(tz).strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="reporte_expedientes_prestados_{ts}.pdf"'
        return response

    except Exception as e:
        logger.error(f"Error en exportar_reporte_pdf: {e}", exc_info=True)
        return JsonResponse({"error": f"Error al generar PDF: {str(e)}"}, status=500)
