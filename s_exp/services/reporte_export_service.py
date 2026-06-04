"""
Servicio de exportación de reportes (módulo s_exp).
=============================================================================

Genera el reporte "Expedientes Prestados" (matriz de Áreas x Motivos) en
formato Excel (openpyxl) y PDF (reportlab). Las vistas solo delegan aquí, de
modo que toda la lógica pesada de armado de documentos vive fuera de views.py.

Notas de diseño:
  - El conteo se obtiene de forma RELACIONAL: las áreas salen de
    SolicitudPrestamo.servicio_unidad (FK), no de texto plano.
  - Las fechas/horas se muestran en hora local (UTC-6) con fmt_local; la BD
    sigue guardando en UTC como el resto del sistema.
  - Cada función pública recibe el `request`, valida permisos y devuelve un
    HttpResponse (descarga) o un JsonResponse de error.
"""
import logging
from datetime import datetime, time as _dtime
from io import BytesIO

from django.http import JsonResponse, HttpResponse
from django.utils import timezone

from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.platypus import (
    BaseDocTemplate, PageTemplate, Frame, Paragraph, Table, TableStyle, Spacer
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    Workbook = None

from s_exp.models import SolicitudPrestamo
from s_exp.services.permisos import es_exp_admin, get_unidad_usuario
from s_exp.services.formato import fmt_local

logger = logging.getLogger("s_exp")


def obtener_datos_reporte_areas_motivos(fecha_inicio='', fecha_fin=''):
    """
    Construye una matriz de áreas (filas) x motivos (columnas) con conteos.

    Retorna: {
        'areas': ['Area1', 'Area2', ...],
        'motivos': ['Motivo1', 'Motivo2', ...],
        'datos': [[count, count, ...], ...],  // filas = áreas, columnas = motivos
        'totales_filas': [total_area1, total_area2, ...],
        'totales_columnas': [total_motivo1, total_motivo2, ...],
        'total_general': int
    }
    """
    # Filtrar solicitudes por rango de fechas (timezone-aware)
    sol_filtros = {}
    if fecha_inicio:
        try:
            d_ini = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            dt_ini = timezone.make_aware(datetime.combine(d_ini, _dtime.min))
            sol_filtros['fecha_creacion__gte'] = dt_ini
        except (ValueError, TypeError):
            pass
    if fecha_fin:
        try:
            d_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            dt_fin = timezone.make_aware(datetime.combine(d_fin, _dtime.max))
            sol_filtros['fecha_creacion__lte'] = dt_fin
        except (ValueError, TypeError):
            pass

    qs_solicitudes = SolicitudPrestamo.objects.filter(**sol_filtros).select_related(
        'motivo', 'servicio_unidad'
    )

    # Obtener todas las áreas únicas. La unidad ahora es relacional
    # (servicio_unidad__nombre_unidad), ya no el texto area_destino eliminado.
    areas_raw = qs_solicitudes.values_list('servicio_unidad__nombre_unidad', flat=True).distinct()
    areas = sorted(set(a or 'Sin Área' for a in areas_raw))

    # Obtener todos los motivos únicos
    motivos_raw = qs_solicitudes.values_list('motivo__nombre', flat=True).distinct()
    motivos = sorted(set(m or 'Sin Motivo' for m in motivos_raw))

    # Construir matriz de conteos
    datos = {}
    for area in areas:
        datos[area] = {}
        for motivo in motivos:
            # Construir filtros de forma segura (vía relación servicio_unidad)
            filtros = {}
            if area == 'Sin Área':
                filtros['servicio_unidad__isnull'] = True
            else:
                filtros['servicio_unidad__nombre_unidad'] = area

            if motivo == 'Sin Motivo':
                filtros['motivo__isnull'] = True
            else:
                filtros['motivo__nombre'] = motivo

            count = qs_solicitudes.filter(**filtros).count()
            datos[area][motivo] = count

    # Construir filas de datos y calcular totales
    matriz_datos = []
    totales_filas = []
    for area in areas:
        fila = [datos[area].get(motivo, 0) for motivo in motivos]
        matriz_datos.append(fila)
        totales_filas.append(sum(fila))

    # Calcular totales por columna
    totales_columnas = []
    for i in range(len(motivos)):
        total = sum(fila[i] for fila in matriz_datos)
        totales_columnas.append(total)

    total_general = sum(totales_filas)

    return {
        'areas': areas,
        'motivos': motivos,
        'datos': matriz_datos,
        'totales_filas': totales_filas,
        'totales_columnas': totales_columnas,
        'total_general': total_general,
    }


def exportar_reporte_excel(request):
    """Exporta el reporte de áreas x motivos a Excel."""
    if not es_exp_admin(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)

    if not Workbook:
        return JsonResponse({"error": "openpyxl no está instalado"}, status=400)

    try:
        fecha_inicio = request.GET.get('fecha_inicio', '')
        fecha_fin = request.GET.get('fecha_fin', '')

        # Obtener datos
        datos_reporte = obtener_datos_reporte_areas_motivos(fecha_inicio, fecha_fin)

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Expedientes"

        # Estilos
        titulo_font = Font(name='Times New Roman', size=14, bold=True)
        encabezado_fill = PatternFill(start_color='008B8B', end_color='008B8B', fill_type='solid')
        encabezado_font = Font(bold=True, color='FFFFFF')
        total_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        total_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Título
        ws['A1'] = "REPORTE EXPEDIENTES PRESTADOS"
        ws['A1'].font = titulo_font
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = center_align

        # Rango de fechas
        fecha_texto = f"Período: {fecha_inicio or 'Inicio'} a {fecha_fin or 'Hoy'}"
        if not fecha_inicio and not fecha_fin:
            fecha_texto = "Período: Todos"
        ws['A2'] = fecha_texto
        ws.merge_cells('A2:F2')
        ws['A2'].alignment = center_align

        ws.append([])  # Espacio

        # Encabezados de tabla
        encabezados = ['Área'] + datos_reporte['motivos'] + ['TOTAL']
        ws.append(encabezados)

        header_row = ws.max_row
        for col in range(1, len(encabezados) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = encabezado_fill
            cell.font = encabezado_font
            cell.border = border
            cell.alignment = center_align

        # Datos
        for idx, area in enumerate(datos_reporte['areas']):
            fila = [area] + datos_reporte['datos'][idx] + [datos_reporte['totales_filas'][idx]]
            ws.append(fila)

            # Aplicar estilos a esta fila
            row_num = ws.max_row
            for col in range(1, len(fila) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.border = border
                cell.alignment = center_align if col > 1 else Alignment(horizontal='left', vertical='center')

        # Fila de totales
        totales_fila = ['TOTAL'] + datos_reporte['totales_columnas'] + [datos_reporte['total_general']]
        ws.append(totales_fila)

        totales_row = ws.max_row
        for col in range(1, len(totales_fila) + 1):
            cell = ws.cell(row=totales_row, column=col)
            cell.fill = total_fill
            cell.font = total_font
            cell.border = border
            cell.alignment = center_align if col > 1 else Alignment(horizontal='left', vertical='center')

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 20
        for i in range(2, len(encabezados) + 1):
            ws.column_dimensions[chr(64 + i)].width = 15

        # Guardar en memoria
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        # Retornar como descarga
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        tz = timezone.get_current_timezone()
        ts = timezone.now().astimezone(tz).strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="reporte_expedientes_prestados_{ts}.xlsx"'
        return response

    except Exception as e:
        logger.error(f"Error en exportar_reporte_excel: {e}", exc_info=True)
        return JsonResponse({"error": "Error al generar Excel"}, status=500)


def exportar_reporte_pdf(request):
    """Exporta el reporte de áreas x motivos a PDF con el mismo encabezado/pie del PDF de solicitudes."""
    if not es_exp_admin(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)

    try:
        from reportlab.lib.units import inch
        from .pdf_solicitud_service import (
            IMG_GOB_SESAL, IMG_HEAC, IMG_FUNDAGES, IMG_SIWIH
        )

        fecha_inicio = request.GET.get('fecha_inicio', '')
        fecha_fin = request.GET.get('fecha_fin', '')

        # Obtener datos desde la BD
        datos_reporte = obtener_datos_reporte_areas_motivos(fecha_inicio, fecha_fin)

        # Datos del usuario que genera el reporte.
        # get_unidad_usuario ya hace la cascada PerfilUnidad → RRHH.
        user = request.user
        usuario_nombre = (f"{user.first_name} {user.last_name}".strip()) or user.username
        usuario_area = get_unidad_usuario(user) or '—'

        # Tamaño de página: 8.5 x 13 pulgadas horizontal (13 ancho x 8.5 alto)
        page_size = (13 * inch, 8.5 * inch)
        margen_top = 3 * cm
        margen_bot = 2.5 * cm
        margen_lat = 1.5 * cm

        ahora = timezone.now()
        fecha_impresion = fmt_local(ahora)  # 24h local

        buf = BytesIO()

        # Canvas personalizado para encabezado/pie con páginas numeradas
        class _PdfCanvas(rl_canvas.Canvas):
            def __init__(self, *args, draw_footer=None, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved_states = []
                self._draw_footer = draw_footer

            def showPage(self):
                self._saved_states.append(dict(self.__dict__))
                self._startPage()

            def save(self):
                total = len(self._saved_states)
                for state in self._saved_states:
                    self.__dict__.update(state)
                    if self._draw_footer:
                        self._draw_footer(self, total)
                    super().showPage()
                super().save()

        doc = BaseDocTemplate(
            buf,
            pagesize=page_size,
            leftMargin=margen_lat, rightMargin=margen_lat,
            topMargin=margen_top, bottomMargin=margen_bot,
            title='Reporte Expedientes Prestados',
        )

        frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='contenido')

        # HEADER - Mismo estilo que PDF de solicitudes (logos + título)
        def dibujar_header(canvas_obj, doc):
            canvas_obj.saveState()
            ancho, alto = doc.pagesize
            y_top = alto - 0.5 * cm

            # GOB_SESAL a la izquierda
            try:
                canvas_obj.drawImage(
                    IMG_GOB_SESAL, 0.5 * cm, y_top - 1.5 * cm,
                    width=6 * cm, height=1.5 * cm, preserveAspectRatio=True, mask='auto'
                )
            except Exception:
                pass

            # Texto centrado
            canvas_obj.setFont('Times-Bold', 11)
            canvas_obj.drawCentredString(
                ancho / 2, y_top - 0.75 * cm,
                'FUNDAGES - HOSPITAL DR. ENRIQUE AGUILAR CERRATO'
            )

            # Logos HEAC y FUNDAGES2 a la derecha
            try:
                canvas_obj.drawImage(
                    IMG_HEAC, ancho - 5 * cm, y_top - 2.0 * cm,
                    width=2.2 * cm, height=2.2 * cm, preserveAspectRatio=True, mask='auto'
                )
                canvas_obj.drawImage(
                    IMG_FUNDAGES, ancho - 2.5 * cm, y_top - 2.0 * cm,
                    width=2.2 * cm, height=2.2 * cm, preserveAspectRatio=True, mask='auto'
                )
            except Exception:
                pass

            canvas_obj.restoreState()

        # FOOTER - Mismo estilo que PDF de solicitudes
        def dibujar_footer(canvas_obj, total_pages):
            canvas_obj.saveState()
            ancho, alto = canvas_obj._pagesize
            y_bot = 1.2 * cm

            canvas_obj.setFont('Helvetica', 8)
            canvas_obj.setFillColor(colors.black)

            # Izquierda: fecha impresión
            canvas_obj.drawString(1.5 * cm, y_bot, f'Impreso: {fecha_impresion}')

            # Centro: página X de Y
            page_num = canvas_obj.getPageNumber()
            canvas_obj.drawCentredString(ancho / 2, y_bot, f'Página {page_num} de {total_pages}')

            # Derecha: SIWIH + logo
            try:
                canvas_obj.drawImage(
                    IMG_SIWIH, ancho - 3.3 * cm, y_bot - 0.1 * cm,
                    width=1.3 * cm, height=0.9 * cm, preserveAspectRatio=True, mask='auto'
                )
            except Exception:
                pass
            canvas_obj.setFont('Helvetica-Bold', 8)
            canvas_obj.drawRightString(ancho - 3.5 * cm, y_bot, 'SIWIH')

            canvas_obj.restoreState()

        doc.addPageTemplates([PageTemplate(id='main', frames=[frame], onPage=dibujar_header)])

        # Estilos tipográficos
        styles = getSampleStyleSheet()
        st_titulo = ParagraphStyle('titulo', parent=styles['Title'],
                                   fontName='Times-Bold', fontSize=16,
                                   alignment=TA_CENTER, spaceAfter=6,
                                   borderBottom=1, borderPadding=4)
        st_periodo = ParagraphStyle('periodo', parent=styles['Normal'],
                                    fontName='Helvetica', fontSize=11,
                                    alignment=TA_CENTER, spaceAfter=10)
        st_usuario_lbl = ParagraphStyle('usr_lbl', parent=styles['Normal'],
                                        fontName='Helvetica-Bold', fontSize=10,
                                        textColor=colors.HexColor('#006464'))
        st_usuario_val = ParagraphStyle('usr_val', parent=styles['Normal'],
                                        fontName='Helvetica', fontSize=10)
        st_tabla_head = ParagraphStyle('tabla_head', parent=styles['Normal'],
                                       fontName='Helvetica-Bold', fontSize=7,
                                       textColor=colors.white, alignment=TA_CENTER, leading=9)
        st_tabla_cell = ParagraphStyle('tabla_cell', parent=styles['Normal'],
                                       fontName='Helvetica', fontSize=10,
                                       alignment=TA_CENTER, leading=12)
        st_tabla_area = ParagraphStyle('tabla_area', parent=styles['Normal'],
                                       fontName='Helvetica-Bold', fontSize=8,
                                       alignment=TA_LEFT, leading=10)
        st_tabla_total = ParagraphStyle('tabla_total', parent=styles['Normal'],
                                        fontName='Helvetica-Bold', fontSize=10,
                                        alignment=TA_CENTER, leading=12)

        elementos = []

        # Título
        elementos.append(Paragraph('Reporte Expedientes Prestados', st_titulo))

        # Período
        fecha_texto = f"Período: del {fecha_inicio or 'inicio'} al {fecha_fin or 'hoy'}"
        elementos.append(Paragraph(fecha_texto, st_periodo))

        # Datos de usuario (quien genera el reporte)
        datos_usuario = [
            [Paragraph('Generado por:', st_usuario_lbl), Paragraph(usuario_nombre, st_usuario_val)],
            [Paragraph('Área:', st_usuario_lbl), Paragraph(usuario_area, st_usuario_val)],
        ]
        t_usuario = Table(datos_usuario, colWidths=[4 * cm, doc.width - 4 * cm])
        t_usuario.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
        ]))
        elementos.append(t_usuario)
        elementos.append(Spacer(1, 14))

        # Construir tabla: Áreas x Motivos (motivos en mayúsculas)
        encabezados = ['ÁREA'] + [str(m).upper() for m in datos_reporte['motivos']] + ['TOTAL']
        filas = [[Paragraph(str(h), st_tabla_head) for h in encabezados]]

        # Filas de datos
        for idx, area in enumerate(datos_reporte['areas']):
            fila = [Paragraph(str(area).upper(), st_tabla_area)]
            for col_idx in range(len(datos_reporte['motivos'])):
                count = datos_reporte['datos'][idx][col_idx]
                fila.append(Paragraph(str(count), st_tabla_cell))
            fila.append(Paragraph(str(datos_reporte['totales_filas'][idx]), st_tabla_total))
            filas.append(fila)

        # Fila de totales
        fila_total = [Paragraph('TOTAL', st_tabla_total)]
        for total_col in datos_reporte['totales_columnas']:
            fila_total.append(Paragraph(str(total_col), st_tabla_total))
        fila_total.append(Paragraph(str(datos_reporte['total_general']), st_tabla_total))
        filas.append(fila_total)

        # Anchos: Área 3cm, Total 2cm, motivos distribuyen el resto
        # Motivos con nombres largos (COMPLICACIONES..., INVESTIGACION) reciben
        # mayor peso para que la primera palabra quepa completa.
        num_motivos = len(datos_reporte['motivos'])
        area_w = 3 * cm
        total_w = 2 * cm
        disponible = doc.width - area_w - total_w

        def _peso_motivo(nombre):
            n = (nombre or '').upper()
            if 'COMPLICACION' in n or 'INVESTIGACI' in n:
                return 1.35
            return 1.0

        pesos = [_peso_motivo(m) for m in datos_reporte['motivos']]
        suma_pesos = sum(pesos) or 1
        motivo_widths = [disponible * (p / suma_pesos) for p in pesos]
        col_widths = [area_w] + motivo_widths + [total_w]

        tabla = Table(filas, colWidths=col_widths, repeatRows=1)

        tabla_styles = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#008b8b')),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#444444')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f1f5f5')]),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#d3d3d3')),
            ('BACKGROUND', (-1, 1), (-1, -2), colors.HexColor('#e8f4f4')),
        ]
        tabla.setStyle(TableStyle(tabla_styles))
        elementos.append(tabla)

        # Build PDF
        def make_canvas(*args, **kwargs):
            return _PdfCanvas(*args, draw_footer=dibujar_footer, **kwargs)

        doc.build(elementos, canvasmaker=make_canvas)

        pdf_bytes = buf.getvalue()
        buf.close()

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        tz = timezone.get_current_timezone()
        ts = timezone.now().astimezone(tz).strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="reporte_expedientes_prestados_{ts}.pdf"'
        return response

    except Exception as e:
        logger.error(f"Error en exportar_reporte_pdf: {e}", exc_info=True)
        return JsonResponse({"error": f"Error al generar PDF: {str(e)}"}, status=500)
