from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image
from django.http import HttpResponse
from django.utils import timezone
from PIL import Image as PILImage
from django.db import connections
from core.utils.utilidades_fechas import mes_nombre, formatear_fecha
from core.constants.domain_constants import LogApp
from core.utils.utilidades_logging import *

class ServiceExcel:
    """
    Servicio de generación de reportes en Excel.
    Contiene variables de estilo compartidas.
    """

    # Encabezados
    ENCABEZADO_GENERAl = ["CONTEO", "PORCENTAJE"]

    ENCABEZADOS_ESPECIFICOS = {
        10: ["REF RECIB ", "RESP", "% RESP ", "DERIV", "% DERIV"],
        11: ["UAPS", "CIS", "SMI ", "ZPP", "TOTAL", "% TOTAL"],
    }

    # Paleta de colores
    COLOR_PRINCIPAL = "0F2F2C"
    COLOR_ZEBRA = "D7DBE7"
    COLOR_TOTAL = "DDDAC3"
    COLOR_SUBTITULO = "B1C1BF"
    COLOR_TEXTO_BLANCO = "FFFFFF"
    COLOR_NEGRO = "000000"

    # Fuentes
    FONT_TITULO = Font(name="Helvetica", bold=True, color=COLOR_NEGRO, size=12)
    FONT_TITULO_TABLA = Font(name="Helvetica", bold=True, color=COLOR_TEXTO_BLANCO, size=12)
    FONT_GENERAL = Font(name="Helvetica", size=11)
    FONT_SUBTITULO = Font(name="Helvetica", bold=True)
    FONT_PIE = Font(name="Helvetica", size=9, bold=True)
    FONT_TOTAL = Font(name="Helvetica", bold=True, size=12)

    # Rellenos
    FILL_TITULO_TABLA = PatternFill(fill_type="solid", fgColor=COLOR_PRINCIPAL)
    FILL_ZEBRA_1 = PatternFill(fill_type="solid", fgColor="FFFFFF")
    FILL_ZEBRA_2 = PatternFill(fill_type="solid", fgColor=COLOR_ZEBRA)
    FILL_TOTAL = PatternFill(fill_type="solid", fgColor=COLOR_TOTAL)
    FILL_SUBTITULO = PatternFill(fill_type="solid", fgColor=COLOR_SUBTITULO)

    # Alineaciones
    ALIGN_TITULO = Alignment(horizontal="center", vertical="center")
    ALIGN_GENERAL = Alignment(horizontal="center", vertical="center")
    ALIGN_PRIMERA_COL = Alignment(horizontal="left", vertical="center")
    ALIGN_TOTAL = Alignment(horizontal="center", vertical="center")
    ALIGN_SUBTITULO = Alignment(horizontal="center", vertical="center")

    # Bordes
    BORDER_TITULO = Border(bottom=Side(style="thick", color=COLOR_SUBTITULO))

    BORDER_TITULO_TABLA = Border(
        left=Side(style="thin", color=COLOR_TEXTO_BLANCO),
        right=Side(style="thin", color=COLOR_TEXTO_BLANCO),
    )

    BORDER_GENERAL = Border(
        left=Side(style="thin", color=COLOR_NEGRO),
        right=Side(style="thin", color=COLOR_NEGRO),
    )

    BORDER_INFERIOR_TOTAL = Border(
        bottom=Side(style="thick", color=COLOR_PRINCIPAL),
        left=Side(style="thin", color=COLOR_NEGRO),
        right=Side(style="thin", color=COLOR_NEGRO),
    )

    BORDER_SUBTITULO_ARRIBA = Border(
        top=Side(style="thin", color=COLOR_PRINCIPAL),
        left=Side(style="thin", color=COLOR_NEGRO),
        right=Side(style="thin", color=COLOR_NEGRO),
    )

    # Alturas de fila
    ROW_HEIGHT_GENERAL = 20
    ROW_HEIGHT_ENCABEZADO = 30
    ROW_HEIGHT_TOTAL = 25
    ROW_HEIGHT_SUBTITULO = 22

    # Logo paths
    LOGO_SESAL_PATH = "core/static/core/img/logo_sesal.jpg"
    LOGO_FUNDAGES_PATH = "core/static/core/img/logo_FUNDAGES_ESCUDO.jpg"

    # ---------------------
    # Helpers
    # ---------------------

    @staticmethod
    def aplicar_estilo_rango(ws, rango, font=None, fill=None, align=None, border=None):
        for fila in ws[rango]:
            for celda in fila:
                if font: celda.font = font
                if fill: celda.fill = fill
                if align: celda.alignment = align
                if border: celda.border = border

    @staticmethod
    def convertir_porcentaje(valor):
        if isinstance(valor, str) and "%" in valor:
            limpio = valor.strip().replace(" ", "").replace("%", "")
            try:
                return float(limpio) / 100
            except:
                return valor
        return valor

    # ---------------------
    # Renderizado
    # ---------------------

    def dibujar_encabezado_excel(ws, col_inicio="A", col_fin="E"):
        logo1 = PILImage.open(ServiceExcel.LOGO_SESAL_PATH).resize((100, 80))
        logo2 = PILImage.open(ServiceExcel.LOGO_FUNDAGES_PATH).resize((180, 80))

        logo1_path = "temp_logo1.png"
        logo2_path = "temp_logo2.png"

        logo1.save(logo1_path)
        logo2.save(logo2_path)

        img1 = Image(logo1_path)
        img2 = Image(logo2_path)

        lineas = [
            "FUNDACIÓN GESTORA DE LA SALUD",
            "HOSPITAL DR. ENRIQUE AGUILAR CERRATO",
            "INTIBUCÁ, INTIBUCÁ, HONDURAS, C.A.",
            "(504) 2783-0242 / 2783-1939",
            "fundagesheac@gmail.com",
        ]

        for fila, texto in enumerate(lineas, start=1):
            ws.merge_cells(f"{col_inicio}{fila}:{col_fin}{fila}")
            celda = ws[f"{col_inicio}{fila}"]
            celda.value = texto
            celda.alignment = ServiceExcel.ALIGN_TITULO
            celda.font = Font(name="Helvetica", size=9, bold=(fila == 1))
            ws.row_dimensions[fila].height = 13

        ws.add_image(img1, f"{col_inicio}1")
        col_num_img = column_index_from_string(col_fin) - 1
        ws.add_image(img2, f"{get_column_letter(col_num_img)}1")

    def dibujar_titulo_excel(ws, titulo, mes, anio, fila, col_inicio="A", col_fin="E"):
        fila += 2

        col_fin_num = column_index_from_string(col_fin)

        for col in range(1, col_fin_num):
            ws.cell(row=fila, column=col).border = ServiceExcel.BORDER_TITULO

        ws.merge_cells(f"{col_inicio}{fila}:{col_fin}{fila}")
        c = ws[f"{col_inicio}{fila}"]
        c.value = titulo
        c.alignment = ServiceExcel.ALIGN_TITULO
        c.font = ServiceExcel.FONT_TITULO
        ws.row_dimensions[fila].height = 22

        fila += 1
        ws.merge_cells(f"{col_inicio}{fila}:{col_fin}{fila}")
        c = ws[f"{col_inicio}{fila}"]
        c.value = f"MES: {mes} { " "*10}  AÑO: {anio}"
        c.alignment = ServiceExcel.ALIGN_TITULO
        c.font = ServiceExcel.FONT_SUBTITULO
        ws.row_dimensions[fila].height = 20

    def dibujar_pie_excel(ws, fechaActual, usuario, usuario_nombre):
        fechaActual = fechaActual[:40].upper()
        usuario_txt = f"{usuario} ({usuario_nombre})"[:40]

        linea1 = '&"Helvetica"&9 ' + f"IMPRESO EL → {fechaActual}      ·      POR → {usuario_txt}"
        linea2 = '&"Helvetica,Italic"&8 SIWIH - Sistema Informatico Web Integral Hospitalario'
        linea3 = '&"Helvetica,Bold"&8 Página &P de &N'

        ws.oddFooter.center.text = f"{linea1}\n{linea2}\n{linea3}"


    def escribir_observacion(ws, texto, fila_actual, col_fin, col_ini="A", alto=35):
        # avanzar 2 filas
        fila_actual += 3

        # merge A..col_fin
        ws.merge_cells(f"{col_ini}{fila_actual}:{col_fin}{fila_actual}")
        ws.row_dimensions[fila_actual].height = alto

        celda = ws[f"{col_ini}{fila_actual}"]
        celda.value = f"Observaciones: {texto}"
        celda.alignment = Alignment(
            wrap_text=True,
            horizontal="justify",
            vertical="top"
        )
        celda.font = Font(name="Helvetica", size=9)

        return fila_actual

    # ---------------------
    # Configuraciones
    # ---------------------

    def configurar_anchos(ws, columnas):
        PERFILES = {
            3: [50, 18, 18, 2],
            6: [40, 15, 15, 15, 15, 15, 1],
            7: [40, 13, 13, 13, 13, 13, 16, 1],
        }

        if columnas not in PERFILES:
            raise ValueError(f"Perfil no definido: {columnas}")

        perfil = PERFILES[columnas]
        anchos = perfil[:-1]
        col_inicio_num = perfil[-1]

        for i, ancho in enumerate(anchos):
            col = get_column_letter(col_inicio_num + i)
            ws.column_dimensions[col].width = ancho

        col_fin_num = columnas + 2 if columnas <= 3 else columnas
        return get_column_letter(col_inicio_num), get_column_letter(col_fin_num)

    def configurar_impresion(ws, fila_encabezado):
        ws.page_setup.paperSize = 1
        ws.page_margins.top = 0.5
        ws.page_margins.header = 0.5
        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.footer = 0.4
        ws.page_margins.bottom = 1.1
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToPage = True
        ws.page_setup.scale = None
        ws.freeze_panes = f"A{fila_encabezado + 1}"
        ws.print_title_rows = f"1:{fila_encabezado}"

    def procesar_header(fila, fila_actual, indices):
        if fila[0] != "__HEADER__":
            return fila
        titulo, numero = fila[1].split("-", 1)
        nueva = [titulo, int(numero)] + fila[2:]
        indices.append(fila_actual)
        return nueva

    # ---------------------
    # Generación Excel
    # ---------------------

    @staticmethod
    def GenerarExcelReferenciaBase(reporte):
        anio = reporte["anio"]
        mes = mes_nombre(reporte["mes"], upper=True)
        titulo = reporte["informe_titulo"]
        informe = reporte["informe"]
        data = reporte["data"]
        etiqueta = reporte["etiqueta"]
        fechaActual = formatear_fecha(timezone.now())
        usuario = reporte["usuario"]
        usuario_nombre = reporte["usuario_nombre"]
        observacion = reporte.get('observacion', '0')

        try:
            if not data:
                log_warning(
                    f"Excel referencia sin datos informe {informe}",
                    app=LogApp.REPORTE
                )
                raise ValueError("No hay datos para generar el reporte")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte"

            columnas = len(data[0])
            col_ini, col_fin = ServiceExcel.configurar_anchos(ws, columnas)

            ServiceExcel.dibujar_encabezado_excel(ws, col_fin=col_fin)
            ServiceExcel.dibujar_titulo_excel(ws, titulo, mes, anio, fila=ws.max_row, col_fin=col_fin)
            ws.append([])

            encabezado = (
                [None, etiqueta.upper()] + ServiceExcel.ENCABEZADO_GENERAl
                if informe in range(1, 10)
                else [etiqueta.upper()] + ServiceExcel.ENCABEZADOS_ESPECIFICOS.get(informe, ServiceExcel.ENCABEZADO_GENERAl)
            )
            ws.append(encabezado)

            ServiceExcel.aplicar_estilo_rango(
                ws,
                f"{col_ini}{ws.max_row}:{('D' if columnas <= 3 else col_fin)}{ws.max_row}",
                font=ServiceExcel.FONT_TITULO_TABLA,
                fill=ServiceExcel.FILL_TITULO_TABLA,
                align=ServiceExcel.ALIGN_TITULO,
                border=ServiceExcel.BORDER_GENERAL,
            )
            ws.row_dimensions[ws.max_row].height = ServiceExcel.ROW_HEIGHT_ENCABEZADO

            fila_enc = ws.max_row
            ServiceExcel.configurar_impresion(ws, fila_enc)

            fila1 = data[0]
            indices_pct = [
                i for i, valor in enumerate(fila1) if isinstance(valor, str) and "%" in valor
            ]

            fila_actual = ws.max_row + 1
            indices_header = []

            for fila in data:
                fila = ServiceExcel.procesar_header(list(fila), fila_actual, indices_header)

                for i in indices_pct:
                    fila[i] = ServiceExcel.convertir_porcentaje(fila[i])

                fila_excel = [None] + fila + [None] if columnas <= 3 else fila
                ws.append(fila_excel)
                ws.row_dimensions[fila_actual].height = ServiceExcel.ROW_HEIGHT_GENERAL

                col_fin_fila = (
                    get_column_letter(ws.max_column - 1) if columnas <= 3 else col_fin
                )

                if fila_actual in indices_header:
                    ServiceExcel.aplicar_estilo_rango(
                        ws,
                        f"{col_ini}{fila_actual}:{col_fin_fila}{fila_actual}",
                        fill=ServiceExcel.FILL_SUBTITULO,
                        border=ServiceExcel.BORDER_SUBTITULO_ARRIBA,
                        font=ServiceExcel.FONT_SUBTITULO,
                        align=ServiceExcel.ALIGN_GENERAL,
                    )
                else:
                    fill = (
                        ServiceExcel.FILL_ZEBRA_1 if fila_actual % 2 == 0 else ServiceExcel.FILL_ZEBRA_2
                    )
                    ServiceExcel.aplicar_estilo_rango(
                        ws,
                        f"{col_ini}{fila_actual}:{col_fin_fila}{fila_actual}",
                        fill=fill,
                        border=ServiceExcel.BORDER_GENERAL,
                        font=ServiceExcel.FONT_GENERAL,
                        align=ServiceExcel.ALIGN_GENERAL,
                    )

                for i in indices_pct:
                    col_excel = i + (2 if columnas <= 3 else 1)
                    celda = ws.cell(row=fila_actual, column=col_excel)
                    celda.font = ServiceExcel.FONT_SUBTITULO
                    celda.number_format = "0.00%"

                col_indent = 2 if columnas <= 3 else 1
                celda = ws.cell(row=fila_actual, column=col_indent)
                celda.alignment = (
                    Alignment(horizontal="center", vertical="center")
                    if fila_actual in indices_header
                    else Alignment(horizontal="left", vertical="center", indent=1)
                )

                fila_actual += 1

            ultima = ws.max_row
            rango_total = f"{col_ini}{ultima}:{col_fin_fila}{ultima}"

            ws.row_dimensions[ultima].height = ServiceExcel.ROW_HEIGHT_TOTAL
            ServiceExcel.aplicar_estilo_rango(
                ws,
                rango_total,
                font=ServiceExcel.FONT_TOTAL,
                fill=ServiceExcel.FILL_TOTAL,
                align=ServiceExcel.ALIGN_TOTAL,
                border=ServiceExcel.BORDER_INFERIOR_TOTAL,
            )

            if observacion != '0':
                ultima = ServiceExcel.escribir_observacion(ws, observacion, ultima, col_fin)



            ServiceExcel.dibujar_pie_excel(ws, fechaActual, usuario, usuario_nombre)

            nombre_archivo = f"{titulo}_{mes}_{anio}.xlsx"

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'

            wb.save(response)
            return response
        
        except ValueError:
            raise

        except Exception:
            log_error(
                f"Error generando Excel referencia informe {reporte.get('informe')} mes {reporte.get('mes')} año {reporte.get('anio')}",
                app=LogApp.REPORTE
            )
            raise
        

    @staticmethod
    def GenerarExcelCatalogo(titulos, data, titulo):
        try:
            if not data:
                log_warning(
                    f"Excel catálogo sin datos: {titulo}",
                    app=LogApp.REPORTE
                )
                raise ValueError("No hay datos para generar el catálogo")
            

            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte"
            columnas = len(data[0])
            col_letra_fin  = get_column_letter(columnas)
            #titulo
            fila = ws.max_row + 1
            for col in range(1, columnas+1):
                ws.cell(row=fila, column=col).border = ServiceExcel.BORDER_TITULO
            ws.merge_cells(f"A{fila}:{col_letra_fin}{fila}")
            c = ws[f"A{fila}"]
            c.value = titulo
            c.alignment = ServiceExcel.ALIGN_TITULO
            c.font = ServiceExcel.FONT_TITULO
            ws.row_dimensions[fila].height = 22


            ws.append([])
            ws.append(titulos)
            fila_titulos = ws.max_row
            ws.auto_filter.ref = f"A{fila_titulos}:{col_letra_fin}{fila_titulos}"
            ws.freeze_panes = f"A{fila_titulos + 1}"


            ServiceExcel.aplicar_estilo_rango(ws, 
                        F"A{ws.max_row}:{col_letra_fin}{ws.max_row}",
                        font=ServiceExcel.FONT_TITULO_TABLA,
                        fill=ServiceExcel.FILL_TITULO_TABLA,
                        align=ServiceExcel.ALIGN_PRIMERA_COL,
                        border=ServiceExcel.BORDER_GENERAL,
                    ) 
            ws.row_dimensions[ws.max_row].height = ServiceExcel.ROW_HEIGHT_ENCABEZADO


            fila_actual = ws.max_row + 1
            maximos = [8] * columnas
            for fila in data:
                fila = list(fila)
                
                ws.append(fila)

                fill = (
                        ServiceExcel.FILL_ZEBRA_1 if fila_actual % 2 == 0 else ServiceExcel.FILL_ZEBRA_2)
                ServiceExcel.aplicar_estilo_rango(
                    ws,
                    f"A{fila_actual}:{col_letra_fin}{fila_actual}",
                    fill=fill,
                    border=ServiceExcel.BORDER_GENERAL,
                    font=ServiceExcel.FONT_GENERAL,
                    align=ServiceExcel.ALIGN_PRIMERA_COL,
                )
                for i in range(0,columnas):
                    valor = str(fila[i]) if fila[i] is not None else ""
                    largo = len(valor)
                    if largo > maximos[i]:
                        maximos[i] = largo
                    
                fila_actual += 1

            factor = 1.2
            padding = 1
            MAX_ANCHO = 75

            for num in range(0,columnas):
                largo = maximos[num]

                if largo > MAX_ANCHO:
                    largo = MAX_ANCHO

                ancho = largo*factor + padding

                ws.column_dimensions[get_column_letter(num+1)].width = ancho
                
            titulo = titulo.replace('/','_').replace(':','-')
            nombre_archivo = f"{titulo}.xlsx"

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'

            wb.save(response)
            return response
        except ValueError:
            raise

        except Exception:
            log_error(
                f"Error al generar el catalogo {titulo or "referencias"}",
                app=LogApp.REPORTE
            )
            raise

    @staticmethod
    def obtener_data_catalogo(SP, fecha_ini, fecha_fin):
        try:
            with connections['default'].cursor() as cursor:
                cursor.callproc(SP, [fecha_ini, fecha_fin])
                data = cursor.fetchall() or []
                columnas = [col[0] for col in cursor.description]
                return columnas, data
        except Exception as e:
            print(f"Error en catalogo: {e}")
            return [], []